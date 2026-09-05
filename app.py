from flask import (
    Flask, render_template, request, redirect,
    url_for, send_file, flash, jsonify, abort, session
)
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
import os as _os
from dotenv import load_dotenv
load_dotenv()
from pathlib import Path
from datetime import datetime, date, timedelta
from zoneinfo import ZoneInfo

_BRT = ZoneInfo("America/Sao_Paulo")
from io import BytesIO
import json
import platform
import re
import subprocess
import unicodedata
import uuid
import collections

from services.gerar_contrato import gerar_docx, gerar_termo_quitacao, gerar_notificacao_avalista, gerar_notificacao_inadimplente, nome_arquivo_saida
from services.gerar_vistoria_entrada_saida import gerar_vistoria_entrada_saida, docx_para_pdf as _docx_para_pdf_es

app = Flask(__name__)
app.secret_key = _os.environ.get("SECRET_KEY", "ativuz-secret-dev-2026")

VISTORIA_APP_URL = _os.environ.get("VISTORIA_APP_URL", "").strip()


@app.context_processor
def _inject_vistoria_app_url():
    return {"vistoria_app_url": VISTORIA_APP_URL}


@app.after_request
def _sem_cache_html(resp):
    """
    Páginas são geradas a cada requisição — servir HTML do cache faz o navegador
    exibir versões antigas depois de um deploy, o que já custou horas de
    diagnóstico. Estáticos seguem cacheáveis.
    """
    if resp.mimetype == "text/html":
        resp.headers["Cache-Control"] = "no-store, must-revalidate"
        resp.headers["Pragma"] = "no-cache"
    return resp


@app.errorhandler(Exception)
def handle_any_error(e):
    import traceback; traceback.print_exc()
    return jsonify({"error": str(e)}), 500


# ── Template filters ──────────────────────────────────────────────────────────

@app.template_filter('brl')
def _fmt_brl(v):
    if v is None:
        return '—'
    neg = v < 0
    s = f"{abs(v):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    return f"({s})" if neg else s


@app.template_filter('pct_fmt')
def _fmt_pct(v):
    if v is None:
        return '—'
    return f"{v * 100:.1f}%"


def _nh(s):
    """Normaliza string: minúsculas, sem acentos."""
    s = unicodedata.normalize("NFD", str(s or "").lower())
    return "".join(c for c in s if unicodedata.category(c) != "Mn")


# ── Autenticação ──────────────────────────────────────────────────────────────

_ROTAS_PUBLICAS = {
    "login", "static", "admin_novo_usuario",
    "api_contratos_ativos", "api_vistoria_importar", "api_contrato_dados",
    "api_inadimplencia_upload",
}

@app.before_request
def verificar_login():
    if request.endpoint in _ROTAS_PUBLICAS:
        return
    if not session.get("usuario"):
        return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if session.get("usuario"):
        return redirect(url_for("dashboard"))
    erro = None
    if request.method == "POST":
        nome  = request.form.get("nome", "").strip()
        senha = request.form.get("senha", "")
        sb = _supabase()
        if not sb:
            erro = "Serviço indisponível. Tente novamente."
        else:
            try:
                from supabase import create_client
                url = _os.environ.get("SUPABASE_URL", "")
                key = _os.environ.get("SUPABASE_KEY", "")
                email = (nome if "@" in nome else f"{nome}@ativuz.com").lower()
                auth_client = create_client(url, key)
                res = auth_client.auth.sign_in_with_password({"email": email, "password": senha})
                session["usuario"] = nome
                return redirect(url_for("dashboard"))
            except Exception as e:
                import traceback; traceback.print_exc()
                print(f"LOGIN ERROR: {e}")
                erro = "Nome ou senha incorretos."
    return render_template("login.html", erro=erro)




@app.route("/logout")
def logout():
    sb = _supabase()
    if sb:
        try:
            sb.auth.sign_out()
        except Exception:
            pass
    session.clear()
    return redirect(url_for("login"))


@app.route("/api/clientes")
def api_clientes():
    import openpyxl
    path = Path(__file__).parent / "planilhas" / "DADOS_CLIENTES_CONS.xlsx"
    if not path.exists():
        return jsonify([])
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < 2:
        return jsonify([])
    # Mapeamento dinâmico por nome de coluna (case-insensitive, ignora acentos)
    import unicodedata
    def _norm(s):
        s = unicodedata.normalize("NFD", str(s or "").lower())
        return "".join(c for c in s if unicodedata.category(c) != "Mn")
    headers = [_norm(h) for h in rows[0]]
    def _col(name):
        n = _norm(name)
        return next((i for i, h in enumerate(headers) if n in h), None)
    i_nome    = _col("cliente")
    i_tel     = _col("telefone")
    i_ano     = _col("ano")
    i_chassi  = _col("chassi")
    i_cor     = _col("cor")
    i_marca   = _col("marca")
    i_modelo  = _col("modelo")
    i_placa   = _col("placa")
    i_end     = _col("endereco")
    i_motor   = _col("motor")
    def _v(row, i): return str(row[i] or "") if i is not None and i < len(row) else ""
    q = request.args.get("q", "").lower().strip()
    clientes = []
    for row in rows[1:]:
        if not (i_nome is not None and i_nome < len(row) and row[i_nome]):
            continue
        nome = str(row[i_nome])
        if q and q not in nome.lower():
            continue
        marca  = _v(row, i_marca)
        modelo = _v(row, i_modelo)
        clientes.append({
            "nome":         nome,
            "telefone":     _v(row, i_tel),
            "endereco":     _v(row, i_end),
            "veiculo":      f"{marca} {modelo}".strip(),
            "placa":        _v(row, i_placa),
            "cor":          _v(row, i_cor),
            "ano":          str(int(row[i_ano])) if i_ano is not None and i_ano < len(row) and row[i_ano] else "",
            "chassi":       _v(row, i_chassi),
            "numero_motor": _v(row, i_motor),
        })
    return jsonify(clientes[:30])


@app.route("/api/todos-telefones")
def api_todos_telefones():
    import openpyxl, unicodedata
    path = Path(__file__).parent / "planilhas" / "DADOS_CLIENTES_CONS.xlsx"
    if not path.exists():
        return jsonify([])
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < 2:
        return jsonify([])
    def _norm(s):
        s = unicodedata.normalize("NFD", str(s or "").lower())
        return "".join(c for c in s if unicodedata.category(c) != "Mn")
    headers = [_norm(h) for h in rows[0]]
    def _col(name):
        n = _norm(name)
        return next((i for i, h in enumerate(headers) if n in h), None)
    i_nome = _col("cliente")
    i_tel  = _col("telefone")
    def _v(row, i): return str(row[i] or "").strip() if i is not None and i < len(row) else ""
    resultado = []
    vistos = set()
    for row in rows[1:]:
        if i_nome is None or i_nome >= len(row) or not row[i_nome]:
            continue
        nome = str(row[i_nome]).strip()
        fone = _v(row, i_tel)
        if not fone or not nome:
            continue
        if "segcomp" in _norm(nome):
            continue
        chave = (nome.lower(), fone)
        if chave in vistos:
            continue
        vistos.add(chave)
        resultado.append({"nome": nome, "telefone": fone})
    return jsonify(resultado)


def _asaas_norm(s):
    import unicodedata
    s = unicodedata.normalize("NFD", str(s or "").lower())
    return "".join(c for c in s if unicodedata.category(c) != "Mn")


# Nomes normalizados dos clientes conhecidos (motoristas)
# ─────────────────────────────────────────────────────────────────────────────
# Frotas sob administração. Cada uma tem seus próprios extratos ASAAS, veículos
# e parâmetros de contrato. A aba "Investidores" mostra a luz-divina (BYD) e a
# "Investidores 2" a joao-paulo (Polos).
# ─────────────────────────────────────────────────────────────────────────────
_FROTA_PADRAO = "luz-divina"

_FROTAS = {
    "luz-divina": {
        "nome":    "Luz Divina",
        "caucao":  3000.0,   # adesão = caução + 1ª semana
        "semana":  1200.0,
        "cota_investidor": 0.85,
        "ocupacao": {
            "TSW-3H63": [("JACKSON CASSIANO VERISSIMO",           "2026-04-27", None)],
            "TSW-3H91": [("TANIELLE GLAUCIANA SOUZA DA SILVA",    "2026-05-11", "2026-08-03"),
                         ("JOSE PEREIRA JUNIOR",                  "2026-08-31", None)],
            "TSW-3I33": [("MARCIANO EZEQUIEL VALDEVINO DA SILVA", "2026-05-11", "2026-08-03")],
            "TSW-3I03": [("ADRIANO TEOTONIO DA SILVA",            "2026-05-04", "2026-06-01"),
                         ("MARCIANO EZEQUIEL VALDEVINO DA SILVA", "2026-08-10", None)],
        },
        "indisponivel": {
            "TSW-3I03": [("Manutenção / carro reserva", "2026-06-08", "2026-08-03")],
            "TSW-3I33": [("Manutenção",                 "2026-08-10", None)],
        },
    },
    "joao-paulo": {
        "nome":    "João Paulo",
        "caucao":  1200.0,
        "semana":  750.0,
        "cota_investidor": 0.85,
        "ocupacao": {
            # STX passou por três motoristas; datas = 1º e último recebimento de cada
            "STX-6G05": [("REGINALDO BENTO DA SILVA",          "2026-04-06", "2026-04-27"),
                         ("PEDRO GABRIEL FASANARO DE OLIVEIRA", "2026-05-11", "2026-06-15"),
                         ("ANNY KATARINA ACIOLE DA SILVA",      "2026-07-13", None)],
            # SSW sempre com o Elionilson (caução paga em 18/03)
            "SSW-1A28": [("ELIONILSON CORDEIRO BARBOSA",        "2026-03-16", None)],
        },
        "indisponivel": {},
    },
}


def _frota_slug(valor=None):
    """Slug válido da frota — cai na padrão quando não reconhecido."""
    slug = (valor or "").strip()
    return slug if slug in _FROTAS else _FROTA_PADRAO


def _frota_cfg(valor=None):
    return _FROTAS[_frota_slug(valor)]


# Cobranças que passam pela mesma conta ASAAS mas não são locação de veículo —
# ficam fora de todos os totais e da tabela por motorista.
# Mesmo cliente cadastrado com nomes diferentes no ASAAS
_ASAAS_ALIAS_MOTORISTA = {
    "67.009.261 elionilson c. barbosa": "ELIONILSON CORDEIRO BARBOSA",
}


_ASAAS_NAO_VEICULO = (
    "gelo e gela conveniencia",
    "juan e ivan conveniencia",
)


_ASAAS_CLIENTES_N = [
    "jackson cassiano verissimo",
    "adriano teotonio da silva",
    "tanielle glauciana souza da silva",
    "marciano ezequiel valdevino da silva",
]

# Mapeamento fatura -> motorista real (caução paga por terceiro)
_ASAAS_PROXY_FATURA = {
    "799563477": "JACKSON CASSIANO VERISSIMO",              # Joel pagou por Jackson
    "803445386": "ADRIANO TEOTONIO DA SILVA",               # Andrier pagou por Adriano
    "811925256": "MARCIANO EZEQUIEL VALDEVINO DA SILVA",    # Andrier pagou por Marciano
    "875101330": "JACKSON CASSIANO VERISSIMO",              # Polliana pagou por Jackson
    "894432964": "JOSE PEREIRA JUNIOR",                     # Andrier pagou R$450 no cartão; com os R$750 do Pix (fatura 894414083) fecha a semana de R$1.200
}

# Pagadores que sempre representam outro motorista (todas as faturas em nome deles)
_ASAAS_PROXY_PAGADOR = {
    "polliana maria gonzalez canejo": "JACKSON CASSIANO VERISSIMO",
    "zippi solucoes de credito":      "JOSE PEREIRA JUNIOR",
}

# Faturas que são caução pura (sem 1ª semana embutida) — motorista real
_ASAAS_FATURAS_CAUCAO = {
    "767966354": "ELIONILSON CORDEIRO BARBOSA",   # caução do Polo cobrada via Ativuz
    "891485512": "JOSE PEREIRA JUNIOR",   # R$2.000 pagos via Zippi
    "891487511": "JOSE PEREIRA JUNIOR",   # R$1.000 pagos pelo próprio
}

# PIX de devolução ao motorista — estornam o que ele já havia pago.
# Adriano devolveu o carro no mesmo dia em que pagou a semana: caução + semana de volta.
_ASAAS_DEVOLUCOES = {
    "1891815437": ("devolucao_caucao",  "ADRIANO TEOTONIO DA SILVA"),
    "1891871820": ("devolucao_aluguel", "ADRIANO TEOTONIO DA SILVA"),
}

# Cobranças recebidas e depois devolvidas ao pagador — não entram em nenhum total
_ASAAS_FATURAS_DEVOLVIDAS = {
    "865166300": "Adesão Alison Ferreira Spindola devolvida via Pix em 24/07/2026",
}


def _asaas_montar_transacao(data, tx_id, tipo, estornado, desc, valor, lancamento="", cfg=None):
    """Classifica uma linha do extrato (xlsx ou OFX) no formato usado pelo painel."""
    import re as _re

    cfg    = cfg or _frota_cfg()
    desc_n = _asaas_norm(desc)
    abs_v  = abs(valor)
    piso_adesao = cfg["caucao"] + cfg["semana"]   # separa adesão de aluguel avulso

    # Nº da fatura, quando houver ("... fatura nr. 123456 NOME")
    m_fat    = _re.search(r"fatura nr\.\s*(\d+)", desc, _re.IGNORECASE)
    fatura   = m_fat.group(1) if m_fat else ""

    if any(p in desc_n for p in _ASAAS_NAO_VEICULO):
        categoria = "nao_veiculo"
    elif tx_id in _ASAAS_DEVOLUCOES:
        categoria = _ASAAS_DEVOLUCOES[tx_id][0]
    elif fatura in _ASAAS_FATURAS_DEVOLVIDAS:
        categoria = "devolvido"
    elif estornado or desc_n.startswith("estorno"):
        categoria = "estorno"
    elif fatura in _ASAAS_FATURAS_CAUCAO:
        categoria = "caucao"
    elif "cobranca recebida" in desc_n:
        categoria = "adesao" if abs_v >= piso_adesao else "aluguel"
    elif "luz divina" in desc_n:
        categoria = "repasse_investidor"
    elif "ativuz" in desc_n:
        categoria = "taxa_ativuz"
    elif "ipva" in desc_n or "seguro" in desc_n:
        categoria = "ipva"
    elif "taxa" in desc_n or "notificacao" in desc_n:
        categoria = "taxa_asaas"
    elif valor < 0 and any(c in desc_n for c in _ASAAS_CLIENTES_N):
        categoria = "reembolso_manutencao"
    else:
        categoria = "outro"

    # Extrai nome do motorista (cobranças)
    motorista = ""
    pagador   = ""
    if categoria in ("devolucao_caucao", "devolucao_aluguel"):
        motorista = _ASAAS_DEVOLUCOES[tx_id][1]
    elif categoria in ("aluguel", "adesao", "caucao"):
        mf = _re.search(r"fatura nr\.\s*(\d+)\s+(.+)$", desc, _re.IGNORECASE)
        if mf:
            pagador   = mf.group(2).strip()
            motorista = (_ASAAS_FATURAS_CAUCAO.get(mf.group(1))
                         or _ASAAS_PROXY_FATURA.get(mf.group(1))
                         or _ASAAS_PROXY_PAGADOR.get(_asaas_norm(pagador))
                         or _ASAAS_ALIAS_MOTORISTA.get(_asaas_norm(pagador))
                         or pagador)

    # Placa do seguro
    placa_seguro = ""
    if categoria == "seguro":
        m = _re.search(r"BYD\s+([A-Z0-9\-]+)", desc, _re.IGNORECASE)
        placa_seguro = m.group(1).upper().replace(" ", "") if m else ""

    return {
        "data":          data,
        "tipo":          tipo,
        "tx_id":         tx_id,
        "descricao":     desc,
        "valor":         valor,
        "lancamento":    lancamento,
        "categoria":     categoria,
        "motorista":     motorista,
        "pagador":       pagador if pagador != motorista else "",
        "placa_seguro":  placa_seguro,
        # Relevante = tudo exceto PIX para terceiros sem relação com a operação
        # e cobranças devolvidas ao pagador
        "relevante":     categoria not in ("outro", "devolvido", "nao_veiculo"),
    }


def _asaas_totais(transacoes, cfg=None):
    """Totalizadores — apenas lançamentos relevantes."""
    def _soma(cat):
        return sum(t["valor"] for t in transacoes if t["categoria"] == cat and t["relevante"])

    # Adesão = caução (valor fixo do contrato) + 1ª semana (o que sobrar)
    cfg = cfg or _frota_cfg()
    _CAUCAO  = cfg["caucao"]
    _adesoes = [t for t in transacoes if t["categoria"] == "adesao" and t["relevante"]]
    caucao_total  = len(_adesoes) * _CAUCAO
    semana_adesao = sum(max(t["valor"] - _CAUCAO, 0) for t in _adesoes)

    # Devoluções ao motorista (valores negativos) abatem o que ele havia pago
    return {
        "total_recebido":       _soma("aluguel") + _soma("adesao") + _soma("caucao")
                                + _soma("devolucao_aluguel") + _soma("devolucao_caucao"),
        "aluguel":              _soma("aluguel") + semana_adesao + _soma("devolucao_aluguel"),
        "caucao":               caucao_total + _soma("caucao") + _soma("devolucao_caucao"),
        "devolucao_aluguel":    _soma("devolucao_aluguel"),
        "devolucao_caucao":     _soma("devolucao_caucao"),
        "taxa_ativuz":          _soma("taxa_ativuz"),
        "reembolso_manutencao": _soma("reembolso_manutencao"),
        "ipva":                 _soma("ipva"),
        "taxa_asaas":           _soma("taxa_asaas"),
        "estorno":              _soma("estorno"),
    }


# (ocupação e indisponibilidade agora vivem em _FROTAS, por frota)


def _asaas_reclassificar(transacoes):
    """Reaplica as regras de motorista/caução em lançamentos já salvos."""
    import re as _re
    for t in transacoes:
        if any(p in _asaas_norm(t.get("descricao", "")) for p in _ASAAS_NAO_VEICULO):
            t["categoria"] = "nao_veiculo"
            t["relevante"] = False
            t["motorista"] = ""
            continue
        alias = _ASAAS_ALIAS_MOTORISTA.get(_asaas_norm(t.get("motorista", "")))
        if alias:
            t["motorista"] = alias
        dev = _ASAAS_DEVOLUCOES.get(t.get("tx_id") or "")
        if dev:
            t["categoria"], t["motorista"] = dev
            t["pagador"] = ""
            t["relevante"] = True
            continue
        desc = t.get("descricao", "")
        mf = _re.search(r"fatura nr\.\s*(\d+)\s+(.+)$", desc, _re.IGNORECASE)
        if not mf or t.get("categoria") not in ("aluguel", "adesao", "caucao"):
            continue
        fatura, pagador = mf.group(1), mf.group(2).strip()
        if fatura in _ASAAS_FATURAS_CAUCAO:
            t["categoria"] = "caucao"
        motorista = (_ASAAS_FATURAS_CAUCAO.get(fatura)
                     or _ASAAS_PROXY_FATURA.get(fatura)
                     or _ASAAS_PROXY_PAGADOR.get(_asaas_norm(pagador))
                     or _ASAAS_ALIAS_MOTORISTA.get(_asaas_norm(pagador))
                     or pagador)
        t["motorista"] = motorista
        t["pagador"]   = pagador if pagador != motorista else ""
    return transacoes


def _asaas_chave(t):
    """Chave de deduplicação: ID da transação ASAAS, com fallback por conteúdo."""
    return t.get("tx_id") or "{}|{}|{}".format(t.get("data"), t.get("descricao"), t.get("valor"))


def _asaas_ler_xlsx(f, cfg=None):
    """Lê o extrato oficial do ASAAS em .xlsx. Retorna (dados, erro)."""
    import openpyxl

    wb = openpyxl.load_workbook(f, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Localiza linha de cabeçalho: busca "Data" em qualquer coluna da linha
    header_row = None
    for i, r in enumerate(rows):
        cells = [str(c or "").strip() for c in r]
        if "Data" in cells and "Descrição" in cells:
            header_row = i
            break
    if header_row is None:
        return None, "Formato de arquivo não reconhecido — certifique-se de usar o extrato oficial do ASAAS (.xlsx ou .ofx)"

    # Extrai período
    periodo = ""
    for r in rows[:header_row]:
        for cell in r:
            s = str(cell or "")
            if "periodo" in _asaas_norm(s):
                periodo = s
                break

    # Saldo inicial e final extraídos do arquivo
    saldo_inicial = None
    saldo_final   = None
    for r in rows:
        desc = str(r[4] or "").strip()
        val  = r[6]
        if desc == "Saldo Inicial" and val is not None:
            try: saldo_inicial = float(val)
            except (TypeError, ValueError): pass
        elif desc == "Saldo Final" and val is not None:
            try: saldo_final = float(val)
            except (TypeError, ValueError): pass
    if saldo_inicial is None:
        saldo_inicial = 0

    transacoes = []
    for r in rows[header_row + 1:]:
        data      = str(r[0] or "").strip()
        tx_id     = str(r[1] or "").strip()   # ID único da transação ASAAS
        tipo      = str(r[2] or "").strip()
        estorn    = str(r[3] or "").strip()
        desc      = str(r[4] or "").strip()
        valor_raw = r[5]
        lancam    = str(r[11] or "").strip()

        if not data or not desc or valor_raw is None:
            continue
        try:
            valor = float(valor_raw)
        except (TypeError, ValueError):
            continue

        transacoes.append(_asaas_montar_transacao(data, tx_id, tipo, estorn, desc, valor, lancam, cfg))

    return {
        "periodo":       periodo,
        "saldo_inicial": saldo_inicial,
        "saldo_final":   saldo_final,
        "transacoes":    transacoes,
    }, None


def _asaas_ler_ofx(f, cfg=None):
    """Lê extrato bancário no formato OFX/QFX (SGML ou XML). Retorna (dados, erro)."""
    import re as _re

    raw = f.read()
    if isinstance(raw, bytes):
        for enc in ("utf-8", "latin-1"):
            try:
                texto = raw.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        else:
            texto = raw.decode("utf-8", "replace")
    else:
        texto = raw

    if "<STMTTRN>" not in texto.upper():
        return None, "Arquivo OFX sem lançamentos (<STMTTRN>) — certifique-se de exportar o extrato completo"

    def _tag(bloco, nome):
        """Valor de uma tag OFX, tolerando SGML (sem fechamento) e XML."""
        m = _re.search(r"<%s>\s*(.*?)\s*(?:</%s>|<|$)" % (nome, nome), bloco, _re.IGNORECASE | _re.DOTALL)
        return m.group(1).strip() if m else ""

    def _data_br(v):
        """'20260605120000[-3:BRT]' -> '05/06/2026'"""
        d = _re.sub(r"\D", "", v or "")[:8]
        return "%s/%s/%s" % (d[6:8], d[4:6], d[0:4]) if len(d) == 8 else ""

    transacoes = []
    for bloco in _re.findall(r"<STMTTRN>(.*?)</STMTTRN>", texto, _re.IGNORECASE | _re.DOTALL):
        data = _data_br(_tag(bloco, "DTPOSTED"))
        try:
            valor = float(_tag(bloco, "TRNAMT").replace(",", "."))
        except ValueError:
            continue
        memo = _tag(bloco, "MEMO")
        nome = _tag(bloco, "NAME")
        desc = memo if len(memo) >= len(nome) else nome
        if not data or not desc:
            continue
        transacoes.append(_asaas_montar_transacao(
            data, _tag(bloco, "FITID"), _tag(bloco, "TRNTYPE"), "", desc, valor, "", cfg))

    if not transacoes:
        return None, "Nenhum lançamento válido encontrado no arquivo OFX"

    dt_ini = _data_br(_tag(texto, "DTSTART"))
    dt_fim = _data_br(_tag(texto, "DTEND"))
    periodo = ""
    if dt_ini and dt_fim:
        periodo = "Período a partir de %s até %s" % (dt_ini, dt_fim)

    saldo_final = None
    m_bal = _re.search(r"<LEDGERBAL>(.*?)(?:</LEDGERBAL>|$)", texto, _re.IGNORECASE | _re.DOTALL)
    if m_bal:
        try:
            saldo_final = float(_tag(m_bal.group(1), "BALAMT").replace(",", "."))
        except ValueError:
            pass

    return {
        "periodo":       periodo,
        "saldo_inicial": 0,
        "saldo_final":   saldo_final,
        "transacoes":    transacoes,
    }, None


@app.route("/api/asaas-parse", methods=["POST"])
def api_asaas_parse():
    f = request.files.get("arquivo")
    if not f:
        return jsonify({"erro": "Nenhum arquivo enviado"}), 400

    slug = _frota_slug(request.args.get("frota") or request.form.get("frota"))
    cfg  = _FROTAS[slug]
    nome = (f.filename or "").lower()
    if nome.endswith((".ofx", ".qfx")):
        dados, erro = _asaas_ler_ofx(f, cfg)
    elif nome.endswith((".xlsx", ".xlsm")):
        dados, erro = _asaas_ler_xlsx(f, cfg)
    else:
        return jsonify({"erro": "Formato não suportado — envie o extrato em .xlsx ou .ofx"}), 400

    if erro:
        return jsonify({"erro": erro}), 400

    # Deduplica dentro do próprio arquivo
    vistas, unicas = set(), []
    for t in dados["transacoes"]:
        k = _asaas_chave(t)
        if k in vistas:
            continue
        vistas.add(k)
        unicas.append(t)
    dados["transacoes"] = unicas
    dados["totais"] = _asaas_totais(unicas, cfg)
    dados["frota"]  = slug

    return jsonify(dados)


@app.route("/api/asaas-salvar", methods=["POST"])
def api_asaas_salvar():
    dados = request.get_json(force=True)
    if not dados:
        return jsonify({"erro": "Sem dados"}), 400
    slug = _frota_slug(dados.get("frota") or request.args.get("frota"))
    sb = _supabase()

    # Remove lançamentos já presentes em extratos salvos (períodos sobrepostos)
    transacoes = dados.get("transacoes", [])
    ja_salvas = set()
    try:
        antigos = _sb_retry(lambda: sb.table("asaas_extratos")
                                      .select("transacoes").eq("frota", slug).execute())
        for row in (antigos.data or []):
            for t in (row.get("transacoes") or []):
                ja_salvas.add(_asaas_chave(t))
    except Exception:
        pass

    novas, vistas = [], set()
    for t in transacoes:
        k = _asaas_chave(t)
        if k in ja_salvas or k in vistas:
            continue
        vistas.add(k)
        novas.append(t)
    duplicadas = len(transacoes) - len(novas)

    if not novas:
        return jsonify({"ok": False, "duplicadas": duplicadas,
                        "erro": "Todos os %d lançamentos deste arquivo já constam em extratos salvos." % duplicadas}), 409

    res = sb.table("asaas_extratos").insert({
        "frota":         slug,
        "periodo":       dados.get("periodo", ""),
        "saldo_inicial": dados.get("saldo_inicial"),
        "saldo_final":   dados.get("saldo_final"),
        "totais":        _asaas_totais(novas, _FROTAS[slug]),
        "transacoes":    novas,
    }).execute()
    row = (res.data or [{}])[0]
    return jsonify({"ok": True, "id": row.get("id"), "duplicadas": duplicadas})


def _sb_retry(consulta, tentativas=3):
    """
    Executa uma consulta ao Supabase com retentativa.

    O cliente httpx falha esporadicamente com ReadError ([Errno 35]) mesmo em
    respostas pequenas; sem retentativa a rota devolvia 500 e a página ficava
    vazia, sem indicar o motivo.
    """
    import time as _time
    ultimo = None
    for tentativa in range(tentativas):
        try:
            return consulta()
        except Exception as exc:
            ultimo = exc
            if tentativa < tentativas - 1:
                _time.sleep(0.3 * (tentativa + 1))
    raise ultimo


@app.route("/api/asaas-extratos")
def api_asaas_extratos():
    sb = _supabase()
    try:
        slug = _frota_slug(request.args.get("frota"))
        res = _sb_retry(lambda: sb.table("asaas_extratos").select("*")
                                  .eq("frota", slug).order("created_at").execute())
    except Exception as exc:
        return jsonify({"erro": "Falha ao ler os extratos: %s" % exc}), 503
    extratos = []
    for row in (res.data or []):
        txs = _asaas_reclassificar(row.get("transacoes", []) or [])
        extratos.append({
            "id":           row["id"],
            "periodo":      row.get("periodo", ""),
            "saldo_inicial": row.get("saldo_inicial"),
            "saldo_final":  row.get("saldo_final"),
            "totais":       _asaas_totais(txs, _FROTAS[slug]),
            "transacoes":   txs,
        })
    return jsonify(extratos)


@app.route("/api/rentabilidade-real")
def api_rentabilidade_real():
    """
    Rentabilidade por veículo com dinheiro real do ASAAS.

    Cruza os recebimentos por motorista (extratos ASAAS) com o histórico de
    ocupação da frota: cada semana de aluguel é atribuída ao veículo
    que o motorista ocupava naquela segunda-feira. Cauções ficam de fora —
    são garantia, não receita do investidor.
    """
    from datetime import timedelta
    slug = _frota_slug(request.args.get("frota"))
    cfg  = _FROTAS[slug]

    # 1) Recebimentos por (motorista, segunda-feira)
    sb = _supabase()
    transacoes, vistas = [], set()
    if sb:
        try:
            res = _sb_retry(lambda: sb.table("asaas_extratos").select("transacoes").order("created_at").execute())
            for row in (res.data or []):
                for t in _asaas_reclassificar(row.get("transacoes") or []):
                    k = _asaas_chave(t)
                    if k in vistas:
                        continue
                    vistas.add(k)
                    transacoes.append(t)
        except Exception:
            pass

    por_motorista = {}   # nome maiúsculo -> {segunda(date): valor de aluguel}
    for t in transacoes:
        if t.get("categoria") not in ("aluguel", "adesao") or not t.get("relevante", True):
            continue
        try:
            d = datetime.strptime(t["data"], "%d/%m/%Y").date()
        except (ValueError, KeyError):
            continue
        seg   = d - timedelta(days=d.weekday())
        # Na adesão só a 1ª semana é aluguel; o restante é caução
        valor = max(t["valor"] - cfg["caucao"], 0) if t["categoria"] == "adesao" else t["valor"]
        nome  = (t.get("motorista") or "").upper()
        por_motorista.setdefault(nome, {})
        por_motorista[nome][seg] = por_motorista[nome].get(seg, 0) + valor

    # Devoluções de aluguel: abatem do veículo que o motorista ocupava, mesmo que o
    # PIX de volta tenha saído depois de ele já ter entregue o carro.
    devolucoes = []
    for t in transacoes:
        if t.get("categoria") != "devolucao_aluguel":
            continue
        try:
            d = datetime.strptime(t["data"], "%d/%m/%Y").date()
        except (ValueError, KeyError):
            continue
        devolucoes.append(((t.get("motorista") or "").upper(), d, t["valor"]))

    # Cada devolução pertence a uma única placa: a última que o motorista ocupou até a data
    devolucao_por_placa = {}
    for nome_dev, data_dev, valor_dev in devolucoes:
        melhor_placa, melhor_de = None, None
        for placa, periodos in cfg["ocupacao"].items():
            for nome, de, _ate in periodos:
                if nome.upper() != nome_dev:
                    continue
                d_de = datetime.strptime(de, "%Y-%m-%d").date()
                if d_de <= data_dev and (melhor_de is None or d_de > melhor_de):
                    melhor_placa, melhor_de = placa, d_de
        if melhor_placa:
            devolucao_por_placa[melhor_placa] = devolucao_por_placa.get(melhor_placa, 0) + valor_dev

    # 2) Atribui cada semana ao veículo ocupado
    modelos = {v["placa"].upper(): v.get("modelo", "") for v in (_ler_sob_administracao()[0] or [])}
    hoje    = date.today()
    veiculos = []

    for placa, periodos in cfg["ocupacao"].items():
        semanas, recebido, ocupantes = 0, 0.0, []
        for nome, de, ate in periodos:
            d_ini = datetime.strptime(de, "%Y-%m-%d").date()
            d_fim = datetime.strptime(ate, "%Y-%m-%d").date() if ate else hoje
            d_fim = min(d_fim, hoje)
            if d_fim < d_ini:
                continue
            pagos = por_motorista.get(nome.upper(), {})
            cur = d_ini
            while cur <= d_fim:
                semanas += 1
                recebido += pagos.get(cur, 0)
                cur += timedelta(weeks=1)
            ocupantes.append({"motorista": nome, "de": de, "ate": ate})

        # Semanas fora de locação por manutenção / uso como carro reserva
        paradas, sem_paradas = [], 0
        for motivo, de, ate in cfg["indisponivel"].get(placa, []):
            d_ini = datetime.strptime(de, "%Y-%m-%d").date()
            d_fim = min(datetime.strptime(ate, "%Y-%m-%d").date() if ate else hoje, hoje)
            n = 0
            cur = d_ini
            while cur <= d_fim:
                n += 1
                cur += timedelta(weeks=1)
            sem_paradas += n
            paradas.append({"motivo": motivo, "de": de, "ate": ate, "semanas": n})

        recebido += devolucao_por_placa.get(placa, 0)

        veiculos.append({
            "semanas_paradas": sem_paradas,
            "paradas":         paradas,
            "placa":              placa,
            "modelo":             modelos.get(placa, ""),
            "semanas_ativas":     semanas,
            "recebido_total":     round(recebido, 2),
            "recebido_investidor": round(recebido * cfg["cota_investidor"], 2),
            "ocupantes":          ocupantes,
        })

    veiculos.sort(key=lambda v: v["placa"])
    return jsonify({"ok": True, "veiculos": veiculos})


@app.route("/api/asaas-extratos/<extrato_id>", methods=["DELETE"])
def api_asaas_excluir(extrato_id):
    sb   = _supabase()
    slug = _frota_slug(request.args.get("frota"))
    # Só apaga se o extrato for da frota que pediu — evita remover o de outro
    # investidor por id trocado.
    alvo = _sb_retry(lambda: sb.table("asaas_extratos").select("id,frota")
                               .eq("id", extrato_id).execute())
    linha = (alvo.data or [None])[0]
    if not linha:
        return jsonify({"ok": False, "erro": "Extrato não encontrado."}), 404
    if (linha.get("frota") or _FROTA_PADRAO) != slug:
        return jsonify({"ok": False, "erro": "Este extrato pertence a outra frota."}), 403
    sb.table("asaas_extratos").delete().eq("id", extrato_id).execute()
    return jsonify({"ok": True})


@app.route("/admin/novo-usuario", methods=["GET", "POST"])
def admin_novo_usuario():
    token_correto = _os.environ.get("ADMIN_TOKEN", "")
    token = request.args.get("token", "")
    if not token_correto or token != token_correto:
        abort(403)
    mensagem = None
    erro = None
    if request.method == "POST":
        nome  = request.form.get("nome", "").strip()
        senha = request.form.get("senha", "")
        if not nome or not senha:
            erro = "Nome e senha são obrigatórios."
        else:
            sb = _supabase()
            if not sb:
                erro = "Supabase não configurado."
            else:
                try:
                    sb.table("usuarios").insert({
                        "nome": nome,
                        "senha_hash": generate_password_hash(senha),
                        "ativo": True,
                    }).execute()
                    mensagem = f"Usuário '{nome}' criado com sucesso!"
                except Exception as exc:
                    erro = f"Erro ao criar usuário: {exc}"
    return f"""
    <!DOCTYPE html><html lang="pt-BR"><head><meta charset="UTF-8">
    <title>Novo Usuário — Admin</title>
    <style>body{{font-family:Inter,sans-serif;background:#f0f2f7;display:flex;
    align-items:center;justify-content:center;min-height:100vh;}}
    .card{{background:#fff;border-radius:14px;padding:2rem;width:360px;
    box-shadow:0 4px 20px rgba(0,0,0,.1);}}
    h1{{font-size:1.1rem;margin-bottom:1.5rem;}}
    label{{font-size:.75rem;font-weight:600;text-transform:uppercase;
    letter-spacing:.05em;color:#475569;display:block;margin-bottom:.3rem;}}
    input{{width:100%;padding:.6rem .8rem;border:1.5px solid #e2e8f0;border-radius:8px;
    font-family:inherit;font-size:.9rem;margin-bottom:1rem;outline:none;}}
    input:focus{{border-color:#4361ee;}}
    button{{width:100%;padding:.7rem;background:#4361ee;color:#fff;border:none;
    border-radius:8px;font-weight:700;font-size:.9rem;cursor:pointer;}}
    .ok{{color:#166534;background:#f0fdf4;border:1px solid #bbf7d0;
    border-radius:8px;padding:.6rem .9rem;font-size:.85rem;margin-bottom:1rem;}}
    .err{{color:#991b1b;background:#fef2f2;border:1px solid #fecaca;
    border-radius:8px;padding:.6rem .9rem;font-size:.85rem;margin-bottom:1rem;}}
    </style></head><body><div class="card">
    <h1>Criar novo usuário</h1>
    {"<div class='ok'>"+mensagem+"</div>" if mensagem else ""}
    {"<div class='err'>"+erro+"</div>" if erro else ""}
    <form method="POST" action="?token={token}">
    <label>Nome</label><input name="nome" required>
    <label>Senha</label><input type="password" name="senha" required>
    <button>Criar usuário</button></form></div></body></html>
    """

# ── Supabase (opcional — só ativa se as env vars estiverem definidas) ─────────

_sb = None

def _supabase():
    global _sb
    if _sb is None:
        url = _os.environ.get("SUPABASE_URL", "")
        key = _os.environ.get("SUPABASE_KEY", "")
        if url and key:
            from supabase import create_client
            _sb = create_client(url, key)
    return _sb

_BASE = Path(__file__).parent
# Vercel (e outros runtimes read-only) só permitem escrita em /tmp.
# No Windows (dev local) usamos paths relativos; em qualquer Unix usamos /tmp.
_TMP_ROOT = Path(".") if platform.system() == "Windows" else Path("/tmp")
UPLOAD_FOLDER    = _TMP_ROOT / "uploads"
CONTRATOS_FOLDER = _TMP_ROOT / "contratos"
TEMP_FOLDER      = _TMP_ROOT / "temp_preview"
DOCX_TEMPLATES   = _BASE / "docx_templates"

UPLOAD_FOLDER.mkdir(exist_ok=True)
CONTRATOS_FOLDER.mkdir(exist_ok=True)
TEMP_FOLDER.mkdir(exist_ok=True)
DOCX_TEMPLATES.mkdir(exist_ok=True)


# ── helpers ───────────────────────────────────────────────

def _docx_bytes_to_html(docx_bytes: bytes) -> str:
    import mammoth, io
    return mammoth.convert_to_html(io.BytesIO(docx_bytes)).value


def _converter_pdf(caminho_docx: str, caminho_pdf: str):
    """Converte .docx para PDF.
    - Windows: usa Word via docx2pdf
    - Linux/Mac com LibreOffice: usa LibreOffice headless
    - Linux/Mac sem LibreOffice (ex: Vercel): usa Google Drive API
    """
    if platform.system() == "Windows":
        import pythoncom
        from docx2pdf import convert
        pythoncom.CoInitialize()
        try:
            convert(caminho_docx, caminho_pdf)
        finally:
            pythoncom.CoUninitialize()
    else:
        import shutil as _sh, tempfile, os
        if _sh.which("libreoffice"):
            # LibreOffice disponível (Docker, Fly.io, Railway, dev local)
            docx_abs = str(Path(caminho_docx).resolve())
            pdf_abs  = str(Path(caminho_pdf).resolve())
            if not Path(docx_abs).exists():
                raise FileNotFoundError(f"DOCX não encontrado: {docx_abs!r}")
            with tempfile.TemporaryDirectory() as work_dir:
                tmp_docx = Path(work_dir) / Path(docx_abs).name
                _sh.copy2(docx_abs, tmp_docx)
                env = {**os.environ, "HOME": work_dir}
                result = subprocess.run(
                    [
                        "libreoffice",
                        "--headless", "--norestore", "--nofirststartwizard",
                        "--convert-to", "pdf",
                        "--outdir", work_dir,
                        str(tmp_docx),
                    ],
                    capture_output=True,
                    env=env,
                )
                gerado = Path(work_dir) / (tmp_docx.stem + ".pdf")
                if not gerado.exists():
                    stderr = result.stderr.decode(errors="replace") if result.stderr else ""
                    stdout = result.stdout.decode(errors="replace") if result.stdout else ""
                    raise RuntimeError(
                        f"LibreOffice (exit {result.returncode}) não gerou PDF. "
                        f"docx={str(tmp_docx)!r} stderr={stderr!r} stdout={stdout!r}"
                    )
                _sh.copy2(gerado, pdf_abs)
        else:
            # Sem LibreOffice — usa mammoth + xhtml2pdf (Vercel / ambientes serverless)
            from services.docx_to_pdf import docx_bytes_to_pdf
            pdf_bytes = docx_bytes_to_pdf(Path(caminho_docx).read_bytes())
            Path(caminho_pdf).write_bytes(pdf_bytes)


def _slugify(texto: str) -> str:
    """Maiúsculas, sem acentos, espaços → underscore, sem caracteres especiais."""
    norm = unicodedata.normalize('NFD', texto)
    norm = ''.join(c for c in norm if unicodedata.category(c) != 'Mn')
    norm = norm.upper().strip()
    norm = re.sub(r'[^A-Z0-9\s]', '', norm)
    norm = re.sub(r'\s+', '_', norm)
    return norm or "SEM_NOME"


def detectar_tipo(filename: str):
    """Retorna o tipo do template com base no nome do arquivo."""
    norm = unicodedata.normalize('NFD', filename.lower())
    norm = ''.join(c for c in norm if unicodedata.category(c) != 'Mn')
    if 'quitacao' in norm:
        return 'quitacao'
    if 'locacao' in norm:
        return 'locacao'
    if 'notificacao' in norm and 'inadimplente' in norm:
        return 'inadimplente'
    if 'notificacao' in norm:
        return 'notificacao'
    return None


def get_templates():
    sb = _supabase()
    if sb:
        try:
            items = sb.storage.from_("documentos").list("templates") or []
            data_files = sorted([f for f in items if not f["name"].endswith(".json")], key=lambda x: x["name"])
            meta_map   = {f["name"]: True for f in items if f["name"].endswith(".json")}
            result = []
            for finfo in data_files:
                fname = finfo["name"]
                stem  = Path(fname).stem
                display_name = stem
                if f"{stem}.json" in meta_map:
                    try:
                        mb = sb.storage.from_("documentos").download(f"templates/{stem}.json")
                        display_name = json.loads(bytes(mb)).get("nome", stem)
                    except Exception:
                        pass
                size_kb = round((finfo.get("metadata") or {}).get("size", 0) / 1024, 1)
                updated = finfo.get("updated_at") or finfo.get("created_at") or ""
                try:
                    dt = datetime.fromisoformat(updated.replace("Z", "+00:00"))
                    data_fmt = dt.astimezone(_BRT).strftime("%d/%m/%Y %H:%M")
                except Exception:
                    data_fmt = ""
                result.append({"filename": fname, "nome": display_name,
                                "tamanho_kb": size_kb, "data": data_fmt})
            return result
        except Exception:
            import traceback; traceback.print_exc()
    # fallback local — une uploads/ e docx_templates/ sem duplicatas
    result = []
    seen = set()
    all_files = sorted(
        list(UPLOAD_FOLDER.glob("*.docx")) + list(UPLOAD_FOLDER.glob("*.xlsx")) +
        list(DOCX_TEMPLATES.glob("*.docx")) + list(DOCX_TEMPLATES.glob("*.xlsx")),
        key=lambda f: f.name,
    )
    for f in all_files:
        if f.name in seen:
            continue
        seen.add(f.name)
        # Busca metadados primeiro em uploads/, depois em docx_templates/
        meta_path = UPLOAD_FOLDER / f"{f.stem}.json"
        if not meta_path.exists():
            meta_path = DOCX_TEMPLATES / f"{f.stem}.json"
        display_name = f.stem
        if meta_path.exists():
            display_name = json.loads(meta_path.read_text(encoding="utf-8")).get("nome", f.stem)
        result.append({
            "filename": f.name,
            "nome": display_name,
            "tamanho_kb": round(f.stat().st_size / 1024, 1),
            "data": datetime.fromtimestamp(f.stat().st_mtime).strftime("%d/%m/%Y %H:%M"),
        })
    return result


def _resolve_template(filename: str):
    """
    Retorna (caminho_local, nome_display, erro).
    Se Supabase disponível, baixa para arquivo temporário.
    Caller deve apagar o temp se caminho_local estiver em TEMP_FOLDER.
    """
    safe = secure_filename(filename)
    stem = Path(safe).stem
    sb   = _supabase()
    if sb:
        try:
            data = sb.storage.from_("documentos").download(f"templates/{safe}")
            if not data:
                return None, filename, "Template não encontrado no Storage."
            TEMP_FOLDER.mkdir(exist_ok=True)
            tmp = TEMP_FOLDER / f"tpl_{uuid.uuid4().hex}{Path(safe).suffix}"
            tmp.write_bytes(bytes(data))
            nome_display = stem
            try:
                mb = sb.storage.from_("documentos").download(f"templates/{stem}.json")
                nome_display = json.loads(bytes(mb)).get("nome", stem)
            except Exception:
                pass
            return str(tmp), nome_display, None
        except Exception as e:
            import traceback; traceback.print_exc()
    # fallback local — tenta uploads/ depois docx_templates/
    local = UPLOAD_FOLDER / safe
    if not local.exists():
        local = DOCX_TEMPLATES / safe
    if not local.exists():
        return None, filename, "Template não encontrado."
    nome_display = stem
    meta_local = UPLOAD_FOLDER / f"{stem}.json"
    if meta_local.exists():
        try:
            nome_display = json.loads(meta_local.read_text(encoding="utf-8")).get("nome", stem)
        except Exception:
            pass
    return str(local), nome_display, None


def _historico_append(locatario_nome: str, template: str, arquivo: str):
    sb = _supabase()
    if not sb:
        return
    try:
        sb.table("historico_docs").insert({
            "locatario_nome": locatario_nome,
            "template": template,
            "arquivo": arquivo,
            "data_hora": datetime.now(_BRT).strftime("%d/%m/%Y %H:%M"),
        }).execute()
    except Exception:
        import traceback; traceback.print_exc()


def _gerar_para_caminho(form, tipo, template_path_str, caminho_saida):
    """Gera o documento para caminho_saida. Retorna nome_pessoa."""
    if tipo == "locacao":
        campos = [
            "locatario_nome", "locatario_rg", "locatario_cpf",
            "locatario_endereco", "locatario_cep", "locatario_telefone",
            "avalista_nome", "avalista_cpf", "avalista_endereco", "avalista_telefone",
            "veiculo_descricao", "veiculo_marca", "veiculo_modelo", "veiculo_ano",
            "veiculo_motor", "veiculo_chassi", "veiculo_cor", "veiculo_placa",
            "contrato_inicio", "contrato_duracao", "valor_semanal",
            "data_dia", "data_mes", "data_ano",
            "testemunha1_nome", "testemunha1_rg", "testemunha1_cpf",
            "testemunha2_nome", "testemunha2_rg", "testemunha2_cpf",
        ]
        dados = {c: form.get(c, "") for c in campos}
        gerar_docx(dados, caminho_saida, template_path=template_path_str)
        return dados["locatario_nome"]

    elif tipo == "notificacao":
        avalista_nome = form.get("avalista_nome_notif", "")
        gerar_notificacao_avalista(
            avalista_nome  = avalista_nome,
            data_contrato  = form.get("data_contrato", ""),
            locatario_nome = form.get("locatario_nome_notif", ""),
            valor_debito   = float(form.get("valor_debito") or 0),
            caminho_saida  = caminho_saida,
            template_path  = template_path_str,
            avalista_cpf   = form.get("avalista_cpf_notif", ""),
        )
        return avalista_nome

    elif tipo == "inadimplente":
        locatario_nome_inad = form.get("locatario_nome_inad", "")
        gerar_notificacao_inadimplente(
            locatario_nome = locatario_nome_inad,
            data_contrato  = form.get("data_contrato_inad", ""),
            valor_debito   = float(form.get("valor_debito_inad") or 0),
            caminho_saida  = caminho_saida,
            template_path  = template_path_str,
        )
        return locatario_nome_inad

    else:  # quitacao
        def _f(campo): return float(form.get(campo) or 0)
        def _i(campo): return int(float(form.get(campo) or 0))
        devedor_nome = form.get("devedor_nome", "")
        gerar_termo_quitacao(
            devedor_nome          = devedor_nome,
            devedor_cpf           = form.get("devedor_cpf", ""),
            placa                 = form.get("placa", ""),
            mes_referencia_fipe   = form.get("mes_referencia_fipe", ""),
            valor_fipe            = _f("valor_fipe"),
            percentual_fipe       = _f("percentual_fipe"),
            meias_diarias         = _f("meias_diarias"),
            entrada               = _f("entrada"),
            num_parcelas_pagas    = _i("num_parcelas_pagas"),
            valor_parcela_paga    = _f("valor_parcela_paga"),
            num_parcelas_semanais = _i("num_parcelas_semanais"),
            valor_parcela_semanal = _f("valor_parcela_semanal"),
            data_primeira_parcela = form.get("data_primeira_parcela", ""),
            data_assinatura       = form.get("data_assinatura", ""),
            caminho_saida         = caminho_saida,
            template_path         = template_path_str,
        )
        return devedor_nome


# ── página 1 — Templates ──────────────────────────────────

@app.route("/")
def dashboard():
    sb = _supabase()
    total_contratos = 0
    total_vistorias = 0
    total_docs = 0
    valor_mensal = "—"
    contratos = []
    if sb:
        try:
            res = sb.table("contratos_locacao").select(
                "id, locatario_nome, veiculo_placa, veiculo_marca, veiculo_modelo, contrato_inicio, valor_semanal",
                count="exact"
            ).order("criado_em", desc=True).limit(5).execute()
            contratos = res.data or []
            total_contratos = res.count or len(contratos)
        except Exception:
            pass
        try:
            rv = sb.table("vistorias").select("id", count="exact").execute()
            total_vistorias = rv.count or 0
        except Exception:
            pass
    if sb:
        try:
            rd = sb.table("historico_docs").select("id", count="exact").eq("deletado", False).execute()
            total_docs = rd.count or 0
        except Exception:
            pass

    # ── Frota summary ──────────────────────────────────────
    frota_total = 0
    frota_valor_fipe = None
    try:
        veiculos_frota, _, _ = _ler_frota_dados()
        frota_total = len(veiculos_frota)
        frota_valor_fipe = _frota_valor_fipe_total(veiculos_frota)
    except Exception:
        pass

    # ── Checklist: docs incompletas ─────────────────────────
    ck_pendentes_total = 0
    ck_pendentes_placas = []
    try:
        veiculos_ck, _ = _ler_veiculos()
        badge_data_ck = {}
        if sb:
            contratos_res = sb.table("checklist_contratos").select("id, contrato").execute()
            if contratos_res.data:
                ids_map = {r["id"]: r["contrato"] for r in contratos_res.data}
                itens_res = sb.table("checklist_itens").select("contrato_id, marcado").execute()
                for item in (itens_res.data or []):
                    cid  = item["contrato_id"]
                    cnum = ids_map.get(cid)
                    if cnum:
                        if cnum not in badge_data_ck:
                            badge_data_ck[cnum] = {"total": 0, "marcados": 0}
                        badge_data_ck[cnum]["total"] += 1
                        if item["marcado"]:
                            badge_data_ck[cnum]["marcados"] += 1
        for v in veiculos_ck:
            if not v.get("contrato"):
                continue
            bd = badge_data_ck.get(v["contrato"])
            if bd is None or (bd["total"] > 0 and bd["marcados"] < bd["total"]):
                ck_pendentes_placas.append(v["placa"])
        ck_pendentes_total = len(ck_pendentes_placas)
    except Exception:
        pass

    # ── Receita mensal real + contratos ativos ─────────────
    contratos_ativos = 0
    total_contratos_planilha = 0
    receita_mensal_real = None
    lista_contratos = []
    ct_app = 0
    ct_terceirizacao = 0
    receita_semanal_app = None
    _MODELOS_EXCLUIR = {"DOLPHIN", "POLO"}
    _PLACA_EXCLUIR   = "QGO-2H58"
    def _excluir_app(c):
        modelo = (c.get('modelo') or '').upper()
        placa  = (c.get('placa')  or '').upper().replace(' ', '')
        return (any(m in modelo for m in _MODELOS_EXCLUIR)
                or placa == _PLACA_EXCLUIR.upper().replace(' ', ''))
    try:
        lista_contratos, _ = _ler_contratos()
        _sync_contratos_supabase(lista_contratos)
        ativos_lst = [c for c in lista_contratos if c['situacao'] == 'EM ANDAMENTO']
        ativos_filtrados = [c for c in ativos_lst if not _excluir_app(c)]
        contratos_ativos = len(ativos_filtrados)
        total_contratos_planilha = len(lista_contratos)
        def _valor_mensal(c):
            val = c.get('valor_locacao') or 0
            tipo = (c.get('tipo_contrato') or '').upper()
            return val * 4 if 'MOTOR' in tipo else val
        s = sum(_valor_mensal(c) for c in ativos_filtrados)
        if s > 0:
            receita_mensal_real = s
        ct_app = sum(
            1 for c in ativos_filtrados
            if 'MOTOR' in (c.get('tipo_contrato') or '').upper()
        )
        ct_terceirizacao = sum(
            1 for c in ativos_filtrados
            if 'TERCEI' in (c.get('tipo_contrato') or '').upper()
        )
        contratos_ativos = ct_app + ct_terceirizacao
        _rec_sem = sum(
            c.get('valor_locacao') or 0
            for c in ativos_filtrados
            if 'MOTOR' in (c.get('tipo_contrato') or '').upper()
        )
        if _rec_sem > 0:
            receita_semanal_app = _rec_sem
    except Exception:
        pass

    # ── Saldo de financiamentos ─────────────────────────────
    saldo_financiamentos = None
    fin_ativos = 0
    try:
        from math import ceil as _ceil
        hoje_fin = datetime.now(_BRT).date()
        rows_fin = (sb.table("financiamentos_contratos").select(
            "valor_parcela,data_vencimento,parcelas_total"
        ).execute().data or []) if sb else []
        saldo = 0.0
        for r in rows_fin:
            try:
                vcto = date.fromisoformat(str(r.get("data_vencimento", ""))[:10])
                dias = (vcto - hoje_fin).days
                restante = _ceil(dias / 30.44) if dias > 0 else 0
                if restante > 0:
                    fin_ativos += 1
                    saldo += restante * float(r["valor_parcela"])
            except Exception:
                continue
        if saldo > 0:
            saldo_financiamentos = saldo
    except Exception:
        pass

    # ── Carteira judicializada ──────────────────────────────
    jud_processos = 0
    jud_valor = None
    try:
        rows_jud = (sb.table("carteira_judicializada").select(
            "status,valor_atual"
        ).execute().data or []) if sb else []
        ativos_jud = [r for r in rows_jud if (r.get("status") or "") == "Ajuizado"]
        jud_processos = len(ativos_jud)
        v = sum(float(r.get("valor_atual") or 0) for r in ativos_jud)
        if v > 0:
            jud_valor = v
    except Exception:
        pass

    inad = _inad_summary()
    contratos_vencendo = _contratos_vencendo(dias_limite=60, contratos=lista_contratos or None)
    return render_template(
        "dashboard.html",
        active="dashboard",
        total_contratos=total_contratos_planilha or total_contratos,
        contratos_ativos=contratos_ativos,
        total_vistorias=total_vistorias,
        total_docs=total_docs,
        valor_mensal=receita_mensal_real,
        contratos=contratos,
        inad=inad,
        frota_total=frota_total,
        frota_valor_fipe=frota_valor_fipe,
        ck_pendentes_total=ck_pendentes_total,
        ck_pendentes_placas=ck_pendentes_placas,
        contratos_vencendo=contratos_vencendo,
        saldo_financiamentos=saldo_financiamentos,
        fin_ativos=fin_ativos,
        jud_processos=jud_processos,
        jud_valor=jud_valor,
        ct_app=ct_app,
        ct_terceirizacao=ct_terceirizacao,
        receita_semanal_app=receita_semanal_app,
    )


@app.route("/templates")
def pagina_templates():
    return render_template("templates.html", templates=get_templates(), active="templates")


@app.route("/upload", methods=["POST"])
def upload_template():
    nome = request.form.get("nome", "").strip()
    arquivo = request.files.get("arquivo")

    if not nome:
        flash("Informe um nome para o template.", "erro")
        return redirect(url_for("pagina_templates"))

    if not arquivo or arquivo.filename == "":
        flash("Selecione um arquivo .docx.", "erro")
        return redirect(url_for("pagina_templates"))

    if not arquivo.filename.lower().endswith((".docx", ".xlsx")):
        flash("Apenas arquivos .docx ou .xlsx são aceitos.", "erro")
        return redirect(url_for("pagina_templates"))

    uid  = uuid.uuid4().hex[:8]
    ext  = Path(secure_filename(arquivo.filename)).suffix.lower()
    safe_stem    = secure_filename(f"{nome}_{uid}")
    storage_path = f"templates/{safe_stem}{ext}"
    meta_path    = f"templates/{safe_stem}.json"

    sb = _supabase()
    if sb:
        try:
            file_bytes = arquivo.read()
            ct = ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                  if ext == ".xlsx"
                  else "application/vnd.openxmlformats-officedocument.wordprocessingml.document")
            sb.storage.from_("documentos").upload(storage_path, file_bytes,
                                                   {"content-type": ct, "upsert": "true"})
            sb.storage.from_("documentos").upload(
                meta_path,
                json.dumps({"nome": nome}, ensure_ascii=False).encode("utf-8"),
                {"content-type": "application/json", "upsert": "true"},
            )
        except Exception as e:
            flash(f"Erro ao salvar template: {e}", "erro")
            return redirect(url_for("pagina_templates"))
    else:
        dest = UPLOAD_FOLDER / f"{safe_stem}{ext}"
        arquivo.save(str(dest))
        (UPLOAD_FOLDER / f"{safe_stem}.json").write_text(
            json.dumps({"nome": nome}, ensure_ascii=False), encoding="utf-8"
        )

    flash(f'Template "{nome}" enviado com sucesso!', "ok")
    return redirect(url_for("pagina_templates"))


@app.route("/templates/excluir/<filename>", methods=["POST"])
def excluir_template(filename):
    safe = secure_filename(filename)
    stem = Path(safe).stem
    sb   = _supabase()
    if sb:
        try:
            sb.storage.from_("documentos").remove([f"templates/{safe}", f"templates/{stem}.json"])
        except Exception:
            import traceback; traceback.print_exc()
    else:
        for p in (UPLOAD_FOLDER / safe, UPLOAD_FOLDER / f"{stem}.json"):
            if p.exists():
                p.unlink()
    flash("Template excluído.", "ok")
    return redirect(url_for("pagina_templates"))


# ── página 2 — Gerar Contrato ─────────────────────────────

@app.route("/gerar")
def pagina_gerar():
    return render_template("gerar.html", templates=get_templates(), active="gerar")


@app.route("/gerar-contrato", methods=["POST"])
def gerar_contrato_route():
    template_filename = request.form.get("template", "")
    if not template_filename:
        flash("Selecione um template.", "erro")
        return redirect(url_for("pagina_gerar"))

    tipo = detectar_tipo(template_filename)
    if tipo is None:
        return jsonify({
            "error": "Template não reconhecido. Renomeie o arquivo com 'locacao', 'quitacao', 'notificacao' ou 'inadimplente' no nome."
        }), 400

    tpl_path, nome_template, tpl_erro = _resolve_template(template_filename)
    if tpl_erro:
        flash(tpl_erro, "erro")
        return redirect(url_for("pagina_gerar"))

    formato = request.form.get("formato", "docx")

    # ── Nome do arquivo de saída ──────────────────────────
    if tipo == "locacao":
        ano        = datetime.now().strftime("%Y")
        nome_saida = f"{ano}_{_slugify(request.form.get('veiculo_placa', ''))}_{_slugify(request.form.get('locatario_nome', ''))}.docx"
    elif tipo == "notificacao":
        data_slug  = datetime.now().strftime("%d.%m.%Y")
        nome_saida = f"NOTIFICACAO_AVALISTA_{_slugify(request.form.get('avalista_nome_notif', ''))}_{data_slug}.docx"
    elif tipo == "inadimplente":
        data_slug  = datetime.now().strftime("%d.%m.%Y")
        nome_saida = f"NOTIFICACAO_INADIMPLENTE_{_slugify(request.form.get('locatario_nome_inad', ''))}_{data_slug}.docx"
    else:  # quitacao
        data_slug  = datetime.now().strftime("%d.%m.%Y")
        nome_saida = f"QUITACAO_DIVIDA_{_slugify(request.form.get('devedor_nome', ''))}_{data_slug}.docx"

    caminho_saida = str(CONTRATOS_FOLDER / nome_saida)

    # ── Gerar documento ───────────────────────────────────
    try:
        nome_pessoa = _gerar_para_caminho(request.form, tipo, tpl_path, caminho_saida)
    except Exception as e:
        return jsonify({"error": f"Erro ao gerar contrato: {e}"}), 500
    finally:
        # remove temp se foi baixado do Storage
        if tpl_path and tpl_path.startswith(str(TEMP_FOLDER)):
            Path(tpl_path).unlink(missing_ok=True)
    try:
        _historico_append(nome_pessoa, nome_template, nome_saida)
    except Exception:
        pass

    # ── Download direto ────────────────────────────────────
    if formato == "pdf":
        nome_pdf = nome_saida.replace(".docx", ".pdf")
        try:
            import base64 as _b64
            docx_b64    = _b64.b64encode(Path(caminho_saida).read_bytes()).decode()
            viewer_html = render_template("pdf_viewer.html",
                                          docx_b64=docx_b64,
                                          nome_arquivo=nome_pdf)
            return jsonify({"viewer_html": viewer_html})
        except Exception as e:
            docx_url = url_for("download_contrato", filename=nome_saida)
            return jsonify({
                "error": f"Erro ao gerar PDF: {e}",
                "docx_url": docx_url,
            }), 422

    return send_file(
        caminho_saida,
        as_attachment=True,
        download_name=nome_saida,
        mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    )


@app.route("/preview-contrato", methods=["POST"])
def preview_contrato():
    import mammoth

    template_filename = request.form.get("template", "")
    if not template_filename:
        return jsonify({"error": "Selecione um template."}), 400

    tipo = detectar_tipo(template_filename)
    if tipo is None:
        return jsonify({"error": "Template não reconhecido."}), 400
    if tipo == "vistoria":
        return jsonify({"error": "Pré-visualização não disponível para vistoria (formato .xlsx)."}), 400

    tpl_path, _, tpl_erro = _resolve_template(template_filename)
    if tpl_erro:
        return jsonify({"error": tpl_erro}), 400

    temp_id = uuid.uuid4().hex
    caminho_temp = str(TEMP_FOLDER / f"{temp_id}.docx")

    try:
        _gerar_para_caminho(request.form, tipo, tpl_path, caminho_temp)
    except Exception as e:
        return jsonify({"error": f"Erro ao gerar pré-visualização: {e}"}), 500
    finally:
        if tpl_path and tpl_path.startswith(str(TEMP_FOLDER)) and tpl_path != caminho_temp:
            Path(tpl_path).unlink(missing_ok=True)

    try:
        with open(caminho_temp, "rb") as f:
            result = mammoth.convert_to_html(f)
        html = result.value
    except Exception as e:
        Path(caminho_temp).unlink(missing_ok=True)
        return jsonify({"error": f"Erro ao converter para HTML: {e}"}), 500

    return jsonify({"html": html, "temp_id": temp_id})


@app.route("/cleanup-temp/<temp_id>", methods=["POST"])
def cleanup_temp(temp_id):
    if not re.match(r'^[0-9a-f]{32}$', temp_id):
        abort(400)
    caminho = TEMP_FOLDER / f"{temp_id}.docx"
    if caminho.exists():
        caminho.unlink()
    return jsonify({"ok": True})


# ── página 3 — Histórico ──────────────────────────────────

@app.route("/historico")
def pagina_historico():
    sb = _supabase()
    historico = []
    if sb:
        try:
            res = sb.table("historico_docs").select("*") \
                .eq("deletado", False) \
                .order("criado_em", desc=True).execute()
            historico = res.data or []
        except Exception:
            pass
    return render_template("historico.html", historico=historico, active="historico")


@app.route("/historico/download/<path:filename>")
def download_contrato(filename):
    caminho = (CONTRATOS_FOLDER / filename).resolve()
    if not str(caminho).startswith(str(CONTRATOS_FOLDER.resolve())):
        abort(400)
    if not caminho.exists():
        flash("Arquivo não encontrado.", "erro")
        return redirect(url_for("pagina_historico"))
    ext = Path(filename).suffix.lower()
    mime = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        if ext == ".xlsx"
        else "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    )
    return send_file(
        str(caminho),
        as_attachment=True,
        download_name=Path(filename).name,
        mimetype=mime,
    )


@app.route("/historico/download-pdf/<path:filename>")
def download_contrato_pdf(filename):
    return redirect(url_for("visualizar_contrato_historico_pdf", filename=filename))


@app.route("/historico/visualizar-pdf/<path:filename>")
def visualizar_contrato_historico_pdf(filename):
    caminho_docx = (CONTRATOS_FOLDER / filename).resolve()
    if not str(caminho_docx).startswith(str(CONTRATOS_FOLDER.resolve())):
        abort(400)
    if not caminho_docx.exists():
        abort(404)
    import base64 as _b64
    docx_b64 = _b64.b64encode(caminho_docx.read_bytes()).decode()
    nome_pdf = Path(filename).stem + ".pdf"
    return render_template("pdf_viewer.html", docx_b64=docx_b64, nome_arquivo=nome_pdf)


@app.route("/historico/contratos/<contrato_id>/visualizar")
def visualizar_contrato_pdf(contrato_id):
    sb = _supabase()
    if not sb:
        abort(503)
    try:
        res = sb.table("contratos_locacao").select("arquivo_path").eq("id", contrato_id).single().execute()
        docx_path  = res.data["arquivo_path"]
        docx_bytes = sb.storage.from_("documentos").download(docx_path)
    except Exception as e:
        return f"Erro ao buscar contrato: {e}", 500
    if not isinstance(docx_bytes, (bytes, bytearray)):
        docx_bytes = getattr(docx_bytes, "content", None) or bytes(docx_bytes)
    import base64 as _b64
    docx_b64 = _b64.b64encode(bytes(docx_bytes)).decode()
    nome_pdf = Path(docx_path).stem + ".pdf"
    return render_template("pdf_viewer.html", docx_b64=docx_b64, nome_arquivo=nome_pdf)


@app.route("/historico/vistorias/<vistoria_id>/visualizar")
def visualizar_vistoria_pdf(vistoria_id):
    sb = _supabase()
    if not sb:
        abort(503)
    try:
        res = sb.table("vistorias").select("*").eq("id", vistoria_id).single().execute()
        registro = res.data
    except Exception as e:
        return f"Erro ao buscar vistoria: {e}", 500
    try:
        docx_bytes, nome_docx = _gerar_docx_vistoria_bytes(registro, sb)
    except Exception as e:
        return f"Erro ao regenerar vistoria: {e}", 500
    import base64 as _b64
    raw = bytes(docx_bytes) if not isinstance(docx_bytes, (bytes, bytearray)) else docx_bytes
    docx_b64 = _b64.b64encode(raw).decode()
    nome_pdf = nome_docx.replace(".docx", ".pdf")
    return render_template("pdf_viewer.html", docx_b64=docx_b64, nome_arquivo=nome_pdf)


@app.route("/historico/excluir/<entry_id>", methods=["POST"])
def excluir_contrato(entry_id):
    sb = _supabase()
    if sb:
        try:
            sb.table("historico_docs").update({"deletado": True}).eq("id", entry_id).execute()
        except Exception:
            import traceback; traceback.print_exc()
    return jsonify({"ok": True})


@app.route("/historico/exportar-excel")
def exportar_historico_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    sb = _supabase()
    historico = []
    if sb:
        try:
            res = sb.table("historico_docs").select("*") \
                .eq("deletado", False) \
                .order("criado_em", desc=True).execute()
            historico = res.data or []
        except Exception:
            pass
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Histórico"

    cabecalho = ["Locatário", "Template", "Data / Hora", "Nome do Arquivo"]
    ws.append(cabecalho)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E3A5F")
        cell.alignment = Alignment(horizontal="center")

    for item in historico:
        ws.append([
            item.get("locatario_nome", ""),
            item.get("template", ""),
            item.get("data_hora", ""),
            item.get("arquivo", ""),
        ])

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    data_hoje = datetime.now().strftime("%d-%m-%Y")
    nome_arquivo = f"HISTORICO_ATIVUZ_{data_hoje}.xlsx"

    return send_file(
        buf,
        as_attachment=True,
        download_name=nome_arquivo,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ── Contrato de Locação — helpers e rotas ────────────────────────────────────

CONTRATO_LOCACAO_TEMPLATE = DOCX_TEMPLATES / "CONTRATO DE LOCAÇÃO EDITADO.docx"


def _salvar_contrato_locacao(insert: dict, caminho_docx: str, storage_path: str, edit_id: str = None):
    """INSERT no Supabase + upload do arquivo. Retorna None em sucesso ou str com erro."""
    import traceback as _tb
    sb = _supabase()
    if not sb:
        return "Supabase não configurado."
    try:
        sb.table("contratos_locacao").insert(insert).execute()
    except Exception as e:
        _tb.print_exc()
        return str(e)

    _old_path = None
    if edit_id:
        try:
            old = sb.table("contratos_locacao").select("arquivo_path").eq("id", edit_id).single().execute()
            _old_path = (old.data or {}).get("arquivo_path")
            sb.table("contratos_locacao").delete().eq("id", edit_id).execute()
        except Exception:
            _tb.print_exc()

    _docx_bytes = Path(caminho_docx).read_bytes()
    try:
        sb.storage.from_("documentos").upload(
            storage_path, _docx_bytes,
            {"content-type": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
             "upsert": "true"},
        )
    except Exception:
        _tb.print_exc()
    if _old_path and _old_path != storage_path:
        try:
            sb.storage.from_("documentos").remove([_old_path])
        except Exception:
            pass
    return None

_MESES_PT = ["janeiro","fevereiro","março","abril","maio","junho",
             "julho","agosto","setembro","outubro","novembro","dezembro"]


@app.route("/contrato-locacao")
def pagina_contrato_locacao():
    agora = datetime.now(_BRT)
    defaults = {
        "data_dia": agora.strftime("%d"),
        "data_mes": _MESES_PT[agora.month - 1],
        "data_ano": agora.strftime("%Y"),
    }
    return render_template("contrato_locacao.html", defaults=defaults, active="contrato_locacao")


@app.route("/contrato-locacao/gerar", methods=["POST"])
def gerar_contrato_locacao_route():
    """Usado pelo fluxo de edição a partir de historico_contratos."""
    if not CONTRATO_LOCACAO_TEMPLATE.exists():
        return jsonify({"error": "Template não encontrado em docx_templates/."}), 404

    campos = [
        "locatario_nome", "locatario_rg", "locatario_cpf",
        "locatario_endereco", "locatario_cep", "locatario_telefone",
        "avalista_nome", "avalista_cpf", "avalista_endereco", "avalista_telefone",
        "veiculo_descricao", "veiculo_marca", "veiculo_modelo", "veiculo_ano",
        "veiculo_motor", "veiculo_chassi", "veiculo_cor", "veiculo_placa",
        "contrato_inicio", "contrato_duracao", "valor_semanal",
        "caucao_valor", "caucao_extenso",
        "data_dia", "data_mes", "data_ano",
        "testemunha1_nome", "testemunha1_rg", "testemunha1_cpf",
        "testemunha2_nome", "testemunha2_rg", "testemunha2_cpf",
    ]
    dados   = {c: request.form.get(c, "") for c in campos}
    edit_id = request.form.get("edit_id", "").strip()

    placa_slug    = _slugify(dados.get("veiculo_placa") or "PLACA")
    nome_slug     = _slugify((dados.get("locatario_nome") or "LOCATARIO").split()[0])
    data_slug     = datetime.now(_BRT).strftime("%d.%m.%Y")
    nome_docx     = f"CONTRATO_LOCACAO_{placa_slug}_{nome_slug}_{data_slug}.docx"
    caminho_saida = str(CONTRATOS_FOLDER / nome_docx)

    try:
        gerar_docx(dados, caminho_saida, template_path=str(CONTRATO_LOCACAO_TEMPLATE))
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": f"Erro ao gerar contrato: {e}"}), 500

    _storage_path = f"contratos/{nome_docx}"
    _insert = {**dados, "arquivo_path": _storage_path}
    # Campos que não existem na tabela contratos_locacao do Supabase
    for _campo_extra in ("caucao_extenso", "caucao_valor"):
        _insert.pop(_campo_extra, None)
    _err = _salvar_contrato_locacao(_insert, caminho_saida, _storage_path, edit_id=edit_id or None)
    if _err:
        return jsonify({"error": f"Erro ao salvar no banco de dados: {_err}"}), 500

    return jsonify({"redirect_url": url_for("historico_contratos")})


@app.route("/historico/contratos")
def historico_contratos():
    sb = _supabase()
    contratos = []
    erro = None
    if sb:
        try:
            res = sb.table("contratos_locacao").select(
                "id, locatario_nome, locatario_cpf, veiculo_placa, veiculo_marca, "
                "veiculo_modelo, contrato_inicio, valor_semanal, arquivo_path, criado_em"
            ).neq("deletado", True).order("criado_em", desc=True).execute()
            contratos = res.data or []

            # contrato_inicio é texto "dd/mm/aaaa": ordena do mais recente para o
            # mais antigo em Python; sem data vai para o fim (mantém criado_em desc)
            def _ini_key(c):
                try:
                    d, m, y = str(c.get("contrato_inicio") or "").split("/")
                    return (1, date(int(y), int(m), int(d)))
                except (ValueError, TypeError):
                    return (0, date.min)
            # sort estável: empates de data mantêm a ordem criado_em desc
            contratos.sort(key=_ini_key, reverse=True)
        except Exception as e:
            erro = str(e)
    else:
        erro = "Supabase não configurado."
    return render_template("historico_contratos.html", contratos=contratos, erro=erro,
                           active="hist_contratos")


@app.route("/historico/contratos/<contrato_id>/excluir", methods=["POST"])
def excluir_contrato_locacao(contrato_id):
    sb = _supabase()
    if sb:
        try:
            sb.table("contratos_locacao").update({"deletado": True}).eq("id", contrato_id).execute()
        except Exception:
            import traceback; traceback.print_exc()
    return jsonify({"ok": True})


@app.route("/historico/contratos/download/<contrato_id>")
def download_contrato_locacao_docx(contrato_id):
    sb = _supabase()
    if not sb:
        abort(503)
    try:
        res = sb.table("contratos_locacao").select("arquivo_path").eq("id", contrato_id).single().execute()
        path = res.data["arquivo_path"]
        signed = sb.storage.from_("documentos").create_signed_url(path, 60)
        return redirect(signed["signedURL"])
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/historico/contratos/download/<contrato_id>/pdf")
def download_contrato_locacao_pdf(contrato_id):
    return redirect(url_for("visualizar_contrato_pdf", contrato_id=contrato_id))


@app.route("/historico/contratos/<contrato_id>/editar")
def editar_contrato_locacao(contrato_id):
    sb = _supabase()
    if not sb:
        flash("Supabase não configurado.", "erro")
        return redirect(url_for("historico_contratos"))
    try:
        res = sb.table("contratos_locacao").select("*").eq("id", contrato_id).single().execute()
        contrato = res.data
    except Exception as e:
        flash(f"Erro ao buscar contrato: {e}", "erro")
        return redirect(url_for("historico_contratos"))
    agora = datetime.now(_BRT)
    defaults = {
        "data_dia": contrato.get("data_dia") or agora.strftime("%d"),
        "data_mes": contrato.get("data_mes") or _MESES_PT[agora.month - 1],
        "data_ano": contrato.get("data_ano") or agora.strftime("%Y"),
    }
    return render_template("contrato_locacao.html", contrato=contrato,
                           edit_id=contrato_id, defaults=defaults,
                           active="hist_contratos")


# ── Vistoria de Entrega ───────────────────────────────────────────────────────

VISTORIA_ES_TEMPLATE = DOCX_TEMPLATES / "VISTORIA_ENTRADA_SAIDA_TEMPLATE.docx"



@app.route("/vistoria/gerar", methods=["POST"])
def gerar_vistoria_route():
    foto_path = None
    try:
        return _gerar_vistoria_impl()
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": f"Erro interno: {e}"}), 500
    finally:
        if foto_path:
            Path(foto_path).unlink(missing_ok=True)


_ANGULOS_FOTO = [
    "frontal", "traseira", "lateral_dir", "lateral_esq",
    "painel", "hodometro", "estepe", "teto",
    "motor", "mala", "dano_1", "dano_2",
]

_CHAVES_ACC = [
    "acc_calotas", "acc_buzina", "acc_doc_crlv", "acc_triangulo", "acc_antena",
    "acc_sensor_re", "acc_som", "acc_tapetes", "acc_limpadores", "acc_chave_roda",
    "acc_vidros_eletricos", "acc_oleo_motor", "acc_alarme", "acc_lampadas", "acc_macaco",
    "acc_estepe", "acc_gnv", "acc_agua", "acc_borr_psg_dir", "acc_borr_mtr_dir",
    "acc_asa_dd", "acc_asa_td", "acc_tapete_mala", "acc_tampa_parachoque",
    "acc_borr_psg_tras", "acc_borr_mtr_tras", "acc_asa_de", "acc_asa_te",
    "acc_bagagito", "acc_lingueta",
]


def _salvar_foto(file_storage):
    if not file_storage or not file_storage.filename:
        return None
    ext = Path(secure_filename(file_storage.filename)).suffix.lower()
    if ext not in ('.jpg', '.jpeg', '.png'):
        return None
    p = TEMP_FOLDER / f"{uuid.uuid4().hex}{ext}"
    file_storage.save(str(p))
    return str(p)


def _salvar_foto_base64(b64_data, mime_type=None):
    """Decodifica uma imagem base64 (payload de API externa) para um arquivo temporário local."""
    import base64 as _b64
    if not b64_data:
        return None
    ext = ".png" if (mime_type or "").lower() == "image/png" else ".jpg"
    p = TEMP_FOLDER / f"{uuid.uuid4().hex}{ext}"
    p.write_bytes(_b64.b64decode(b64_data))
    return str(p)


def _upload_bg(storage_path, docx_bytes, old_storage_path=None):
    import traceback as _tb
    try:
        sb2 = _supabase()
        if not sb2:
            return
        try:
            sb2.storage.from_("documentos").upload(
                storage_path, docx_bytes,
                {"content-type": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                 "upsert": "true"},
            )
        except Exception:
            _tb.print_exc()
        if old_storage_path and old_storage_path != storage_path:
            try:
                sb2.storage.from_("documentos").remove([old_storage_path])
            except Exception:
                pass
    except Exception:
        _tb.print_exc()


def _montar_e_salvar_vistoria(etapa, dados_fixos, campos, fotos_locais, vistoria_id_hint=""):
    """
    Núcleo de geração/gravação de uma vistoria (entrada ou saída): gera o
    .docx a partir do template, sobe fotos e documento pro Storage, grava/
    atualiza a linha em `vistorias`. Usado tanto pelo formulário interno
    (/vistoria/gerar) quanto pela importação externa (/api/vistoria/importar).

    dados_fixos = {contrato_id, cliente_nome, cliente_telefone, cliente_endereco,
                    preenchido_por, veiculo, placa, cor, ano, chassi, numero_motor}
    campos = {hodometro, combustivel, obs, sintomas, acessorios: {chave_sem_sufixo: valor}}
    fotos_locais = {angulo: caminho_temp_local}  — fotos desta etapa
    vistoria_id_hint = id explícito da vistoria a atualizar (só usado na saída;
                        se vazio, busca a mais recente pelo contrato_id)

    Retorna {"vistoria_id", "nome_docx", "status"} ou lança Exception.
    """
    import traceback as _tb
    agora = datetime.now(_BRT)
    contrato_id = dados_fixos["contrato_id"]
    placa_slug  = _slugify(dados_fixos["placa"] or "PLACA")

    if etapa == "entrada":
        dados = {
            **dados_fixos,
            "data_entrada":        agora.strftime("%d/%m/%Y %H:%M"),
            "hodometro_entrada":   campos.get("hodometro", ""),
            "combustivel_entrada": campos.get("combustivel", ""),
            "obs_entrada":         campos.get("obs", ""),
            "sintomas_entrada":    campos.get("sintomas", ""),
            "responsavel_entrada": dados_fixos["preenchido_por"],
            "acessorios_entrada":  {k: (campos.get("acessorios") or {}).get(k, "") for k in _CHAVES_ACC},
            "fotos_entrada":       fotos_locais,
        }

        data_slug    = agora.strftime("%d.%m.%Y")
        nome_docx    = f"VISTORIA_{placa_slug}_{data_slug}.docx"
        caminho_docx = str(CONTRATOS_FOLDER / nome_docx)

        pasta_fotos   = f"vistorias/fotos/{placa_slug}_{data_slug}"
        foto_s_paths  = {}
        foto_s_bytes  = {}
        for angulo, local_p in fotos_locais.items():
            ext = Path(local_p).suffix.lower() or ".jpg"
            s_path = f"{pasta_fotos}/{angulo}{ext}"
            foto_s_paths[angulo] = s_path
            try:
                foto_s_bytes[s_path] = Path(local_p).read_bytes()
            except Exception:
                pass

        try:
            resumo = gerar_vistoria_entrada_saida(
                dados,
                caminho_saida=caminho_docx,
                template_path=str(VISTORIA_ES_TEMPLATE),
            )
        except Exception as e:
            raise Exception(f"Erro ao gerar vistoria (entrada): {e}")
        finally:
            for p in fotos_locais.values():
                Path(p).unlink(missing_ok=True)

        fotos_entrada_db = [f"{ang}:{pth}" for ang, pth in foto_s_paths.items()]

        _storage_path = f"vistorias/{nome_docx}"
        sb = _supabase()
        if sb:
            try:
                sb.table("vistorias").insert({
                    "contrato_id":          contrato_id or None,
                    "cliente":              dados["cliente_nome"],
                    "telefone":             dados["cliente_telefone"],
                    "endereco":             dados["cliente_endereco"],
                    "preenchido_por":       dados["preenchido_por"],
                    "veiculo":              dados["veiculo"],
                    "placa":                dados["placa"],
                    "cor":                  dados["cor"],
                    "ano":                  dados["ano"],
                    "chassi":               dados["chassi"],
                    "numero_motor":         dados["numero_motor"],
                    "data_hora":            agora.strftime("%d/%m/%Y %H:%M"),
                    "data_entrada":         dados["data_entrada"],
                    "hodometro_entrada":    dados["hodometro_entrada"],
                    "combustivel_entrada":  dados["combustivel_entrada"],
                    "obs_entrada":          dados["obs_entrada"],
                    "sintomas_entrada":     dados["sintomas_entrada"],
                    "responsavel_entrada":  dados["responsavel_entrada"],
                    "acessorios_entrada":   dados["acessorios_entrada"],
                    "fotos_entrada":        fotos_entrada_db,
                    "status":               resumo["status"],
                    "arquivo_entrada_path": _storage_path,
                    "arquivo_path":         _storage_path,
                }).execute()
            except Exception:
                _tb.print_exc()
            _upload_bg(_storage_path, Path(caminho_docx).read_bytes())
            if foto_s_bytes:
                try:
                    sb2 = _supabase()
                    if sb2:
                        for s_path, data in foto_s_bytes.items():
                            ext2 = Path(s_path).suffix.lower()
                            ct = "image/png" if ext2 == ".png" else "image/jpeg"
                            try:
                                sb2.storage.from_("documentos").upload(
                                    s_path, data, {"content-type": ct, "upsert": "true"})
                            except Exception:
                                _tb.print_exc()
                except Exception:
                    _tb.print_exc()

        try:
            _historico_append(dados["cliente_nome"], "VISTORIA", nome_docx)
        except Exception:
            _tb.print_exc()

        return {"vistoria_id": None, "nome_docx": nome_docx, "status": resumo["status"]}

    if etapa == "saida":
        vistoria_id = vistoria_id_hint
        sb = _supabase()
        registro = {}
        caminho_docx_anterior = None

        if sb:
            try:
                if vistoria_id:
                    res = sb.table("vistorias").select("*").eq("id", vistoria_id).execute()
                else:
                    res = (sb.table("vistorias").select("*")
                             .eq("contrato_id", contrato_id)
                             .order("criado_em", desc=True)
                             .limit(1)
                             .execute())
                if res.data:
                    registro = res.data[0]
                    vistoria_id = registro.get("id", vistoria_id)
                    caminho_docx_anterior = (registro.get("arquivo_entrada_path")
                                             or registro.get("arquivo_path"))
            except Exception:
                _tb.print_exc()

        fotos_saida = fotos_locais

        fotos_entrada_recuperadas = {}
        if sb:
            for item in (registro.get("fotos_entrada") or []):
                try:
                    angulo, s_path = item.split(":", 1)
                    data_foto = sb.storage.from_("documentos").download(s_path)
                    if data_foto:
                        ext_rec = Path(s_path).suffix.lower() or ".jpg"
                        tmp_rec = TEMP_FOLDER / f"{uuid.uuid4().hex}{ext_rec}"
                        tmp_rec.write_bytes(bytes(data_foto))
                        fotos_entrada_recuperadas[angulo] = str(tmp_rec)
                except Exception:
                    _tb.print_exc()

        dados = {
            "contrato_id":      contrato_id or registro.get("contrato_id", ""),
            "cliente_nome":     registro.get("cliente", dados_fixos["cliente_nome"]),
            "cliente_telefone": registro.get("telefone", dados_fixos["cliente_telefone"]),
            "cliente_endereco": registro.get("endereco", dados_fixos["cliente_endereco"]),
            "preenchido_por":   registro.get("preenchido_por", dados_fixos["preenchido_por"]),
            "veiculo":          registro.get("veiculo", dados_fixos["veiculo"]),
            "placa":            registro.get("placa", dados_fixos["placa"]),
            "cor":              registro.get("cor", dados_fixos["cor"]),
            "ano":              registro.get("ano", dados_fixos["ano"]),
            "chassi":           registro.get("chassi", dados_fixos["chassi"]),
            "numero_motor":     registro.get("numero_motor", dados_fixos["numero_motor"]),
            "data_entrada":        registro.get("data_entrada", ""),
            "hodometro_entrada":   registro.get("hodometro_entrada", ""),
            "combustivel_entrada": registro.get("combustivel_entrada", ""),
            "obs_entrada":         registro.get("obs_entrada", ""),
            "sintomas_entrada":    registro.get("sintomas_entrada", ""),
            "responsavel_entrada": registro.get("responsavel_entrada", ""),
            "acessorios_entrada":  registro.get("acessorios_entrada") or {},
            "fotos_entrada":       fotos_entrada_recuperadas,
            "data_saida":        agora.strftime("%d/%m/%Y %H:%M"),
            "hodometro_saida":   campos.get("hodometro", ""),
            "combustivel_saida": campos.get("combustivel", ""),
            "obs_saida":         campos.get("obs", ""),
            "sintomas_saida":    campos.get("sintomas", ""),
            "responsavel_saida": dados_fixos["preenchido_por"],
            "acessorios_saida":  {k: (campos.get("acessorios") or {}).get(k, "") for k in _CHAVES_ACC},
            "fotos_saida":       fotos_saida,
        }

        placa_saida      = registro.get("placa") or dados_fixos.get("placa") or "PLACA"
        data_slug_saida  = agora.strftime("%d.%m.%Y")
        pasta_fotos_saida = f"vistorias/fotos/{_slugify(placa_saida)}_{data_slug_saida}_saida"
        foto_saida_s_paths = {}
        foto_saida_s_bytes = {}
        for angulo, local_p in fotos_saida.items():
            ext = Path(local_p).suffix.lower() or ".jpg"
            s_path = f"{pasta_fotos_saida}/{angulo}{ext}"
            foto_saida_s_paths[angulo] = s_path
            try:
                foto_saida_s_bytes[s_path] = Path(local_p).read_bytes()
            except Exception:
                pass

        if caminho_docx_anterior:
            nome_docx    = Path(caminho_docx_anterior).name
            caminho_docx = str(CONTRATOS_FOLDER / nome_docx)
        else:
            nome_docx    = f"VISTORIA_{_slugify(placa_saida)}_{data_slug_saida}.docx"
            caminho_docx = str(CONTRATOS_FOLDER / nome_docx)

        try:
            resumo = gerar_vistoria_entrada_saida(
                dados,
                caminho_saida=caminho_docx,
                template_path=str(VISTORIA_ES_TEMPLATE),
            )
        except Exception as e:
            raise Exception(f"Erro ao gerar vistoria (saida): {e}")
        finally:
            for p in fotos_saida.values():
                Path(p).unlink(missing_ok=True)
            for p in fotos_entrada_recuperadas.values():
                Path(p).unlink(missing_ok=True)

        fotos_saida_db = [f"{ang}:{pth}" for ang, pth in foto_saida_s_paths.items()]

        _storage_path = f"vistorias/{nome_docx}"
        if sb and vistoria_id:
            try:
                sb.table("vistorias").update({
                    "data_saida":            dados["data_saida"],
                    "hodometro_saida":       dados["hodometro_saida"],
                    "combustivel_saida":     dados["combustivel_saida"],
                    "obs_saida":             dados["obs_saida"],
                    "sintomas_saida":        dados["sintomas_saida"],
                    "responsavel_saida":     dados["responsavel_saida"],
                    "acessorios_saida":      dados["acessorios_saida"],
                    "fotos_saida":           fotos_saida_db,
                    "status":                resumo["status"],
                    "divergencias":          [list(d) for d in resumo["divergencias"]],
                    "arquivo_completo_path": _storage_path,
                    "arquivo_path":          _storage_path,
                }).eq("id", vistoria_id).execute()
            except Exception:
                _tb.print_exc()
            _upload_bg(_storage_path, Path(caminho_docx).read_bytes())
            if foto_saida_s_bytes:
                try:
                    sb2 = _supabase()
                    if sb2:
                        for s_path, data in foto_saida_s_bytes.items():
                            ext2 = Path(s_path).suffix.lower()
                            ct = "image/png" if ext2 == ".png" else "image/jpeg"
                            try:
                                sb2.storage.from_("documentos").upload(
                                    s_path, data, {"content-type": ct, "upsert": "true"})
                            except Exception:
                                _tb.print_exc()
                except Exception:
                    _tb.print_exc()

        try:
            _historico_append(dados["cliente_nome"], "VISTORIA", nome_docx)
        except Exception:
            _tb.print_exc()

        return {"vistoria_id": vistoria_id, "nome_docx": nome_docx, "status": resumo["status"]}

    raise Exception("Etapa de vistoria inválida.")


def _gerar_vistoria_impl():
    etapa = request.form.get("etapa", "").strip()  # "entrada" | "saida" | "" (legado)

    dados_fixos = {
        "contrato_id":      request.form.get("contrato_id", "").strip(),
        "cliente_nome":     request.form.get("cliente_nome", ""),
        "cliente_telefone": request.form.get("cliente_telefone", ""),
        "cliente_endereco": request.form.get("cliente_endereco", ""),
        "preenchido_por":   request.form.get("preenchido_por", ""),
        "veiculo":          request.form.get("veiculo", "").strip(),
        "placa":            request.form.get("placa", "").upper().strip(),
        "cor":              request.form.get("cor", ""),
        "ano":              request.form.get("ano", ""),
        "chassi":           request.form.get("chassi", ""),
        "numero_motor":     request.form.get("numero_motor", ""),
    }

    if etapa not in ("entrada", "saida"):
        return jsonify({"error": "Etapa de vistoria inválida."}), 400

    fotos_locais = {}
    for angulo in _ANGULOS_FOTO:
        p = _salvar_foto(request.files.get(f"foto_{etapa}_{angulo}"))
        if p:
            fotos_locais[angulo] = p

    campos = {
        "hodometro":   request.form.get(f"hodometro_{etapa}", ""),
        "combustivel": request.form.get(f"combustivel_{etapa}", ""),
        "obs":         request.form.get(f"obs_{etapa}", ""),
        "sintomas":    request.form.get(f"sintomas_{etapa}", ""),
        "acessorios":  {k: request.form.get(f"{k}_{etapa}", "") for k in _CHAVES_ACC},
    }

    try:
        _montar_e_salvar_vistoria(
            etapa, dados_fixos, campos, fotos_locais,
            vistoria_id_hint=request.form.get("vistoria_id", "").strip(),
        )
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500

    return jsonify({"redirect_url": url_for("historico_vistorias")})


# ── API externa de Vistoria (app Google Apps Script) ──────────────────────────
# Rotas chamadas pelo app de vistoria mobile (fora do Dashboard, sem sessão de
# login) — autenticadas por token compartilhado, não por login de usuário.

def _checar_token_vistoria():
    token_esperado = _os.environ.get("VISTORIA_API_TOKEN", "")
    token_recebido = request.headers.get("X-Vistoria-Token", "")
    return bool(token_esperado) and token_recebido == token_esperado


@app.route("/api/contratos/ativos")
def api_contratos_ativos():
    if not _checar_token_vistoria():
        return jsonify({"error": "Token inválido."}), 401

    sb = _supabase()
    if not sb:
        return jsonify({"error": "Supabase não configurado."}), 500

    try:
        res = (sb.table("contratos_frota")
                 .select("contrato_comercial, cliente, placa, modelo, situacao")
                 .eq("situacao", "EM ANDAMENTO")
                 .order("cliente")
                 .execute())
        contratos = [{
            "id":      r["contrato_comercial"],
            "cliente": r.get("cliente") or "",
            "placa":   r.get("placa") or "",
            "modelo":  r.get("modelo") or "",
        } for r in (res.data or []) if r.get("contrato_comercial")]
        return jsonify({"contratos": contratos})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/contrato/dados")
def api_contrato_dados():
    """
    Busca telefone/endereço/cor/ano/chassi/nº motor de um contrato pela
    placa, direto da planilha DADOS_CLIENTES_CONS.xlsx (mesma fonte do
    /api/clientes). Usada pelo app de vistoria mobile pra pré-preencher o
    formulário sem precisar ler/OCR nenhum documento do Drive.
    """
    if not _checar_token_vistoria():
        return jsonify({"error": "Token inválido."}), 401

    placa = request.args.get("placa", "").strip()
    if not placa:
        return jsonify({"found": False, "reason": "Placa vazia"})

    def _norm_placa(s):
        return "".join(c for c in str(s or "").upper() if c.isalnum())

    placa_norm = _norm_placa(placa)

    path = Path(__file__).parent / "planilhas" / "DADOS_CLIENTES_CONS.xlsx"
    if not path.exists():
        return jsonify({"found": False, "reason": "Planilha não encontrada"})

    import openpyxl
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < 2:
        return jsonify({"found": False, "reason": "Planilha vazia"})

    def _norm(s):
        s = unicodedata.normalize("NFD", str(s or "").lower())
        return "".join(c for c in s if unicodedata.category(c) != "Mn")
    headers = [_norm(h) for h in rows[0]]
    def _col(name):
        n = _norm(name)
        return next((i for i, h in enumerate(headers) if n in h), None)
    i_tel    = _col("telefone")
    i_ano    = _col("ano")
    i_chassi = _col("chassi")
    i_cor    = _col("cor")
    i_placa  = _col("placa")
    i_end    = _col("endereco")
    i_motor  = _col("motor")
    def _v(row, i): return str(row[i] or "") if i is not None and i < len(row) else ""

    if i_placa is None:
        return jsonify({"found": False, "reason": "Coluna Placa não encontrada na planilha"})

    for row in rows[1:]:
        if _norm_placa(_v(row, i_placa)) == placa_norm:
            return jsonify({
                "found":       True,
                "phone":       _v(row, i_tel),
                "address":     _v(row, i_end),
                "color":       _v(row, i_cor),
                "year":        str(int(row[i_ano])) if i_ano is not None and i_ano < len(row) and row[i_ano] else "",
                "chassis":     _v(row, i_chassi),
                "motorNumber": _v(row, i_motor),
            })

    return jsonify({"found": False, "reason": "Placa não encontrada na planilha"})


@app.route("/api/vistoria/importar", methods=["POST"])
def api_vistoria_importar():
    """
    Recebe uma vistoria feita pelo app mobile (Google Apps Script) e a
    processa exatamente como o formulário interno: gera o .docx, sobe pro
    Storage, grava/atualiza a linha em `vistorias`.

    Payload esperado (JSON):
    {
      etapa: "entrada" | "saida",
      contratoId, vistoriaId (opcional, só na saída),
      clientName, phone, address, filledBy,
      vehicle, plate, color, year, chassis, motorNumber,
      odometer, fuelLevel, observations, symptoms,
      accessories: [{item, status}],   // na mesma ordem de ACCESSORY_ITEMS/_CHAVES_ACC
      photos: [{category, mimeType, data}],  // category = um dos _ANGULOS_FOTO, data em base64
      clientSignature: {mimeType, data}, responsibleSignature: {mimeType, data}
    }
    """
    if not _checar_token_vistoria():
        return jsonify({"error": "Token inválido."}), 401

    payload = request.get_json(silent=True) or {}
    etapa = (payload.get("etapa") or "").strip()
    if etapa not in ("entrada", "saida"):
        return jsonify({"error": "Etapa inválida."}), 400

    dados_fixos = {
        "contrato_id":      (payload.get("contratoId") or "").strip(),
        "cliente_nome":     payload.get("clientName") or "",
        "cliente_telefone": payload.get("phone") or "",
        "cliente_endereco": payload.get("address") or "",
        "preenchido_por":   payload.get("filledBy") or "",
        "veiculo":          (payload.get("vehicle") or "").strip(),
        "placa":            (payload.get("plate") or "").upper().strip(),
        "cor":              payload.get("color") or "",
        "ano":              payload.get("year") or "",
        "chassi":           payload.get("chassis") or "",
        "numero_motor":     payload.get("motorNumber") or "",
    }

    # accessories chega como [{item, status}] na mesma ordem de _CHAVES_ACC
    accessories_list = payload.get("accessories") or []
    acessorios = {chave: (accessories_list[i].get("status", "") if i < len(accessories_list) else "")
                  for i, chave in enumerate(_CHAVES_ACC)}

    campos = {
        "hodometro":   payload.get("odometer") or "",
        "combustivel": payload.get("fuelLevel") or "",
        "obs":         payload.get("observations") or "",
        "sintomas":    payload.get("symptoms") or "",
        "acessorios":  acessorios,
    }

    fotos_locais = {}
    for foto in (payload.get("photos") or []):
        angulo = foto.get("category")
        if angulo not in _ANGULOS_FOTO:
            continue
        p = _salvar_foto_base64(foto.get("data"), foto.get("mimeType"))
        if p:
            fotos_locais[angulo] = p

    try:
        resultado = _montar_e_salvar_vistoria(
            etapa, dados_fixos, campos, fotos_locais,
            vistoria_id_hint=(payload.get("vistoriaId") or "").strip(),
        )
    except Exception as e:
        import traceback; traceback.print_exc()
        for p in fotos_locais.values():
            Path(p).unlink(missing_ok=True)
        return jsonify({"error": str(e)}), 500

    # Assinaturas: guardadas à parte no Storage (o .docx ainda não as exibe)
    sb = _supabase()
    if sb:
        import base64 as _b64
        for who, chave_payload in (("cliente", "clientSignature"), ("responsavel", "responsibleSignature")):
            sig = payload.get(chave_payload) or {}
            if sig.get("data"):
                try:
                    ext = ".png" if (sig.get("mimeType") or "").lower() == "image/png" else ".jpg"
                    s_path = f"vistorias/assinaturas/{dados_fixos['placa'] or 'PLACA'}_{etapa}_{who}{ext}"
                    sb.storage.from_("documentos").upload(
                        s_path, _b64.b64decode(sig["data"]),
                        {"content-type": "image/png" if ext == ".png" else "image/jpeg", "upsert": "true"})
                except Exception:
                    import traceback; traceback.print_exc()

    return jsonify({
        "ok": True,
        "nome_docx": resultado["nome_docx"],
        "vistoria_id": resultado["vistoria_id"],
    })


# ── Histórico de Vistorias (Supabase) ─────────────────────────────────────────

@app.route("/historico/vistorias")
def historico_vistorias():
    sb = _supabase()
    vistorias = []
    erro = None
    if sb:
        try:
            res = sb.table("vistorias").select(
                "id, cliente, placa, veiculo, preenchido_por, data_hora, criado_em, arquivo_path, status, contrato_id"
            ).neq("deletado", True).order("criado_em", desc=True).execute()
            vistorias = res.data or []
        except Exception as e:
            erro = str(e)
    else:
        erro = "Supabase não configurado (SUPABASE_URL / SUPABASE_KEY ausentes)."
    return render_template("historico_vistorias.html", vistorias=vistorias, erro=erro, active="hist_vistorias")


@app.route("/historico/vistorias/<vistoria_id>/excluir", methods=["POST"])
def excluir_vistoria(vistoria_id):
    sb = _supabase()
    if sb:
        try:
            sb.table("vistorias").update({"deletado": True}).eq("id", vistoria_id).execute()
        except Exception:
            import traceback; traceback.print_exc()
    return jsonify({"ok": True})


def _reconstruir_dados_vistoria(registro: dict, sb) -> tuple:
    """
    Monta o dict `dados` para gerar_vistoria_entrada_saida a partir de um
    registro do banco. Baixa as fotos individuais do Storage quando disponíveis.
    Retorna (dados, lista_de_caminhos_temp_para_limpar).
    """
    r = registro
    temps: list[str] = []

    def _baixar_fotos(fotos_db):
        # fotos_db pode ser ["angulo:storage_path", ...] (novo) ou None/[] (sem fotos)
        result: dict[str, str] = {}
        if not sb:
            return result
        for item in (fotos_db or []):
            try:
                angulo, s_path = item.split(":", 1)
                data = sb.storage.from_("documentos").download(s_path)
                if data:
                    ext = Path(s_path).suffix.lower() or ".jpg"
                    tmp = TEMP_FOLDER / f"{uuid.uuid4().hex}{ext}"
                    tmp.write_bytes(bytes(data))
                    result[angulo] = str(tmp)
                    temps.append(str(tmp))
            except Exception:
                import traceback as _trc; _trc.print_exc()
        return result

    dados = {
        "contrato_id":       r.get("contrato_id", ""),
        "cliente_nome":      r.get("cliente", ""),
        "cliente_telefone":  r.get("telefone", ""),
        "cliente_endereco":  r.get("endereco", ""),
        "preenchido_por":    r.get("preenchido_por", ""),
        "veiculo":           r.get("veiculo", ""),
        "placa":             r.get("placa", ""),
        "cor":               r.get("cor", ""),
        "ano":               str(r.get("ano") or ""),
        "chassi":            r.get("chassi", ""),
        "numero_motor":      r.get("numero_motor", ""),
        # entrada — novos campos com fallback para os antigos
        "data_entrada":        r.get("data_entrada") or r.get("data_hora", ""),
        "hodometro_entrada":   r.get("hodometro_entrada") or r.get("hodometro_entrega", ""),
        "combustivel_entrada": r.get("combustivel_entrada") or r.get("combustivel", ""),
        "obs_entrada":         r.get("obs_entrada") or r.get("obs_gerais", ""),
        "sintomas_entrada":    r.get("sintomas_entrada") or r.get("desc_sintomas", ""),
        "responsavel_entrada": r.get("responsavel_entrada", ""),
        "acessorios_entrada":  r.get("acessorios_entrada") or r.get("acessorios") or {},
        "fotos_entrada":       _baixar_fotos(r.get("fotos_entrada")),
        # saída
        "data_saida":          r.get("data_saida", ""),
        "hodometro_saida":     r.get("hodometro_saida") or r.get("hodometro_retorno", ""),
        "combustivel_saida":   r.get("combustivel_saida", ""),
        "obs_saida":           r.get("obs_saida", ""),
        "sintomas_saida":      r.get("sintomas_saida", ""),
        "responsavel_saida":   r.get("responsavel_saida", ""),
        "acessorios_saida":    r.get("acessorios_saida") or {},
        "fotos_saida":         _baixar_fotos(r.get("fotos_saida")),
    }
    return dados, temps


def _gerar_docx_vistoria_bytes(registro: dict, sb) -> tuple:
    """
    Reconstrói e regenera o DOCX de uma vistoria usando o template atual.
    Retorna (docx_bytes, nome_sugerido_do_arquivo).
    """
    dados, temps = _reconstruir_dados_vistoria(registro, sb)
    placa_slug = _slugify(dados.get("placa") or "PLACA")
    nome_docx  = f"VISTORIA_{placa_slug}.docx"
    tmp_docx   = TEMP_FOLDER / f"{uuid.uuid4().hex}.docx"
    TEMP_FOLDER.mkdir(exist_ok=True)
    try:
        gerar_vistoria_entrada_saida(
            dados,
            caminho_saida=str(tmp_docx),
            template_path=str(VISTORIA_ES_TEMPLATE),
        )
        return tmp_docx.read_bytes(), nome_docx
    finally:
        tmp_docx.unlink(missing_ok=True)
        for p in temps:
            Path(p).unlink(missing_ok=True)


@app.route("/historico/vistorias/download/<vistoria_id>")
def download_vistoria_supabase(vistoria_id):
    sb = _supabase()
    if not sb:
        abort(503)
    try:
        res = sb.table("vistorias").select("*").eq("id", vistoria_id).single().execute()
        registro = res.data
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    try:
        docx_bytes, nome_docx = _gerar_docx_vistoria_bytes(registro, sb)
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": f"Erro ao regenerar vistoria: {e}"}), 500
    return send_file(
        BytesIO(docx_bytes),
        as_attachment=True,
        download_name=nome_docx,
        mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    )


@app.route("/historico/vistorias/download/<vistoria_id>/pdf")
def download_vistoria_pdf(vistoria_id):
    return redirect(url_for("visualizar_vistoria_pdf", vistoria_id=vistoria_id))



@app.route("/vistoria/download/<nome>")
def baixar_vistoria(nome):
    caminho = CONTRATOS_FOLDER / nome
    if not caminho.exists() or caminho.parent.resolve() != CONTRATOS_FOLDER.resolve():
        abort(404)
    ext = caminho.suffix.lower()
    mime = "application/pdf" if ext == ".pdf" else \
           "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    return send_file(str(caminho), as_attachment=True, download_name=nome, mimetype=mime)


@app.route("/debug/libreoffice")
def debug_libreoffice():
    """Diagnóstico: testa conversão de um DOCX mínimo a PDF."""
    import tempfile, os, shutil, platform as _pf
    lines = [f"platform={_pf.system()}"]
    # versão do LO
    try:
        r = subprocess.run(["libreoffice", "--version"], capture_output=True, timeout=10)
        lines.append(f"version={r.stdout.decode(errors='replace').strip()}")
    except Exception as e:
        lines.append(f"version_error={e}")
    # criação de DOCX mínimo
    try:
        from docx import Document
        with tempfile.TemporaryDirectory() as wd:
            p_docx = Path(wd) / "test.docx"
            p_pdf  = Path(wd) / "test.pdf"
            doc = Document(); doc.add_paragraph("Teste LibreOffice"); doc.save(str(p_docx))
            _converter_pdf(str(p_docx), str(p_pdf))
            lines.append(f"pdf_ok={p_pdf.exists()} size={p_pdf.stat().st_size if p_pdf.exists() else 0}")
    except Exception as e:
        lines.append(f"convert_error={e}")
    return "<br>".join(lines)


# ── Controle de Inadimplência ─────────────────────────────────────────────────

def _brl(v):
    """Format float as Brazilian currency (R$ 1.234,56)."""
    try:
        v = float(v)
    except (TypeError, ValueError):
        v = 0.0
    parts = f"{v:,.2f}".split(".")
    return "R$ " + parts[0].replace(",", ".") + "," + parts[1]


def _parse_valor_excel(raw):
    if isinstance(raw, (int, float)):
        return float(raw)
    if not raw:
        return 0.0
    s = str(raw).replace("R$", "").replace(" ", "").strip()
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except (ValueError, TypeError):
        return 0.0


def _contratos_vencendo(dias_limite: int = 60, contratos=None):
    """Return EM ANDAMENTO contracts expiring within `dias_limite` days or already expired, sorted asc."""
    try:
        if contratos is None:
            contratos, _ = _ler_contratos()
    except Exception:
        return []
    resultado = []
    for c in contratos:
        if c.get("situacao") != "EM ANDAMENTO":
            continue
        dias_rest = c.get("dias_vencer")
        if dias_rest is None or dias_rest > dias_limite:
            continue
        resultado.append({
            "placa":          c["placa"],
            "cliente":        c["cliente"],
            "modelo":         c["modelo"],
            "termino":        c["termino_previsto"],
            "dias_restantes": dias_rest,
            "vencido":        dias_rest < 0,
        })
    resultado.sort(key=lambda x: x["dias_restantes"])
    return resultado


def _ler_inad_dados():
    """
    Fonte única: lê CONTAS-A-RECEBER.xlsx e retorna
    (registros_vencidos, registros_a_vencer, erro_leitura).
    Usada tanto por _inad_summary() quanto por pagina_inadimplencia().
    """
    from urllib.parse import quote as _url_quote
    from collections import Counter
    import openpyxl

    _base         = Path(__file__).parent / "planilhas"
    xlsx_path     = _base / "CONTAS-A-RECEBER.xlsx"
    clientes_path = _base / "DADOS_CLIENTES_CONS.xlsx"

    _tel_map = {}
    if clientes_path.exists():
        try:
            _wb_c = openpyxl.load_workbook(str(clientes_path), read_only=True, data_only=True)
            for row in _wb_c.active.iter_rows(min_row=2, values_only=True):
                nome_c = str(row[0] or "").strip()
                fone_c = str(row[11] or "").strip()
                if nome_c and fone_c:
                    digits = "".join(c for c in fone_c if c.isdigit())
                    if len(digits) >= 10:
                        _tel_map[_nh(nome_c)] = "55" + digits
            _wb_c.close()
        except Exception:
            pass

    hoje = date.today()
    registros_vencidos = []
    registros_a_vencer = []
    erro_leitura = None

    if not xlsx_path.exists():
        return registros_vencidos, registros_a_vencer, (
            "Planilha não encontrada em planilhas/. "
            "Salve o arquivo como CONTAS-A-RECEBER.xlsx nessa pasta."
        )

    try:
        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        header_idx = 0
        for ri, row in enumerate(rows[:10]):
            nh_row = [_nh(str(c or "")) for c in row]
            if sum(1 for t in ["receber de", "vencimento", "valor"]
                   if any(t in n for n in nh_row)) >= 2:
                header_idx = ri
                break

        header    = rows[header_idx]
        data_rows = rows[header_idx + 1:]

        def _ci(keyword):
            nk = _nh(keyword)
            return next((i for i, h in enumerate(header)
                         if h is not None and nk in _nh(str(h))), None)

        i_nome  = _ci("receber de (fantasia)") or _ci("receber de")
        i_valor = _ci("valor previsto") or _ci("valor")
        i_venc  = _ci("data de vencimento") or _ci("vencimento")
        i_sit   = _ci("situacao (data de vencimento)") or _ci("situacao")
        i_tipo  = _ci("tipo de fatura") or _ci("tipo")
        i_doc   = _ci("numero do documento") or _ci("documento")
        i_unid  = _ci("unidade")

        _NOMES_EXCLUIDOS = {
            "MARCELO BENTO DE ARAUJO", "M&S DO TRABALHO",
            "SEGCOMP", "SEGCOMP TECNOLOGIA LTDA", "NEW CHARGER",
        }

        # Placas de veículos de investidores: só entram manutenção,
        # lançamento avulso e taxa de administração — locação e caução ficam de fora.
        _PLACAS_INVESTIDOR = {"TWS", "QGO"}
        _TIPOS_PERMITIDOS_INV = ("manutencao", "avulso", "taxa")

        def _excluir_placa_investidor(unidade, tipo_fatura):
            unid_up = str(unidade or "").upper()
            if not any(p in unid_up for p in _PLACAS_INVESTIDOR):
                return False          # placa normal → incluir sempre
            tipo_norm = _nh(str(tipo_fatura or ""))
            return not any(t in tipo_norm for t in _TIPOS_PERMITIDOS_INV)

        name_counts = Counter()
        for row in data_rows:
            if i_nome is not None and i_nome < len(row) and row[i_nome]:
                n = str(row[i_nome]).strip()
                if n and n.upper() not in _NOMES_EXCLUIDOS:
                    name_counts[n] += 1

        def _get(row, idx):
            return row[idx] if idx is not None and idx < len(row) else None

        for row in data_rows:
            nome_raw = _get(row, i_nome)
            if not nome_raw:
                continue
            nome = str(nome_raw).strip()
            if not nome or nome.upper() in _NOMES_EXCLUIDOS:
                continue

            valor    = _parse_valor_excel(_get(row, i_valor))
            venc_raw = _get(row, i_venc)
            sit_raw  = _get(row, i_sit)
            tipo_raw = _get(row, i_tipo)
            doc_raw  = _get(row, i_doc)
            unid_raw = _get(row, i_unid)

            venc_date = None
            if venc_raw:
                if isinstance(venc_raw, datetime):
                    venc_date = venc_raw.date()
                elif isinstance(venc_raw, date):
                    venc_date = venc_raw
                else:
                    venc_str = str(venc_raw).strip().upper()
                    if venc_str == "HOJE":
                        venc_date = hoje
                    else:
                        for fmt in ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"]:
                            try:
                                venc_date = datetime.strptime(venc_str, fmt).date()
                                break
                            except (ValueError, TypeError):
                                pass
            if venc_date is None:
                continue

            situacao    = _nh(str(sit_raw  or ""))
            tipo_fatura = str(tipo_raw or "").strip()
            num_doc     = str(doc_raw  or "").strip()
            unidade     = str(unid_raw or "").strip()
            for s in ("None", "nan", ""):
                if num_doc == s: num_doc = ""
                if unidade == s: unidade = ""

            reincidente = name_counts[nome] > 1
            is_fatura   = _nh(tipo_fatura) == "fatura"
            data_fmt    = venc_date.strftime("%d/%m/%Y")

            # ── A VENCER ──────────────────────────────────────────────────────
            if "a vencer" in situacao and venc_date > hoje:
                if _excluir_placa_investidor(unidade, tipo_fatura):
                    continue
                dias_ate = (venc_date - hoje).days
                registros_a_vencer.append({
                    "nome":            nome,
                    "num_doc":         num_doc,
                    "unidade":         unidade,
                    "data_vencimento": data_fmt,
                    "dias_ate":        dias_ate,
                    "reincidente":     reincidente,
                    "tipo_fatura":     tipo_fatura,
                    "valor_s":         _brl(valor),
                    "_valor":          valor,
                })
                continue

            # ── VENCIDO ou vence hoje ─────────────────────────────────────────
            if "a vencer" in situacao and venc_date == hoje:
                dias = 0
            else:
                dias = (hoje - venc_date).days
                if dias < 0:
                    continue

            if dias == 0:    etapa, etapa_cls = "Hoje",         "stage-d0"
            elif dias == 1:  etapa, etapa_cls = "Terça-feira",  "stage-d1"
            elif dias == 2:  etapa, etapa_cls = "Quarta-feira", "stage-d2"
            elif dias == 3:  etapa, etapa_cls = "Quinta-feira", "stage-d3"
            elif dias == 4:  etapa, etapa_cls = "Sexta-feira",  "stage-d4"
            elif dias <= 6:  etapa, etapa_cls = "D+5",          "stage-d5"
            elif dias <= 9:  etapa, etapa_cls = "D+7",          "stage-d7"
            elif dias <= 14: etapa, etapa_cls = "D+10",         "stage-d10"
            else:            etapa, etapa_cls = "D+15",         "stage-d15"

            if dias == 0:    proxima = "Enviar lembrete de vencimento"
            elif dias == 1:  proxima = "Aviso de atraso — tem até o final do dia para pagar, caso contrário amanhã entram os juros"
            elif dias == 2:  proxima = "Juros aplicado — a partir de amanhã inicia a contagem dos juros de mora"
            elif dias == 3:  proxima = "Juros de mora em contagem — regularize hoje para evitar suspensão do serviço"
            elif dias == 4:  proxima = "Aviso final — regularize até hoje ou o serviço será suspenso"
            elif dias <= 6:  proxima = "Serviço suspenso — exigir comprovante de pagamento para reativação"
            elif dias <= 9:  proxima = "Encaminhar para cobrança jurídica extrajudicial"
            elif dias <= 14: proxima = "Negativação no SPC/Serasa + encaminhamento jurídico"
            else:            proxima = "Processo judicial iniciado — recolhimento imediato do veículo"

            if _excluir_placa_investidor(unidade, tipo_fatura):
                continue

            tem_fatura = bool(num_doc)
            multa      = valor * 0.10 if (tem_fatura and dias >= 2) else 0.0
            juros_mora = (valor + multa) * 0.00033 * dias if (tem_fatura and dias >= 3) else 0.0
            juros      = multa + juros_mora
            total      = valor + juros
            pausar     = total * 0.5

            dias_label = (f"{dias} dia{'s' if dias != 1 else ''} de atraso"
                          if dias > 0 else "Vence hoje")
            dias_s = f"{dias} dia{'s' if dias != 1 else ''}"

            if dias == 0:
                msg = f"Oi, {nome}! 😊 Passando para avisar que sua parcela de *{_brl(valor)}* vence *hoje*. Qualquer dúvida, é só chamar!\n\n\n*Ativuz Veículos*"
            elif dias == 1:
                msg = f"{nome}, sua parcela de *{_brl(valor)}* venceu ontem (vencimento: {data_fmt}). O valor atualizado é *{_brl(total)}*. Assim que puder, regularize para evitar encargos adicionais.\n\n\n*Ativuz Veículos*"
            elif dias == 2:
                msg = f"{nome}, seu pagamento de *{_brl(total)}* ainda está em aberto (vencimento: {data_fmt}). Caso tenha alguma dúvida ou dificuldade, entre em contato antes que os encargos aumentem.\n\n\n*Ativuz Veículos*"
            elif dias == 3:
                msg = f"{nome}, seu pagamento está em aberto há *{dias_s}* (vencimento: {data_fmt}). Valor atualizado: *{_brl(total)}*. Caso haja algum imprevisto, entre em contato — mas precisamos regularizar em breve para evitar a suspensão do serviço.\n\n\n*Ativuz Veículos*"
            elif dias == 4:
                msg = f"{nome}, sua parcela de *{_brl(valor)}* segue em aberto há *{dias_s}* (vencimento: {data_fmt}). Valor atualizado: *{_brl(total)}*. Regularize o quanto antes para evitar a suspensão do veículo.\n\n\n*Ativuz Veículos*"
            elif dias <= 6:
                msg = f"{nome}, infelizmente precisamos suspender o serviço por inadimplência, conforme contrato. Valor atualizado: *{_brl(total)}*. Para reativação, basta regularizar o pagamento. Estamos à disposição.\n\n\n*Ativuz Veículos*"
            elif dias <= 9:
                msg = f"{nome}, seu débito de *{_brl(total)}* está em aberto há {dias_s}. Esta é uma notificação formal com prazo de *48 horas* para regularização antes de tomarmos as próximas medidas previstas em contrato.\n\n\n*Ativuz Veículos*"
            elif dias <= 14:
                msg = f"{nome}, informamos que seu débito foi encaminhado para negativação e assessoria jurídica. Valor atualizado: *{_brl(total)}*.\n\n\n*Ativuz Veículos*"
            else:
                msg = f"{nome}, comunicamos que serão iniciados os procedimentos de protesto em cartório e execução contratual. Valor atualizado: *{_brl(total)}*.\n\n\n*Ativuz Veículos*"

            mostrar_pausar = is_fatura and 1 <= dias <= 2
            if mostrar_pausar:
                if dias == 1:
                    msg_pausar = f"{nome}, sua parcela de *{_brl(valor)}* está em aberto (vencimento: {data_fmt}). O valor atualizado é *{_brl(total)}*. 📌 Pague *{_brl(pausar)}* hoje e quite *{_brl(pausar)}* até a sexta-feira desta semana. ⚠️ Juros de 0,5% ao dia continuam correndo sobre o saldo restante. Sem pagamento até sexta, a cobrança retoma no sábado. ⚠️ Não se trata de desconto. O valor total do débito permanece integral.\n\n\n*Ativuz Veículos*"
                else:
                    msg_pausar = f"{nome}, sua parcela de *{_brl(valor)}* está em aberto há *2 dias* (vencimento: {data_fmt}). O valor atualizado é *{_brl(total)}*. 📌 Pague *{_brl(pausar)}* hoje e quite *{_brl(pausar)}* até a sexta-feira desta semana. ⚠️ Juros de 0,5% ao dia continuam correndo sobre o saldo restante. Sem pagamento até sexta, a cobrança retoma no sábado. ⚠️ Não se trata de desconto. O valor total do débito permanece integral.\n\n\n*Ativuz Veículos*"
            else:
                msg_pausar = None

            _fone = _tel_map.get(_nh(nome), "")
            wa_cobranca = f"https://wa.me/{_fone}?text=" + _url_quote(msg)
            wa_pausar   = ((f"https://wa.me/{_fone}?text=" + _url_quote(msg_pausar))
                           if mostrar_pausar else None)

            registros_vencidos.append({
                "nome":             nome,
                "num_doc":          num_doc,
                "unidade":          unidade,
                "tipo_fatura":      tipo_fatura,
                "data_vencimento":  data_fmt,
                "dias_atraso":      dias,
                "dias_label":       dias_label,
                "reincidente":      reincidente,
                "is_fatura":        is_fatura,
                "tem_multa":        dias >= 2,
                "etapa":            etapa,
                "etapa_cls":        etapa_cls,
                "proxima_acao":     proxima,
                "situacao_key":     "vence-hoje" if dias == 0 else "vencido",
                "wa_cobranca":      wa_cobranca,
                "wa_pausar":        wa_pausar,
                "msg_cobranca_txt": msg,
                "msg_pausar_txt":   msg_pausar,
                "valor_s":          _brl(valor),
                "multa_s":          _brl(multa),
                "juros_mora_s":     _brl(juros_mora),
                "juros_s":          _brl(juros),
                "total_s":          _brl(total),
                "pausar_s":         _brl(pausar),
                "_valor":           valor,
                "_multa":           multa,
                "_juros_mora":      juros_mora,
                "_juros":           juros,
                "_total":           total,
                "_fone":            _fone,
            })

    except Exception:
        import traceback; traceback.print_exc()
        erro_leitura = "Erro ao ler a planilha."

    return registros_vencidos, registros_a_vencer, erro_leitura


def _inad_summary():
    """Resumo de inadimplência para o dashboard (delega a _ler_inad_dados)."""
    try:
        registros_vencidos, _, _ = _ler_inad_dados()
    except Exception:
        return None

    if not registros_vencidos:
        return {"total_s": _brl(0), "casos": 0, "hoje": 0,
                "por_etapa": {}, "recentes": [], "total_raw": 0}

    total_raw    = sum(r["_total"] for r in registros_vencidos)
    total_orig   = sum(r["_valor"] for r in registros_vencidos)
    nomes_unicos = {r["nome"] for r in registros_vencidos}
    hoje_count   = sum(1 for r in registros_vencidos if r["dias_atraso"] == 0)

    etapas = ["Hoje", "Terça-feira", "Quarta-feira", "Quinta-feira",
              "Sexta-feira", "D+5", "D+7", "D+10", "D+15"]
    por_etapa = {e: 0 for e in etapas}
    for r in registros_vencidos:
        if r["etapa"] in por_etapa:
            por_etapa[r["etapa"]] += 1

    recentes = sorted(registros_vencidos, key=lambda r: r["dias_atraso"], reverse=True)[:8]
    recentes_slim = [
        {
            "nome":    r["nome"],
            "placa":   r["num_doc"] or r["unidade"] or "—",
            "venc":    r["data_vencimento"],
            "dias":    r["dias_atraso"],
            "total_s": r["total_s"],
            "etapa":   r["etapa"],
        }
        for r in recentes
    ]

    # Consolidado por cliente (top 8 maiores devedores)
    _dev: dict = {}
    for r in registros_vencidos:
        n = r["nome"]
        if n not in _dev:
            _dev[n] = {"nome": n, "casos": 0, "total": 0.0}
        _dev[n]["casos"] += 1
        _dev[n]["total"] += r["_total"]
    top_devedores = sorted(_dev.values(), key=lambda x: x["total"], reverse=True)[:8]
    for d in top_devedores:
        d["total_s"] = _brl(d["total"])

    return {
        "total_s":      _brl(total_raw),
        "total_orig_s": _brl(total_orig),
        "total_raw":    total_raw,
        "casos":        len(nomes_unicos),
        "hoje":         hoje_count,
        "por_etapa":    por_etapa,
        "recentes":     recentes_slim,
        "top_devedores": top_devedores,
    }


@app.route("/inadimplencia")
def pagina_inadimplencia():
    from collections import Counter

    registros_vencidos, registros_a_vencer, erro_leitura = _ler_inad_dados()
    _sync_contas_supabase()

    registros_vencidos.sort(key=lambda r: r["dias_atraso"], reverse=True)
    registros_a_vencer.sort(key=lambda r: r["dias_ate"])

    hoje = date.today()
    total_vencidos        = len(registros_vencidos)
    total_a_vencer_cnt    = len(registros_a_vencer)
    total_valor_orig      = _brl(sum(r["_valor"] for r in registros_vencidos))
    total_valor_atual     = _brl(sum(r["_total"] for r in registros_vencidos))
    total_a_vencer_val    = _brl(sum(r["_valor"] for r in registros_a_vencer))
    criticos              = sum(1 for r in registros_vencidos if r["dias_atraso"] >= 7)
    reincidentes_criticos = sum(1 for r in registros_vencidos
                                if r["dias_atraso"] >= 7 and r["reincidente"])

    _vencido_raw   = sum(r["_valor"] for r in registros_vencidos)
    _a_vencer_raw  = sum(r["_valor"] for r in registros_a_vencer)
    # Base semanal = snapshot da segunda desta semana (ou mais recente disponível)
    _base_semanal = 0.0
    _taxa_debug   = ""
    _sb = _supabase()
    if _sb:
        _segunda_atual = (hoje - timedelta(days=hoje.weekday())).isoformat()
        try:
            # 1º: snapshot exato desta segunda
            _snap = (_sb.table("inad_snapshots")
                       .select("semana,total_valor")
                       .eq("semana", _segunda_atual)
                       .limit(1)
                       .execute())
            if _snap.data:
                _base_semanal = float(_snap.data[0].get("total_valor") or 0)
                _taxa_debug = f"snapshot {_segunda_atual}"
            else:
                # 2º: snapshot mais recente de qualquer segunda anterior
                _snaps_all = (_sb.table("inad_snapshots")
                                .select("semana,total_valor")
                                .lte("semana", _segunda_atual)
                                .order("semana", desc=True)
                                .limit(10)
                                .execute())
                for _s in (_snaps_all.data or []):
                    _d = date.fromisoformat(_s["semana"])
                    if _d.weekday() == 0:  # só segundas
                        _base_semanal = float(_s.get("total_valor") or 0)
                        _taxa_debug = f"snapshot {_s['semana']} (anterior)"
                        break
        except Exception:
            pass
    # Numerador = Total Original (todos os vencidos, valor original sem multa/juros)
    _overdue_raw   = _vencido_raw
    taxa_inadimplencia = round(_overdue_raw / _base_semanal * 100, 1) if _base_semanal > 0 else None

    _EXCLUIR_OCORR = {"segcomp", "onevo", "new charger", "m&s", "marcelo bento de araujo"}
    _nome_cnt = Counter(
        r["nome"] for r in registros_vencidos
        if _nh(r["nome"]) not in _EXCLUIR_OCORR
    )
    _nome_valor = {}
    for r in registros_vencidos:
        if _nh(r["nome"]) not in _EXCLUIR_OCORR:
            _nome_valor[r["nome"]] = _nome_valor.get(r["nome"], 0.0) + r["_total"]

    if _nome_cnt:
        critico_ocorr_nome = _nome_cnt.most_common(1)[0][0]
        critico_ocorr_qtd  = _nome_cnt.most_common(1)[0][1]
    else:
        critico_ocorr_nome, critico_ocorr_qtd = "—", 0

    if _nome_valor:
        _cv = max(_nome_valor, key=_nome_valor.get)
        critico_valor_nome  = _cv
        critico_valor_total = _brl(_nome_valor[_cv])
    else:
        critico_valor_nome, critico_valor_total = "—", "—"

    obs_map = {}
    try:
        sb_obs = _supabase()
        if sb_obs:
            obs_res = sb_obs.table("inad_observacoes").select("chave,texto").execute()
            obs_map = {r["chave"]: r["texto"] for r in (obs_res.data or [])}
    except Exception:
        pass

    return render_template(
        "inadimplencia.html",
        registros=registros_vencidos,
        registros_a_vencer=registros_a_vencer,
        total_registros=total_vencidos,
        total_a_vencer=total_a_vencer_cnt,
        total_valor_orig=total_valor_orig,
        total_valor_atual=total_valor_atual,
        total_a_vencer_val=total_a_vencer_val,
        criticos=criticos,
        reincidentes=reincidentes_criticos,
        critico_ocorr_nome=critico_ocorr_nome,
        critico_ocorr_qtd=critico_ocorr_qtd,
        critico_valor_nome=critico_valor_nome,
        critico_valor_total=critico_valor_total,
        taxa_inadimplencia=taxa_inadimplencia,
        taxa_numerador=round(_overdue_raw, 2),
        taxa_denominador=round(_base_semanal, 2),
        taxa_debug=_taxa_debug,
        erro_leitura=erro_leitura,
        hoje=hoje.strftime("%d/%m/%Y"),
        active="inadimplencia",
        obs_map=obs_map,
    )


@app.route("/inadimplencia/exportar")
def exportar_inadimplencia():
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border
    from openpyxl.cell.cell import MergedCell
    from collections import defaultdict

    _base      = Path(__file__).parent / "planilhas"
    xlsx_path  = _base / "CONTAS-A-RECEBER.xlsx"
    modelo     = _base / "Template_Inadimplencia.xlsx"
    hoje       = date.today()
    hoje_str   = hoje.strftime("%d/%m/%Y")
    hoje_fname = hoje.strftime("%d.%m.%y")

    # ── Re-lê e calcula registros (mesma lógica de pagina_inadimplencia) ──────
    registros = []
    if xlsx_path.exists():
        try:
            wb_src = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
            ws_src = wb_src.active
            rows   = list(ws_src.iter_rows(values_only=True))
            wb_src.close()

            header_idx = 0
            for ri, row in enumerate(rows[:10]):
                nh_row = [_nh(str(c or "")) for c in row]
                if sum(1 for t in ["receber de","vencimento","valor"] if any(t in n for n in nh_row)) >= 2:
                    header_idx = ri; break

            header    = rows[header_idx]
            data_rows = rows[header_idx + 1:]

            def _ci(kw):
                nk = _nh(kw)
                return next((i for i,h in enumerate(header) if h and nk in _nh(str(h))), None)

            i_nome  = _ci("receber de (fantasia)") or _ci("receber de")
            i_valor = _ci("valor previsto") or _ci("valor")
            i_venc  = _ci("data de vencimento") or _ci("vencimento")
            i_sit   = _ci("situacao (data de vencimento)") or _ci("situacao")
            i_doc   = _ci("numero do documento") or _ci("documento")
            i_unid  = _ci("unidade")

            def _gv(row, idx):
                return row[idx] if idx is not None and idx < len(row) else None

            for row in data_rows:
                nome_raw = _gv(row, i_nome)
                if not nome_raw or not str(nome_raw).strip():
                    continue
                nome  = str(nome_raw).strip()
                if nome.upper() in {"MARCELO BENTO DE ARAUJO"}:
                    continue
                valor = _parse_valor_excel(_gv(row, i_valor))
                if valor <= 0:
                    continue

                venc_raw  = _gv(row, i_venc)
                sit_raw   = _nh(str(_gv(row, i_sit) or ""))
                doc_raw   = _gv(row, i_doc)
                num_doc   = str(doc_raw or "").strip()
                for s in ("None", "nan", ""):
                    if num_doc == s: num_doc = ""
                venc_date = None
                if venc_raw:
                    if isinstance(venc_raw, datetime): venc_date = venc_raw.date()
                    elif isinstance(venc_raw, date):   venc_date = venc_raw
                    else:
                        for fmt in ["%d/%m/%Y","%Y-%m-%d","%d-%m-%Y"]:
                            try: venc_date = datetime.strptime(str(venc_raw).strip(), fmt).date(); break
                            except (ValueError, TypeError): pass
                if venc_date is None:
                    continue
                if "a vencer" in sit_raw and venc_date > hoje:
                    continue
                dias = 0 if venc_date == hoje else (hoje - venc_date).days
                if dias < 0:
                    continue

                tem_fatura = bool(num_doc)
                multa      = valor * 0.10 if (tem_fatura and dias >= 2) else 0.0
                juros_mora = (valor + multa) * 0.00033 * dias if (tem_fatura and dias >= 3) else 0.0
                juros      = multa + juros_mora
                total      = valor + juros

                _ETAPA_SHORT = {
                    "Hoje": "Hoje", "Terça-feira": "Terça", "Quarta-feira": "Quarta",
                    "Quinta-feira": "Quinta", "Sexta-feira": "Sexta",
                    "D+5": "D+5", "D+7": "D+7", "D+10": "D+10", "D+15": "D+15",
                }
                if   dias == 0:     etapa, proxima = "Hoje",         "Enviar lembrete de vencimento"
                elif dias == 1:     etapa, proxima = "Terça-feira",  "Aviso de atraso — tem até o final do dia para pagar, caso contrário amanhã entram os juros"
                elif dias == 2:     etapa, proxima = "Quarta-feira", "Juros aplicado — a partir de amanhã inicia a contagem dos juros de mora"
                elif dias == 3:     etapa, proxima = "Quinta-feira", "Juros de mora em contagem — regularize hoje para evitar suspensão do serviço"
                elif dias == 4:     etapa, proxima = "Sexta-feira",  "Aviso final — regularize até hoje ou o serviço será suspenso"
                elif dias <= 6:     etapa, proxima = "D+5",  "Serviço suspenso — exigir comprovante de pagamento para reativação"
                elif dias <= 9:     etapa, proxima = "D+7",  "Encaminhar para cobrança jurídica extrajudicial"
                elif dias <= 14:    etapa, proxima = "D+10", "Negativação no SPC/Serasa + encaminhamento jurídico"
                else:               etapa, proxima = "D+15", "Processo judicial iniciado — recolhimento imediato do veículo"

                registros.append({
                    "nome": nome, "etapa": etapa, "etapa_short": _ETAPA_SHORT.get(etapa, etapa),
                    "proxima": proxima, "vencimento": venc_date.strftime("%d/%m/%Y"), "dias": dias,
                    "valor": valor, "juros": juros, "total": total,
                })
        except Exception:
            import traceback; traceback.print_exc()

    # ── Gera Excel ────────────────────────────────────────────────────────────
    def _fill(hex6): return PatternFill("solid", fgColor=hex6)
    def _align(h="center", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    F_ROW_ODD = _fill("F0F4FA")
    F_ROW_EVN = _fill("FFFFFF")
    F_VALOR   = _fill("EFF6FF")
    F_JUROS   = _fill("FFF7ED")
    F_TOTAL_C = _fill("1E3A5F")
    FMT_BRL   = '"R$"\\ #,##0.00'

    # Cor de fundo por etapa (coluna Etapa)
    _ETAPA_BG = {
        "Hoje": "2563EB", "Terça": "1D4ED8", "Quarta": "1E40AF",
        "Quinta": "D97706", "Sexta": "B45309",
        "D+5": "DC2626", "D+7": "B91C1C", "D+10": "991B1B", "D+15": "7F1D1D",
    }
    _ETAPA_PRIO = {
        "Hoje": 1, "Terça": 2, "Quarta": 3, "Quinta": 4, "Sexta": 5,
        "D+5": 6, "D+7": 7, "D+10": 8, "D+15": 9,
    }

    F_NONE = PatternFill(fill_type=None)  # sem preenchimento (limpa fill)

    def _safe_set(cell, **kwargs):
        if isinstance(cell, MergedCell):
            return
        for attr, val in kwargs.items():
            setattr(cell, attr, val)

    def _unmerge_area(ws, min_row, max_row, min_col, max_col):
        to_remove = [
            m for m in list(ws.merged_cells.ranges)
            if m.min_row <= max_row and m.max_row >= min_row
            and m.min_col <= max_col and m.max_col >= min_col
        ]
        for m in to_remove:
            ws.unmerge_cells(str(m))

    def _clear_rows(ws, from_row, to_row, min_col, max_col):
        """Limpa valor e fill de células numa faixa de linhas."""
        _unmerge_area(ws, from_row, to_row, min_col, max_col)
        for r in range(from_row, to_row + 1):
            for c in range(min_col, max_col + 1):
                cell = ws.cell(row=r, column=c)
                if not isinstance(cell, MergedCell):
                    cell.value = None
                    cell.fill  = F_NONE

    wb = openpyxl.load_workbook(str(modelo))

    total_valor = sum(r["valor"] for r in registros)
    total_juros = sum(r["juros"] for r in registros)
    total_total = sum(r["total"] for r in registros)
    n = len(registros)

    F_TOT_FONT = Font(color="FFFFFF", bold=True, size=10)

    # ── Aba 1: Resumo Executivo ───────────────────────────────────────────────
    ws1 = wb["Resumo Executivo"]

    _safe_set(ws1["B3"],
              value=f"ATIVUZ VEÍCULOS  ·  Gerado em: {hoje_str}  ·  Todos os títulos — ordem por maior valor")

    D_INI = 10
    D_LIM = 300
    registros_val = sorted(registros, key=lambda x: x["valor"], reverse=True)
    D_FIM = D_INI + n - 1 if n > 0 else D_INI

    _clear_rows(ws1, D_INI, D_LIM, 2, 8)

    for i, rec in enumerate(registros_val):
        r    = D_INI + i
        base = F_ROW_ODD if i % 2 == 0 else F_ROW_EVN
        es   = rec["etapa_short"]
        F_ET = _fill(_ETAPA_BG.get(es, "374151"))
        F_EF = Font(color="FFFFFF", bold=True, size=9)

        _safe_set(ws1.cell(r, 2), value=rec["nome"],       fill=base,    font=Font(size=10),              alignment=_align("left"))
        _safe_set(ws1.cell(r, 3), value=es,                fill=F_ET,    font=F_EF,                       alignment=_align("center"))
        _safe_set(ws1.cell(r, 4), value=rec["vencimento"], fill=base,    font=Font(size=10),              alignment=_align("center"))
        _safe_set(ws1.cell(r, 5), value=rec["dias"],       fill=base,    font=Font(size=10),              alignment=_align("center"))
        _safe_set(ws1.cell(r, 6), value=rec["valor"],      fill=F_VALOR, font=Font(size=10), number_format=FMT_BRL, alignment=_align("right"))
        _safe_set(ws1.cell(r, 7), value=rec["juros"],      fill=F_JUROS, font=Font(size=10), number_format=FMT_BRL, alignment=_align("right"))
        _safe_set(ws1.cell(r, 8), value=rec["total"],      fill=F_ROW_EVN, font=Font(bold=True, size=10), number_format=FMT_BRL, alignment=_align("right"))

    T_ROW1 = D_FIM + 1
    _safe_set(ws1.cell(T_ROW1, 2), value="TOTAL", fill=F_TOTAL_C, font=F_TOT_FONT, alignment=_align("center"))
    for col in [3, 4, 5]:
        _safe_set(ws1.cell(T_ROW1, col), fill=F_TOTAL_C)
    _safe_set(ws1.cell(T_ROW1, 6), value=f"=SUM(F{D_INI}:F{D_FIM})" if n else 0,
              fill=F_TOTAL_C, font=F_TOT_FONT, number_format=FMT_BRL, alignment=_align("right"))
    _safe_set(ws1.cell(T_ROW1, 7), value=f"=SUM(G{D_INI}:G{D_FIM})" if n else 0,
              fill=F_TOTAL_C, font=F_TOT_FONT, number_format=FMT_BRL, alignment=_align("right"))
    _safe_set(ws1.cell(T_ROW1, 8), value=f"=SUM(H{D_INI}:H{D_FIM})" if n else 0,
              fill=F_TOTAL_C, font=F_TOT_FONT, number_format=FMT_BRL, alignment=_align("right"))

    # Exclui linhas abaixo do TOTAL
    if T_ROW1 + 1 <= D_LIM:
        ws1.delete_rows(T_ROW1 + 1, D_LIM - T_ROW1)

    _safe_set(ws1["B7"], value=n)
    _safe_set(ws1["C7"], value=total_valor, number_format=FMT_BRL)
    _safe_set(ws1["F7"], value=total_juros, number_format=FMT_BRL)
    _safe_set(ws1["H7"], value=total_total, number_format=FMT_BRL)

    # ── Aba 2: Detalhamento por Cliente ──────────────────────────────────────
    ws2 = wb["Detalhamento por Cliente"]

    _safe_set(ws2["B3"],
              value=f"Valores consolidados por cliente — ordem alfabética  ·  {hoje_str}")

    clientes_map = defaultdict(lambda: {"etapa_pior": "", "prio": 0, "titulos": 0, "valor": 0.0, "juros": 0.0})
    for rec in registros:
        c = clientes_map[rec["nome"]]
        c["titulos"] += 1
        c["valor"]   += rec["valor"]
        c["juros"]   += rec["juros"]
        p = _ETAPA_PRIO.get(rec["etapa_short"], 0)
        if p > c["prio"]:
            c["prio"] = p
            c["etapa_pior"] = rec["etapa_short"]

    clientes_sorted = sorted(clientes_map.items(), key=lambda x: x[0])
    nc = len(clientes_sorted)

    D_INI2 = 10
    D_FIM2 = D_INI2 + nc - 1 if nc > 0 else D_INI2

    _clear_rows(ws2, D_INI2, D_LIM, 2, 7)

    for i, (nome, g) in enumerate(clientes_sorted):
        r    = D_INI2 + i
        base = F_ROW_ODD if i % 2 == 0 else F_ROW_EVN
        F_ET = _fill(_ETAPA_BG.get(g["etapa_pior"], "374151"))
        F_EF = Font(color="FFFFFF", bold=True, size=9)
        total_cli = g["valor"] + g["juros"]

        _safe_set(ws2.cell(r, 2), value=nome,            fill=base,    font=Font(size=10),              alignment=_align("left"))
        _safe_set(ws2.cell(r, 3), value=g["etapa_pior"], fill=F_ET,    font=F_EF,                       alignment=_align("center"))
        _safe_set(ws2.cell(r, 4), value=g["titulos"],    fill=base,    font=Font(size=10),              alignment=_align("center"))
        _safe_set(ws2.cell(r, 5), value=g["valor"],      fill=F_VALOR, font=Font(size=10), number_format=FMT_BRL, alignment=_align("right"))
        _safe_set(ws2.cell(r, 6), value=g["juros"],      fill=F_JUROS, font=Font(size=10), number_format=FMT_BRL, alignment=_align("right"))
        _safe_set(ws2.cell(r, 7), value=total_cli,       fill=F_ROW_EVN, font=Font(bold=True, size=10), number_format=FMT_BRL, alignment=_align("right"))

    T_ROW2 = D_FIM2 + 1
    _safe_set(ws2.cell(T_ROW2, 2), value="TOTAL GERAL", fill=F_TOTAL_C, font=F_TOT_FONT, alignment=_align("center"))
    for col in [3, 4]:
        _safe_set(ws2.cell(T_ROW2, col), fill=F_TOTAL_C)
    _safe_set(ws2.cell(T_ROW2, 5), value=f"=SUM(E{D_INI2}:E{D_FIM2})" if nc else 0,
              fill=F_TOTAL_C, font=F_TOT_FONT, number_format=FMT_BRL, alignment=_align("right"))
    _safe_set(ws2.cell(T_ROW2, 6), value=f"=SUM(F{D_INI2}:F{D_FIM2})" if nc else 0,
              fill=F_TOTAL_C, font=F_TOT_FONT, number_format=FMT_BRL, alignment=_align("right"))
    _safe_set(ws2.cell(T_ROW2, 7), value=f"=SUM(G{D_INI2}:G{D_FIM2})" if nc else 0,
              fill=F_TOTAL_C, font=F_TOT_FONT, number_format=FMT_BRL, alignment=_align("right"))

    # Exclui linhas abaixo do TOTAL
    if T_ROW2 + 1 <= D_LIM:
        ws2.delete_rows(T_ROW2 + 1, D_LIM - T_ROW2)

    _safe_set(ws2["B7"], value=nc)
    _safe_set(ws2["D7"], value=total_valor, number_format=FMT_BRL)
    _safe_set(ws2["F7"], value=total_total, number_format=FMT_BRL)

    # ── Aba 3: Análise por Etapa ──────────────────────────────────────────────
    ws3 = wb["Análise por Etapa"]

    _safe_set(ws3["B3"],
              value=f"Resumo consolidado por etapa de cobrança  ·  {hoje_str}")

    ETAPA_ROWS3 = {
        "Hoje": 10, "Terça": 11, "Quarta": 12, "Quinta": 13, "Sexta": 14,
        "D+5": 15, "D+7": 16, "D+10": 17, "D+15": 18,
    }
    agrup = defaultdict(lambda: {"n": 0, "valor": 0.0, "juros": 0.0})
    for rec in registros:
        g = agrup[rec["etapa_short"]]
        g["n"] += 1; g["valor"] += rec["valor"]; g["juros"] += rec["juros"]

    for etapa_s, row_num in ETAPA_ROWS3.items():
        g = agrup[etapa_s]
        total_e = g["valor"] + g["juros"]
        _safe_set(ws3.cell(row_num, 3), value=g["n"],      alignment=_align("center"))
        _safe_set(ws3.cell(row_num, 4), value=g["valor"],  number_format=FMT_BRL, fill=F_VALOR)
        _safe_set(ws3.cell(row_num, 5), value=g["juros"],  number_format=FMT_BRL, fill=F_JUROS)
        _safe_set(ws3.cell(row_num, 6), value=total_e,     number_format=FMT_BRL)
        _safe_set(ws3.cell(row_num, 9), value=total_e,     number_format=FMT_BRL)

    # Total row (row 19 já tem label "TOTAL" no template)
    _safe_set(ws3.cell(19, 3), value="=SUM(C10:C18)", fill=F_TOTAL_C, font=F_TOT_FONT, alignment=_align("center"))
    _safe_set(ws3.cell(19, 4), value="=SUM(D10:D18)", fill=F_TOTAL_C, font=F_TOT_FONT, number_format=FMT_BRL, alignment=_align("right"))
    _safe_set(ws3.cell(19, 5), value="=SUM(E10:E18)", fill=F_TOTAL_C, font=F_TOT_FONT, number_format=FMT_BRL, alignment=_align("right"))
    _safe_set(ws3.cell(19, 6), value="=SUM(F10:F18)", fill=F_TOTAL_C, font=F_TOT_FONT, number_format=FMT_BRL, alignment=_align("right"))

    etapas_ativas = sum(1 for g in agrup.values() if g["n"] > 0)
    _safe_set(ws3["B7"], value=etapas_ativas)
    _safe_set(ws3["C7"], value=n)
    _safe_set(ws3["E7"], value=total_total, number_format=FMT_BRL)

    # ── Serve o arquivo ───────────────────────────────────────────────────────
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    nome_arquivo = f"Relatório_Inadimplência_{hoje_fname}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=nome_arquivo,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _commitar_arquivo_github(repo, path, conteudo, mensagem, branch="main"):
    """
    Sobe/atualiza um arquivo direto no repositório via API de Contents do
    GitHub, sem depender do sistema de arquivos do Vercel (somente leitura
    em produção — um f.save() local ali falha ou não persiste). Cada commit
    criado aqui dispara redeploy automático, já que o Vercel está integrado
    ao repositório. Lança exceção com o motivo em caso de falha.
    """
    import requests as _req
    import base64 as _b64

    token = _os.environ.get("GITHUB_TOKEN", "")
    if not token:
        raise Exception("GITHUB_TOKEN não configurado.")

    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/vnd.github+json",
    }
    url = f"https://api.github.com/repos/{repo}/contents/{path}"

    resp = _req.get(url, headers=headers, params={"ref": branch}, timeout=15)
    sha = resp.json().get("sha") if resp.status_code == 200 else None

    payload = {
        "message": mensagem,
        "content": _b64.b64encode(conteudo).decode(),
        "branch": branch,
    }
    if sha:
        payload["sha"] = sha

    put = _req.put(url, headers=headers, json=payload, timeout=15)
    if put.status_code not in (200, 201):
        raise Exception(f"HTTP {put.status_code}: {put.text}")


@app.route("/inadimplencia/upload", methods=["POST"])
def inadimplencia_upload():
    f = request.files.get("planilha")
    if not f or not f.filename:
        flash("Nenhum arquivo selecionado.", "error")
        return redirect(url_for("pagina_inadimplencia"))
    if not f.filename.lower().endswith(".xlsx"):
        flash("Apenas arquivos .xlsx são aceitos.", "error")
        return redirect(url_for("pagina_inadimplencia"))

    conteudo = f.read()

    if _os.environ.get("GITHUB_TOKEN"):
        try:
            _commitar_arquivo_github(
                repo="grupoativuz/Dashboard-Ativuz",
                path="planilhas/CONTAS-A-RECEBER.xlsx",
                conteudo=conteudo,
                mensagem="chore: atualiza CONTAS-A-RECEBER.xlsx via upload no Dashboard",
            )
            flash("Planilha enviada! O Dashboard vai atualizar em cerca de 1 minuto, "
                  "quando o novo deploy no Vercel terminar.", "success")
        except Exception as e:
            flash(f"Falha ao enviar a planilha pro GitHub: {e}", "error")
        return redirect(url_for("pagina_inadimplencia"))

    # Sem GITHUB_TOKEN (ex: rodando local): grava direto no disco, como antes.
    dest = Path(__file__).parent / "planilhas" / "CONTAS-A-RECEBER.xlsx"
    dest.write_bytes(conteudo)
    flash("Planilha atualizada! Os dados abaixo refletem o arquivo enviado.", "success")
    return redirect(url_for("pagina_inadimplencia"))


def _checar_token_inadimplencia():
    token_esperado = _os.environ.get("INADIMPLENCIA_API_TOKEN", "")
    token_recebido = request.headers.get("X-Inadimplencia-Token", "")
    return bool(token_esperado) and token_recebido == token_esperado


@app.route("/api/inadimplencia/upload", methods=["POST"])
def api_inadimplencia_upload():
    """
    Recebe CONTAS-A-RECEBER.xlsx da automação do Gmail (Apps Script, sem sessão
    de login) e commita no GitHub. Autenticado por token compartilhado, mesmo
    esquema de /api/contratos/ativos.
    """
    if not _checar_token_inadimplencia():
        return jsonify({"error": "Token inválido."}), 401

    f = request.files.get("planilha")
    if not f or not f.filename:
        return jsonify({"error": "Nenhum arquivo enviado."}), 400

    if not _os.environ.get("GITHUB_TOKEN"):
        return jsonify({"error": "GITHUB_TOKEN não configurado no servidor."}), 500

    try:
        _commitar_arquivo_github(
            repo="grupoativuz/Dashboard-Ativuz",
            path="planilhas/CONTAS-A-RECEBER.xlsx",
            conteudo=f.read(),
            mensagem="chore: atualiza CONTAS-A-RECEBER.xlsx via automação (Gmail)",
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    return jsonify({"ok": True})


# ── Linha do Tempo: snapshot diário ──────────────────────────────────────────

_DIAS_PT = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]


def _fmt_dia(d):
    return f"{_DIAS_PT[d.weekday()]} {d.strftime('%d/%m')}"


@app.route("/inadimplencia/snapshot", methods=["POST"])
def inadimplencia_snapshot():
    """Salva snapshot do dia atual com os dados atuais da planilha."""
    sb = _supabase()
    if not sb:
        return jsonify({"ok": False, "erro": "Supabase não configurado."}), 500

    try:
        registros_vencidos, _, _ = _ler_inad_dados()
        hoje = date.today()

        nomes_unicos = {r["nome"] for r in registros_vencidos}
        total_casos  = len(nomes_unicos)
        total_valor  = sum(r["_valor"] for r in registros_vencidos)  # valor original, sem multa/juros
        criticos     = sum(1 for r in registros_vencidos if r["dias_atraso"] >= 7)
        base_semanal = sum(r["_valor"] for r in registros_vencidos if r["dias_atraso"] == 0)

        etapas = ["Hoje", "Terça-feira", "Quarta-feira", "Quinta-feira",
                  "Sexta-feira", "D+5", "D+7", "D+10", "D+15"]
        por_etapa = {e: 0 for e in etapas}
        for r in registros_vencidos:
            if r["etapa"] in por_etapa:
                por_etapa[r["etapa"]] += 1

        dia_str  = hoje.isoformat()
        existing = sb.table("inad_snapshots").select("id").eq("semana", dia_str).execute()
        payload  = {
            "semana":       dia_str,
            "total_casos":  total_casos,
            "total_valor":  round(total_valor, 2),
            "criticos":     criticos,
            "por_etapa":    por_etapa,
            "base_semanal": round(base_semanal, 2),
        }
        if existing.data:
            sb.table("inad_snapshots").update(payload).eq("semana", dia_str).execute()
            msg = f"Snapshot de {_fmt_dia(hoje)} atualizado."
        else:
            sb.table("inad_snapshots").insert(payload).execute()
            msg = f"Snapshot de {_fmt_dia(hoje)} salvo."

        return jsonify({"ok": True, "msg": msg, "semana": dia_str,
                        "total_casos": total_casos, "total_valor": round(total_valor, 2),
                        "criticos": criticos, "base_semanal": round(base_semanal, 2)})
    except Exception:
        import traceback; traceback.print_exc()
        return jsonify({"ok": False, "erro": "Erro ao salvar snapshot."}), 500


@app.route("/inadimplencia/historico", methods=["GET"])
def inadimplencia_historico():
    """Retorna snapshots históricos como JSON para o gráfico."""
    sb = _supabase()
    if not sb:
        return jsonify([])
    try:
        res = sb.table("inad_snapshots") \
                .select("semana,total_casos,total_valor,criticos,por_etapa") \
                .order("semana", desc=False) \
                .execute()
        rows = res.data or []
        for r in rows:
            d = date.fromisoformat(r["semana"])
            r["semana_fmt"]      = _fmt_dia(d)
            r["total_valor_fmt"] = _brl(float(r["total_valor"]))
        return jsonify(rows)
    except Exception:
        return jsonify([])


@app.route("/inadimplencia/historico/manual", methods=["POST"])
def inadimplencia_historico_manual():
    """Insere ou atualiza um snapshot manual de dia anterior."""
    sb = _supabase()
    if not sb:
        return jsonify({"ok": False, "erro": "Supabase não configurado."}), 500
    try:
        data        = request.get_json(force=True)
        dia_raw     = str(data.get("semana", "")).strip()
        total_casos = int(data.get("total_casos", 0))
        total_valor = float(str(data.get("total_valor", "0")).replace(",", "."))
        criticos    = int(data.get("criticos", 0))

        dia     = date.fromisoformat(dia_raw)
        dia_str = dia.isoformat()

        payload = {
            "semana":      dia_str,
            "total_casos": total_casos,
            "total_valor": round(total_valor, 2),
            "criticos":    criticos,
            "por_etapa":   {},
        }
        existing = sb.table("inad_snapshots").select("id").eq("semana", dia_str).execute()
        if existing.data:
            sb.table("inad_snapshots").update(payload).eq("semana", dia_str).execute()
        else:
            sb.table("inad_snapshots").insert(payload).execute()

        return jsonify({"ok": True, "semana": dia_str,
                        "semana_fmt": _fmt_dia(dia)})
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)}), 400


@app.route("/inadimplencia/historico/delete", methods=["POST"])
def inadimplencia_historico_delete():
    """Remove um snapshot pelo campo semana (YYYY-MM-DD)."""
    sb = _supabase()
    if not sb:
        return jsonify({"ok": False, "erro": "Supabase não configurado."}), 500
    try:
        data    = request.get_json(force=True)
        dia_str = str(data.get("semana", "")).strip()
        date.fromisoformat(dia_str)
        sb.table("inad_snapshots").delete().eq("semana", dia_str).execute()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)}), 400


_MESES_PT = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
             "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
_MESES_PT_CURTO = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun",
                   "Jul", "Ago", "Set", "Out", "Nov", "Dez"]


def _nome_mes_label(mes, ano, acumulado=False):
    if acumulado:
        return f"Jan–{_MESES_PT_CURTO[mes - 1]} {ano}"
    return f"{_MESES_PT[mes - 1]} {ano}"


# ── Configurações DRE ────────────────────────────────────────────────────────
# IR/CSLL: manter 0.0 para Simples Nacional; alterar aqui para mudar regime.
_DRE_IR_CSLL = 0.0

# Códigos cujo Tipo no sistema (ENTRADA/SAÍDA) é oposto ao tratamento no DRE.
# Ao adicionar novos, revisar o grupo de destino no _DRE_LAYOUT abaixo.
_DRE_CODIGOS_SINAL_INVERTIDO = frozenset([
    "02.02.06.006",  # SAÍDA → receita operacional (Combustível Reembolsável)
    "01.01.02.008",  # ENTRADA → dedução (Desconto Concedido a Clientes)
    "02.04.06.004",  # SAÍDA → receita financeira (Desconto Pgto Boletos)
])

# ── DRE Layout ────────────────────────────────────────────────────────────────
# sign: +1 = receita, -1 = despesa.  item tuple: (codigo, label).
# Códigos tratados como string — nunca converter para número.

_DRE_LAYOUT = [
    {"id": "rb", "label": "(+) Receita Operacional Bruta", "grupos": [
        {"id": "rb-loc",  "label": "Locação", "sign": +1, "itens": [
            ("01.01.01.001", "Locação"),
            ("01.01.01.002", "KM Excedente"),
            ("01.01.01.003", "Multa Atraso Pagamento"),
            ("01.01.01.004", "Multa Quebra de Contrato"),
            ("01.01.01.005", "Acordo/Renegociação"),
            ("01.01.01.006", "Taxa de Adm. de Veículos"),
        ]},
        {"id": "rb-reimb", "label": "Reembolsos", "sign": +1, "itens": [
            ("01.01.02.001", "Manutenção Reembolsável"),
            ("01.01.02.002", "Entrada de Multa de Trânsito"),
            ("01.01.02.004", "Multa Dev. Antecipada"),
            ("01.01.02.005", "Reembolso de Sinistro"),
            ("01.01.02.006", "Outros Reembolsos"),
            ("02.02.06.006", "Combustível Reembolsável"),  # SAÍDA no sistema → sinal invertido
        ]},
        {"id": "rb-fi", "label": "Reembolsos — Frota Investidores", "sign": +1, "itens": [
            ("01.01.02.009", "Reembolso de Manutenções — FI"),
            ("01.01.02.010", "Reembolso de Multas — FI"),
            ("01.01.02.011", "Reembolso Desp. Operacionais — FI"),
            ("01.02.01.005", "Recebimentos — Frota Investidores"),
        ]},
    ]},
    {"id": "ded", "label": "(-) Deduções", "grupos": [
        {"id": "ded-imp",  "label": "Impostos", "sign": -1,
         "note": "incide sobre o faturamento — Simples Nacional", "itens": [
            ("02.01.01.001", "PIS"),
            ("02.01.01.005", "Simples Nacional"),
            ("02.01.01.006", "Outros Impostos"),
        ]},
        {"id": "ded-desc", "label": "Descontos", "sign": -1, "itens": [
            ("01.01.02.008", "Desconto Concedido a Clientes"),  # ENTRADA → sinal invertido
        ]},
    ]},
    # subtotal: receita_liquida
    {"id": "custos", "label": "(-) Custos — Custo Direto da Operação/Frota", "grupos": [
        {"id": "c-lic",   "label": "Licenciamento", "sign": -1, "itens": [
            ("02.02.01.001", "Emplacamento"),
            ("02.02.01.002", "Transferência Veicular"),
            ("02.02.01.003", "IPVA"),
            ("02.02.01.005", "Taxa de Licenciamento"),
            ("02.02.01.006", "Despesas com Cartório"),
            ("02.02.01.007", "Transferência Veicular"),
            ("02.02.01.008", "Taxa Bombeiros"),
            ("02.02.01.009", "Vistoria"),
            ("02.02.01.010", "Documentação Veicular"),
        ]},
        {"id": "c-desp",  "label": "Honorários Despachante", "sign": -1, "itens": [
            ("02.02.02.001", "Honorários Despachante"),
        ]},
        {"id": "c-seg",   "label": "Seguro/Assistência 24h", "sign": -1, "itens": [
            ("02.02.03.001", "Seguro Total"),
            ("02.02.03.003", "Reserva Operacional p/ Sinistros"),
            ("02.02.03.004", "Guincho"),
            ("02.02.03.005", "Rastreador Veicular"),
        ]},
        {"id": "c-sub",   "label": "Sublocação/Transporte", "sign": -1, "itens": [
            ("02.02.04.002", "Frete Não Reembolsável"),
            ("02.02.04.003", "Multas Não Reembolsáveis"),
            ("02.02.04.004", "Taxa de Frete"),
            ("02.02.04.005", "Gasolina"),
            ("02.02.04.006", "Uber ou App de Transporte"),
        ]},
        {"id": "c-man",   "label": "Manutenção", "sign": -1, "itens": [
            ("02.02.05.001", "Manutenção Preventiva"),
            ("02.02.05.002", "Manutenção Corretiva"),
            ("02.02.05.006", "Compra Equipamento GNV"),
            ("02.02.05.007", "Equipamentos (Sensor, SIM, etc.)"),
            ("02.02.05.008", "Lavagem Veicular Não Reembolsável"),
            ("02.02.05.009", "Compra de Peças"),
            ("02.02.05.010", "Compra de Pneus"),
        ]},
        {"id": "c-reimb", "label": "Despesas Reembolsáveis", "sign": -1, "itens": [
            ("02.02.06.001", "Saída de Multa de Trânsito"),
            ("02.02.06.002", "Saída de Sinistro"),
            ("02.02.06.010", "Reembolso de Clientes"),
        ]},
        {"id": "c-fi",    "label": "Despesas c/ Frota de Investidores", "sign": -1, "itens": [
            ("02.02.07.01", "Manutenções — Frota Investidores"),
            ("02.02.07.02", "Sinistros — Frota Investidores"),
            ("02.02.07.03", "Desp. Operacionais — FI"),
            ("02.02.07.04", "Desp. Operacionais — FI"),
        ]},
    ]},
    # subtotal: margem
    {"id": "sga", "label": "(-) SG&A — Despesas Gerais e Administrativas", "grupos": [
        {"id": "s-sal", "label": "Salários", "sign": -1, "itens": [
            ("02.03.01.001", "Salário"),
            ("02.03.01.002", "Adiantamento Salarial"),
            ("02.03.01.003", "Férias"),
            ("02.03.01.004", "13° Salário"),
            ("02.03.01.005", "Rescisão"),
            ("02.03.01.006", "Prêmios"),
            ("02.03.01.007", "Comissão"),
            ("02.03.01.10",  "ASO e Saúde e Seg. do Trabalho"),
        ]},
        {"id": "s-ben", "label": "Benefícios", "sign": -1, "itens": [
            ("02.03.02.001", "VT"),
            ("02.03.02.003", "VR"),
            ("02.03.02.005", "Assistência Médica"),
            ("02.03.02.006", "P.C.M.S.O"),
            ("02.03.02.007", "Treinamento"),
            ("02.03.02.008", "Outros Benefícios"),
        ]},
        {"id": "s-imp", "label": "Impostos Folha", "sign": -1, "itens": [
            ("02.03.03.001", "INSS"),
            ("02.03.03.002", "FGTS"),
        ]},
        {"id": "s-pro", "label": "Pró-labore", "sign": -1,
         "highlight": True, "note": "remuneração dos sócios", "itens": [
            ("02.03.04.001", "Pró-labore Folha"),
        ]},
        {"id": "s-com", "label": "Despesas Comerciais", "sign": -1, "itens": [
            ("02.04.01.001", "Marketing"),
        ]},
        {"id": "s-ocu", "label": "Ocupação", "sign": -1, "itens": [
            ("02.04.02.001", "Aluguel de Imóveis"),
            ("02.04.02.003", "IPTU"),
            ("02.04.02.004", "Água"),
            ("02.04.02.005", "Luz"),
            ("02.04.02.006", "Manutenção Predial"),
        ]},
        {"id": "s-sup", "label": "Suprimentos", "sign": -1, "itens": [
            ("02.04.03.001", "Telefone"),
            ("02.04.03.007", "Bens de Pequeno Valor"),
        ]},
        {"id": "s-adm", "label": "Despesas Administrativas", "sign": -1, "itens": [
            ("02.04.04.001", "Desp. Administrativas e de Escritório"),
            ("02.04.04.004", "Taxas e Despesas Legais"),
            ("02.04.04.005", "Outras Despesas"),
            ("02.04.04.007", "Despesas Jurídicas"),
        ]},
        {"id": "s-svc", "label": "Serviços Prestados", "sign": -1, "itens": [
            ("02.04.05.001", "Honorários Advocatícios"),
            ("02.04.05.003", "Softwares"),
            ("02.04.05.004", "Órgãos de Proteção ao Crédito"),
            ("02.04.05.005", "Assessoria Administrativa"),
            ("02.04.05.006", "Honorários de Consultoria"),
            ("02.04.05.007", "Serviços Contábeis"),
            ("02.04.05.008", "Serviços de Limpeza"),
            ("02.04.05.009", "Serviços Manut. Máquinas e Equip."),
        ]},
        {"id": "s-out", "label": "Outras Saídas", "sign": -1, "itens": [
            ("02.04.07.002", "Outras Saídas"),
        ]},
    ]},
    # subtotal: ebitda; depreciação=0; subtotal: ebit
    {"id": "rfin", "label": "(-) Resultado Financeiro", "grupos": [
        {"id": "rf-desp", "label": "Despesas Bancárias e Financeiras", "sign": -1, "itens": [
            ("02.04.06.001", "Tarifa Bancária"),
            ("02.04.06.003", "Juros e Multas Bancárias Pagos"),
            ("02.04.06.005", "Taxa Maquineta"),
            ("03.01.03.002", "Consórcio Contemplado Juros"),
            ("03.01.03.004", "Financiamento Juros"),
            ("04.01.02.002", "Pgto Juros sobre Mútuos"),
        ]},
        {"id": "rf-rec",  "label": "Receitas Financeiras", "sign": +1, "itens": [
            ("03.03.01.003", "Rendimento de Aplicações"),
            ("02.04.06.004", "Desconto Pgto Boletos"),  # SAÍDA no sistema → receita financeira
        ]},
    ]},
    {"id": "rnop", "label": "(-) Resultados Não Operacionais", "grupos": [
        {"id": "rnop-out", "label": "Outras Entradas", "sign": +1, "itens": [
            ("01.02.01.002", "Depósitos Não Identificados"),
            ("01.02.01.003", "Outras Entradas"),
        ]},
        {"id": "rnop-inv", "label": "Outros Investimentos", "sign": +1, "itens": [
            ("03.03.02.001", "Outros Investimentos"),
        ]},
    ]},
    # subtotal: lucro_liquido ── abaixo: fluxo de caixa
    {"id": "inv", "label": "(-) Investimentos", "grupos": [
        {"id": "inv-venda", "label": "Venda de Veículos", "sign": +1, "itens": [
            ("01.02.01.004", "Venda de Veículo"),
        ]},
        {"id": "inv-comp",  "label": "Compra de Veículos", "sign": -1, "itens": [
            ("03.01.02.001", "Compra de Veículos à Vista"),
            ("03.01.02.002", "Entrada Compra Veículo"),
            ("03.01.02.003", "Adiantamento de Consórcio"),
        ]},
        {"id": "inv-fin",   "label": "Financiamentos Veiculares", "sign": -1, "itens": [
            ("03.01.03.001", "Consórcio Contemplado"),
            ("03.01.03.003", "Pagamento de Financiamento"),
            ("03.01.03.005", "Consórcio Parcela Não Contemplada"),
            ("03.01.03.006", "Quitação Antecipada de Parcelas"),
        ]},
        {"id": "inv-imob",  "label": "Outros Imobilizados", "sign": -1, "itens": [
            ("03.02.01.001", "Instalações"),
            ("03.02.01.002", "Computadores e Periféricos"),
            ("03.02.01.003", "Móveis e Utensílios"),
            ("03.02.01.004", "Sistemas e Softwares"),
            ("03.02.01.005", "Outras Imobilizações"),
        ]},
        {"id": "inv-obra",  "label": "Construção da Oficina", "sign": -1, "itens": [
            ("02.04.08.01", "Compra de Material para Oficina"),
            ("02.04.08.02", "Pagamento da Mão de Obra"),
            ("02.04.08.03", "Aluguel de Equipamentos"),
        ]},
        {"id": "inv-aplic", "label": "Aplicações Financeiras", "sign": -1, "itens": [
            ("03.03.01.001", "Aplicações Financeiras"),
        ]},
        {"id": "inv-resg",  "label": "Resgate de Aplicação", "sign": +1, "itens": [
            ("03.03.01.002", "Resgate de Aplicação Financeira"),
        ]},
    ]},
    {"id": "financ", "label": "(-) Financiamentos", "grupos": [
        {"id": "fin-ent",      "label": "Entradas de Mútuos", "sign": +1, "itens": [
            ("04.01.01.001", "Entrada de Mútuos"),
            ("04.01.01.002", "Entrada Caução"),
        ]},
        {"id": "fin-pgto",     "label": "Pgto de Mútuos", "sign": -1, "itens": [
            ("04.01.02.001", "Saída de Mútuos"),
            ("04.01.02.003", "Saída Caução"),
        ]},
        {"id": "fin-reimb-in", "label": "Entrada de Reembolso", "sign": +1, "itens": [
            ("04.04.02.01", "Entrada de Reembolso"),
        ]},
        {"id": "fin-reimb-out","label": "Saída de Reembolso", "sign": -1, "itens": [
            ("04.04.02.02", "Saída de Reembolso"),
        ]},
    ]},
    {"id": "aporte", "label": "(+) Aporte de Capital", "grupos": [
        {"id": "aporte-g", "label": "Aporte de Capital", "sign": +1, "itens": [
            ("04.04.01.003", "Aporte de Capital"),
        ]},
    ]},
    # subtotal: fluxo_acionista
    {"id": "distrib", "label": "(-) Distribuição de Resultado", "grupos": [
        {"id": "distrib-g",     "label": "Distribuição de Resultado", "sign": -1, "itens": [
            ("04.04.01.002", "Distribuição de Resultado"),
        ]},
        {"id": "distrib-lucro", "label": "Distribuição de Lucros Mensal", "sign": -1, "itens": [
            ("02.03.04.002", "Distribuição de Lucros Mensal"),
        ]},
    ]},
    # subtotal: fluxo_livre
]


# ── Categorias de DRE (espelho da vinculação feita no Blue Fleet) ────────────
# 2026-08-22: por decisão da diretoria, as naturezas do bloco FROTA INVESTIDORES
# foram realocadas para as categorias operacionais (o bloco deixou de existir).
# Fonte: Administração > Sistema > Categorias de DRE. Naturezas sem categoria
# vinculada no ERP ficam fora deste mapa e não entram nos indicadores.
_DRE_CATEGORIA_POR_NATUREZA = {
    "01.01.01.001": "RECEITAS DE LOCAÇÃO",
    "01.01.01.002": "RECEITAS ADICIONAIS",
    "01.01.01.003": "RECEITAS ADICIONAIS",
    "01.01.01.004": "RECEITAS ADICIONAIS",
    "01.01.01.005": "RECEITAS ADICIONAIS",
    "01.01.01.006": "RECEITAS ADICIONAIS",
    "01.01.01.007": "RECEITAS DE LOCAÇÃO",
    "01.01.02.001": "REEMBOLSOS",
    "01.01.02.002": "REEMBOLSOS",
    "01.01.02.004": "REEMBOLSOS",
    "01.01.02.005": "REEMBOLSOS",
    "01.01.02.006": "REEMBOLSOS",
    "01.01.02.008": "DEDUÇÕES",
    "01.01.02.009": "REEMBOLSOS",
    "01.01.02.010": "REEMBOLSOS",
    "01.01.02.011": "REEMBOLSOS",
    "01.02.01.002": "RESULTADO NÃO OPERACIONAL",
    "01.02.01.003": "RESULTADO NÃO OPERACIONAL",
    "01.02.01.004": "VENDA DE VEÍCULOS",
    "02.01.01.005": "DEDUÇÕES",
    "02.01.01.006": "DEDUÇÕES",
    "02.02.01.001": "CUSTOS DIRETOS DA FROTA",
    "02.02.01.002": "CUSTOS DIRETOS DA FROTA",
    "02.02.01.003": "CUSTOS DIRETOS DA FROTA",
    "02.02.01.005": "CUSTOS DIRETOS DA FROTA",
    "02.02.01.006": "CUSTOS DIRETOS DA FROTA",
    "02.02.01.007": "CUSTOS DIRETOS DA FROTA",
    "02.02.01.008": "CUSTOS DIRETOS DA FROTA",
    "02.02.01.009": "CUSTOS DIRETOS DA FROTA",
    "02.02.01.010": "CUSTOS DIRETOS DA FROTA",
    "02.02.02.001": "CUSTOS DIRETOS DA FROTA",
    "02.02.03.001": "CUSTOS DIRETOS DA FROTA",
    "02.02.03.004": "CUSTOS DIRETOS DA FROTA",
    "02.02.03.005": "CUSTOS DIRETOS DA FROTA",
    "02.02.04.003": "CUSTOS DIRETOS DA FROTA",
    "02.02.04.004": "CUSTOS DIRETOS DA FROTA",
    "02.02.04.005": "CUSTOS DIRETOS DA FROTA",
    "02.02.04.006": "CUSTOS DIRETOS DA FROTA",
    "02.02.05.001": "CUSTOS OPERACIONAIS",
    "02.02.05.002": "CUSTOS OPERACIONAIS",
    "02.02.05.006": "CUSTOS OPERACIONAIS",
    "02.02.05.007": "CUSTOS OPERACIONAIS",
    "02.02.05.008": "CUSTOS OPERACIONAIS",
    "02.02.05.009": "CUSTOS OPERACIONAIS",
    "02.02.05.010": "CUSTOS OPERACIONAIS",
    "02.02.06.001": "CUSTOS DIRETOS DA FROTA",
    "02.02.06.002": "CUSTOS DIRETOS DA FROTA",
    "02.02.06.006": "CUSTOS DIRETOS DA FROTA",
    "02.02.06.010": "CUSTOS DIRETOS DA FROTA",
    "02.02.07.01": "CUSTOS OPERACIONAIS",
    "02.02.07.03": "CUSTOS DIRETOS DA FROTA",
    "02.02.07.04": "CUSTOS DIRETOS DA FROTA",
    "02.03.01.001": "ADMINISTRATIVAS",
    "02.03.01.002": "ADMINISTRATIVAS",
    "02.03.01.003": "ADMINISTRATIVAS",
    "02.03.01.004": "ADMINISTRATIVAS",
    "02.03.01.005": "ADMINISTRATIVAS",
    "02.03.01.006": "ADMINISTRATIVAS",
    "02.03.01.007": "ADMINISTRATIVAS",
    "02.03.02.001": "ADMINISTRATIVAS",
    "02.03.02.003": "ADMINISTRATIVAS",
    "02.03.02.005": "ADMINISTRATIVAS",
    "02.03.02.006": "ADMINISTRATIVAS",
    "02.03.02.007": "ADMINISTRATIVAS",
    "02.03.02.008": "ADMINISTRATIVAS",
    "02.03.03.001": "ADMINISTRATIVAS",
    "02.03.03.002": "ADMINISTRATIVAS",
    "02.03.04.001": "ADMINISTRATIVAS",
    "02.04.01.001": "COMERCIAIS",
    "02.04.02.001": "ADMINISTRATIVAS",
    "02.04.02.003": "ADMINISTRATIVAS",
    "02.04.02.004": "ADMINISTRATIVAS",
    "02.04.02.005": "ADMINISTRATIVAS",
    "02.04.02.006": "ADMINISTRATIVAS",
    "02.04.03.001": "ADMINISTRATIVAS",
    "02.04.04.001": "ADMINISTRATIVAS",
    "02.04.04.004": "ADMINISTRATIVAS",
    "02.04.04.005": "ADMINISTRATIVAS",
    "02.04.04.007": "ADMINISTRATIVAS",
    "02.04.05.001": "ADMINISTRATIVAS",
    "02.04.05.003": "TECNOLOGIA",
    "02.04.05.004": "ADMINISTRATIVAS",
    "02.04.05.006": "ADMINISTRATIVAS",
    "02.04.05.007": "ADMINISTRATIVAS",
    "02.04.05.008": "ADMINISTRATIVAS",
    "02.04.05.009": "ADMINISTRATIVAS",
    "02.04.06.001": "DESPESAS FINANCEIRAS",
    "02.04.06.003": "DESPESAS FINANCEIRAS",
    "02.04.06.005": "DESPESAS FINANCEIRAS",
    "02.04.08.01": "CONSTRUÇÃO DA OFICINA",
    "02.04.08.02": "CONSTRUÇÃO DA OFICINA",
    "02.04.08.03": "CONSTRUÇÃO DA OFICINA",
    "03.01.02.001": "COMPRA DE VEÍCULOS",
    "03.01.02.002": "COMPRA DE VEÍCULOS",
    "03.01.02.003": "COMPRA DE VEÍCULOS",
    "03.01.03.001": "FINANCIAMENTOS VEICULARES",
    "03.01.03.002": "DESPESAS FINANCEIRAS",
    "03.01.03.003": "FINANCIAMENTOS VEICULARES",
    "03.01.03.004": "DESPESAS FINANCEIRAS",
    "03.01.03.005": "FINANCIAMENTOS VEICULARES",
    "03.02.01.001": "OUTROS IMOBILIZADOS",
    "03.02.01.002": "OUTROS IMOBILIZADOS",
    "03.02.01.003": "OUTROS IMOBILIZADOS",
    "03.02.01.004": "OUTROS IMOBILIZADOS",
    "03.02.01.005": "OUTROS IMOBILIZADOS",
    "03.03.01.001": "APLICAÇÃO FINANCEIRA",
    "03.03.01.002": "APLICAÇÃO FINANCEIRA",
    "03.03.01.003": "RECEITAS FINANCEIRAS",
    "03.03.02.001": "RESULTADO NÃO OPERACIONAL",
    "04.01.01.001": "ENTRADA DE MUTUOS",
    "04.01.01.002": "CAUÇÕES DE CLIENTES",
    "04.01.02.001": "SAÍDA DE MUTUOS",
    "04.01.02.002": "DESPESAS FINANCEIRAS",
    "04.01.02.003": "CAUÇÕES DE CLIENTES",
    "04.04.01.003": "APORTE DE CAPITAL",
    "02.01.01.001": "ADMINISTRATIVAS",
    "02.04.06.004": "CUSTOS OPERACIONAIS",
    "04.04.02.01": "ENTRADA DE MUTUOS",
    "04.04.02.02": "SAÍDA DE MUTUOS",
    "02.03.04.002": "DISTRIBUIÇÃO DE RESULTADO",
    "04.04.01.002": "DISTRIBUIÇÃO DE RESULTADO",
    "02.02.04.002": "CUSTOS DIRETOS DA FROTA",
    "02.04.05.005": "ADMINISTRATIVAS",
    "02.03.01.10": "ADMINISTRATIVAS",
    "02.04.03.007": "ADMINISTRATIVAS",
    "02.04.07.002": "ADMINISTRATIVAS",
    "01.01.02.012": "REPASSE A TERCEIROS",   # transferência da conta do investidor para a conta pessoal dele
    "01.02.01.005": "RECEITAS ADICIONAIS",
    "03.01.03.006": "FINANCIAMENTOS VEICULARES",
}


def _dre_categoria(codigo):
    """Categoria de DRE de uma natureza; None quando não vinculada no ERP."""
    return _DRE_CATEGORIA_POR_NATUREZA.get(str(codigo or "").strip())


# Grupos de unidade para visualização do DRE.
# Ativuz e AZ são analisadas sempre juntas; investidores ficam separados.
_DRE_GRUPO_PADRAO = "ativuz_az"
_DRE_GRUPOS = {
    "ativuz_az":  "Ativuz + AZ",
    "joao_paulo": "João Paulo Consórcios",
    "luz_divina": "Luz Divina",
}


def _dre_classificar_grupo(unidade, conta):
    """Grupo de visualização a partir da Unidade declarada; cai na Conta quando vazia."""
    u = str(unidade or "").strip().upper()
    if "JOAO PAULO" in u or "JOÃO PAULO" in u:
        return "joao_paulo"
    if "LUZ DIVINA" in u:
        return "luz_divina"
    if u:                                  # ATIVUZ VEÍCULOS / AZ EMPREENDIMENTOS
        return "ativuz_az"
    # Unidade em branco (preenchimento ainda em andamento no ERP): as contas de
    # frota de investidores são exclusivas deles, o resto é Ativuz+AZ.
    c = str(conta or "").strip().upper()
    if "JOAO PAULO" in c or "JOÃO PAULO" in c:
        return "joao_paulo"
    if "LUZ DIVINA" in c:
        return "luz_divina"
    return "ativuz_az"


def _dre_ler_lancamentos(filtro_tipo=None, grupo=_DRE_GRUPO_PADRAO):
    """
    Lê arquivos '*.xlsx' de lançamentos por natureza da pasta planilhas/dre/.

    Formato (exportação do Blue Fleet, relatório "Lançamentos por natureza"):
      row 2: filtro — contém "PAGAMENTO" ou "REFERÊNCIA"/"VENCIMENTO"
      row 5: cabeçalho nomeado
      row 6+: dados

    As colunas são localizadas pelo NOME do cabeçalho, não pela posição: o ERP
    exporta layouts diferentes (7 colunas no relatório compacto, 41 no completo)
    e ambos precisam ser lidos.

    filtro_tipo: "pagamento" | "referencia" | None (todos)
    grupo:       chave de _DRE_GRUPOS | None (todos)
    """
    import openpyxl

    pasta = Path(__file__).resolve().parent / "planilhas" / "dre"
    # macOS usa NFD nos nomes de arquivo; "natureza" não tem acento e funciona em ambos
    arquivos = [p for p in sorted(pasta.glob("*.xlsx"))
                if "natureza" in p.name.lower()]
    if not arquivos:
        return []

    def _parse_natureza(raw):
        s = str(raw or "").strip()
        if " - " in s:
            cod, label = s.split(" - ", 1)
            return cod.strip(), label.strip()
        return s, s

    def _detectar_tipo(rows):
        filtro = str(rows[1][0] or "").upper() if len(rows) > 1 else ""
        # "REFERÊNCIA" e "VENCIMENTO" são verificados antes porque o filtro
        # de referência pode conter "PAGAMENTO" na descrição
        if "REFER" in filtro or "VENCIMENTO" in filtro:
            return "referencia"
        if "PAGAMENTO" in filtro:
            return "pagamento"
        return "desconhecido"

    def _achar_cabecalho(rows):
        """Índice da linha de cabeçalho (a que contém 'Tipo de Lançamento')."""
        for i, row in enumerate(rows[:15]):
            nomes = [str(c or "").strip().lower() for c in row]
            if "tipo de lançamento" in nomes or "tipo de lancamento" in nomes:
                return i, {n: j for j, n in enumerate(nomes) if n}
        return None, {}

    def _parse_data(raw):
        if isinstance(raw, datetime):
            return raw
        if isinstance(raw, str) and raw.strip():
            try:
                return datetime.fromisoformat(raw.strip()[:10])
            except Exception:
                return None
        return None

    def _ler_arquivo(path):
        registros = []
        try:
            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()

            tipo_arquivo = _detectar_tipo(rows)
            i_hdr, cols = _achar_cabecalho(rows)
            if i_hdr is None:
                return registros

            def col(row, *nomes, default=None):
                for n in nomes:
                    j = cols.get(n)
                    if j is not None and j < len(row):
                        return row[j]
                return default

            # A data relevante depende do tipo de data usado na exportação.
            if tipo_arquivo == "pagamento":
                campos_data = ("data de pagamento ou recebimento", "pago em",
                               "data de vencimento")
            else:
                campos_data = ("data de vencimento", "data de competência",
                               "data prevista")

            for row in rows[i_hdr + 1:]:
                tipo = str(col(row, "tipo de lançamento", "tipo de lancamento") or "").strip().upper()
                if tipo not in ("ENTRADA", "SAÍDA", "SAIDA"):
                    continue

                dt = _parse_data(col(row, *campos_data))
                if dt is None:
                    continue

                descricao = str(col(row, "descrição", "descricao") or "").strip()
                nat_raw   = col(row, "natureza")
                val_raw   = col(row, "valor alocado na natureza", "valor")

                def _num(x):
                    try:
                        return float(str(x).replace(",", ".").replace(" ", ""))
                    except Exception:
                        return None

                valor = _num(val_raw)
                if valor is None:
                    continue
                if valor == 0:
                    # Desconto concedido: o ERP zera o valor alocado e registra o
                    # montante em "Descontos", rateado pelo percentual da natureza.
                    desc = _num(col(row, "descontos")) or 0.0
                    pct  = _num(col(row, "percentual alocado na natureza"))
                    if desc and pct:
                        valor = desc * pct
                if valor == 0:
                    continue

                cod, nat_label = _parse_natureza(nat_raw)

                unidade = str(col(row, "unidade") or "").strip()
                conta   = str(col(row, "conta") or "").strip()
                num     = col(row, "número do lançamento", "numero do lançamento",
                              "numero do lancamento")
                cliente = str(col(row, "pagar para ou receber de") or "").strip()

                registros.append({
                    "tipo_arquivo": tipo_arquivo,
                    "tipo":         tipo,
                    "descricao":    descricao,
                    "dt":           dt,
                    "codigo":       cod,
                    "natureza":     nat_label,
                    "valor":        abs(valor) * (-1 if tipo in ("SAÍDA", "SAIDA") else 1),
                    "unidade":      unidade,
                    "conta":        conta,
                    "grupo":        _dre_classificar_grupo(unidade, conta),
                    "num":          str(num or "").strip(),
                    "cliente":      cliente,
                })
        except Exception:
            import traceback; traceback.print_exc()
        return registros

    vistos = set()
    result = []
    for arq in sorted(arquivos, key=lambda p: p.stat().st_mtime):
        for reg in _ler_arquivo(arq):
            if filtro_tipo and reg["tipo_arquivo"] != filtro_tipo:
                continue
            if grupo and reg["grupo"] != grupo:
                continue
            # Um mesmo lançamento pode ser rateado em várias naturezas, e lançamentos
            # distintos podem ter descrição/data/natureza idênticas (folha, benefícios).
            # Com o número do lançamento a chave é exata; sem ele, cai no modo antigo.
            if reg["num"]:
                chave = (reg["tipo_arquivo"], reg["num"], reg["codigo"], reg["valor"])
            else:
                chave = (reg["tipo_arquivo"], reg["descricao"], reg["dt"].date(), reg["codigo"])
            if chave in vistos:
                continue
            vistos.add(chave)
            result.append(reg)
    return result


# Regras: (keywords, codigo_sugerido_saida, codigo_sugerido_entrada)
# None = mesmo código para os dois tipos
_NATUREZA_KEYWORDS = [
    (["gasolina","etanol","combustivel","combustível","abastec"],
     "02.02.04.005 - GASOLINA",              None),
    (["manutenção preventiva","manutencao preventiva"],
     "02.02.05.001 - MANUTENÇÃO PREVENTIVA",  None),
    (["manutenção corretiva","manutencao corretiva"],
     "02.02.05.002 - MANUTENÇÃO CORRETIVA",   None),
    (["compra de peç","compra de pneu"],
     "02.02.05.009 - COMPRA DE PEÇAS",        None),
    (["lavagem"],
     "02.02.05.008 - LAVAGEM VEICULAR NÃO REEMBOLSÁVEL", None),
    (["seguro total","seguradora","apólice","apolice"],
     "02.02.03.001 - SEGURO TOTAL",           None),
    (["guincho"],
     "02.02.03.004 - GUINCHO",                None),
    (["rastreador"],
     "02.02.03.005 - RASTREADOR VEICULAR",    None),
    (["saída de multa","saida de multa","multa de trânsito","multa de transito"],
     "02.02.06.001 - SAÍDA DE MULTA DE TRÂNSITO",
     "01.01.02.002 - ENTRADA DE MULTA DE TRANSITO"),
    (["ipva"],
     "02.02.01.003 - IPVA",                   None),
    (["licenciamento","crlv"],
     "02.02.01.005 - TAXA DE LICENCIAMENTO",  None),
    (["emplacamento"],
     "02.02.01.001 - EMPLACAMENTO",           None),
    (["transferência veicular","transferencia veicular"],
     "02.02.01.002 - TRANSFERÊNCIA VEICULAR", None),
    (["salário quinzenal","salario quinzenal","adiantamento salarial","adto salarial"],
     "02.03.01.001 - SALARIO",                None),
    (["férias","ferias"],
     "02.03.01.003 - FÉRIAS",                 None),
    (["13° salario","13 salario","décimo terceiro"],
     "02.03.01.004 - 13° SALARIO",            None),
    (["rescisão","rescisao"],
     "02.03.01.005 - RESCISAO",               None),
    (["comissão","comissao"],
     "02.03.01.007 - COMISSÃO",               None),
    (["vale transporte"," vt "],
     "02.03.02.001 - VT",                     None),
    (["vale refeição","vale refeicao"," vr "],
     "02.03.02.003 - VR",                     None),
    (["assistência médica","assistencia medica","plano de saúde"],
     "02.03.02.005 - ASSISTENCIA MEDICA",     None),
    (["inss"],
     "02.03.03.001 - INSS",                   None),
    (["fgts"],
     "02.03.03.002 - FGTS",                   None),
    (["distribuição de lucro","distribuicao de lucro"],
     "02.03.04.002 - DISTRIBUIÇÃO DE LUCROS MENSAL", None),
    (["simples nacional"],
     "02.01.01.005 - SIMPLES NACIONAL",       None),
    (["pis ","cofins"],
     "02.01.01.001 - PIS",                    None),
    (["aluguel de imóvel","aluguel de imovel","aluguel do imovel"],
     "02.04.02.001 - ALUGUEL DE IMOVEIS",     None),
    (["marketing","publicidade","propaganda"],
     "02.04.01.001 - MARKETING",              None),
    (["tarifa bancária","tarifa bancaria","taxa bancária"],
     "02.04.06.001 - TARIFA BANCARIA",        None),
    (["juros bancário","juros bancario","multa bancária"],
     "02.04.06.003 - JUROS E MULTAS BANCÁRIAS PAGOS", None),
    (["honorários advocatícios","honorarios advocaticios"],
     "02.04.05.001 - HONORÁRIOS ADVOCATÍCIOS", None),
    (["serviços contábeis","servicos contabeis","contabilidade"],
     "02.04.05.007 - SERVIÇOS CONTABEIS",     None),
    (["softwares","assinatura","sistema"],
     "02.04.05.003 - SOFTWARES",              None),
    (["financiamento parcela","pagamento de financiamento","parcela financiamento"],
     "03.01.03.003 - PAGAMENTO DE FINANCIAMENTO", None),
    # Consórcio contemplado — deve vir ANTES da regra genérica
    (["consorcio contemplado","consórcio contemplado"],
     "03.01.03.001 - CONSORCIO CONTEMPLADO", None),
    # Consórcio genérico — só sinaliza se a natureza atual não for nenhum código de consórcio
    (["consórcio","consorcio"],
     "03.01.03.005 - CONSÓRCIO PARCELA NÃO CONTEMPLADA", None),
    (["aporte de capital"],
     "04.04.01.003 - APORTE DE CAPITAL",      None),
    (["reembolso"],
     "04.04.02.02 - SAÍDA DE REEMBOLSO",
     "04.04.02.01 - ENTRADA DE REEMBOLSO"),
]


def _dre_correcoes_aceitas():
    """Carrega correções aceitas do Supabase: {(codigo_atual, descricao): (codigo_novo, natureza_nova)}"""
    try:
        sb = _supabase()
        if not sb:
            return {}
        res = sb.table("dre_correcoes").select("codigo_atual,descricao,codigo_novo,natureza_nova").execute()
        return {(r["codigo_atual"], r["descricao"]): (r["codigo_novo"], r["natureza_nova"])
                for r in (res.data or [])}
    except Exception:
        return {}


def _dre_ajustes_natureza(lancamentos):
    """
    Verifica se a descrição sugere uma natureza diferente da atribuída.
    Exclui lançamentos já corrigidos via aceite. Sugere código exato.
    """
    aceitas = _dre_correcoes_aceitas()
    suspeitos = []
    vistos_desc = set()

    for l in lancamentos:
        chave_corr = (l["codigo"], l["descricao"])
        if chave_corr in aceitas:
            continue   # já foi corrigido

        desc_lower = l["descricao"].lower()
        tipo_e = l["tipo"] in ("SAÍDA", "SAIDA")

        for keywords, cod_sug_saida, cod_sug_entrada in _NATUREZA_KEYWORDS:
            if any(kw in desc_lower for kw in keywords):
                cod_sug = cod_sug_saida if tipo_e else (cod_sug_entrada or cod_sug_saida)
                if cod_sug is None:
                    break
                # extrai só o código (antes do " - ")
                prefixo_sug = cod_sug.split(" - ")[0]
                if l["codigo"] == prefixo_sug:
                    break  # já está correto

                # Códigos irmãos válidos — mesma subcategoria, não sinalizar
                def _prefixo(c, n=3):
                    return ".".join(c.split(".")[:n])
                if _prefixo(l["codigo"]) == _prefixo(prefixo_sug):
                    break  # mesmo grupo (ex: 03.01.03.001 vs 03.01.03.005)

                # deduplica por (descricao, codigo_atual, sugestão)
                chave_dup = (l["descricao"], l["codigo"], prefixo_sug)
                if chave_dup in vistos_desc:
                    break
                vistos_desc.add(chave_dup)

                suspeitos.append({
                    "descricao":      l["descricao"],
                    "codigo_atual":   l["codigo"],
                    "natureza_atual": f"{l['codigo']} - {l['natureza']}",
                    "codigo_novo":    prefixo_sug,
                    "natureza_nova":  cod_sug,
                    "dt":             l["dt"].strftime("%d/%m/%Y"),
                    "valor_s":        _brl(abs(l["valor"])),
                    "tipo":           l["tipo"],
                })
                break
    return suspeitos


# ─── Plano de exibição do DRE (hierarquia visual) ────────────────────────────
# Tipos de bloco:
#   "section"  → cabeçalho de seção
#   "group"    → grupo expansível com lista de contas
#   "account"  → conta avulsa (sem grupo pai)
#   "subtotal" → linha calculada; "buckets" = lista de chaves a somar
#   "info"     → linha de texto sem valor numérico
#   "pct"      → linha percentual; valor = sum(buckets) / b[ref] * 100
_DRE_DISPLAY_PLAN = [
    # ── RECEITA OPERACIONAL ───────────────────────────────────────────
    {"t":"section", "label":"(+) RECEITA OPERACIONAL", "bucket":"rob"},
    {"t":"group", "bucket":"rob", "cod":"01.01.01", "label":"LOCAÇÃO", "contas":[
        ("01.01.01.001","LOCAÇÃO"),("01.01.01.002","KM EXCEDENTE"),
        ("01.01.01.003","MULTA ATRASO PAGAMENTO"),("01.01.01.004","MULTA QUEBRA DE CONTRATO"),
        ("01.01.01.005","ACORDO/RENEGOCIAÇÃO"),("01.01.01.006","TAXA DE ADMINISTRAÇÃO DE VEÍCULOS"),
    ]},
    {"t":"group", "bucket":"rob", "cod":"01.01.02", "label":"REEMBOLSOS", "contas":[
        ("01.01.02.001","MANUTENÇÃO REEMBOLSÁVEL"),("01.01.02.002","ENTRADA DE MULTA DE TRÂNSITO"),
        ("01.01.02.004","MULTA DEV. ANTECIPADA"),("01.01.02.005","REEMBOLSO DE SINISTRO"),
        ("01.01.02.006","OUTROS REEMBOLSOS"),
        ("01.01.02.009","REEMBOLSO DE MANUTENÇÕES — FROTA DE INVESTIDORES"),
        ("01.01.02.010","REEMBOLSO DE MULTAS — FROTA DE INVESTIDORES"),
        ("01.01.02.011","REEMBOLSO DE OUTRAS DESPESAS — FROTA DE INVESTIDORES"),
    ]},
    {"t":"subtotal","label":"RECEITA OPERACIONAL BRUTA","buckets":["rob"]},

    # ── DEDUÇÕES ──────────────────────────────────────────────────────
    {"t":"section","label":"(-) DEDUÇÕES","bucket":"ded"},
    {"t":"group","bucket":"ded","cod":"02.01.01","label":"IMPOSTOS","contas":[
        ("02.01.01.001","PIS"),("02.01.01.005","SIMPLES NACIONAL"),("02.01.01.006","OUTROS IMPOSTOS"),
    ]},
    {"t":"account","bucket":"ded","cod":"01.01.02.008","label":"(-) DESCONTO CONCEDIDO A CLIENTES","invert":True},
    {"t":"subtotal","label":"RECEITA OPERACIONAL LÍQUIDA","buckets":["rob","ded"]},

    # ── CUSTOS ────────────────────────────────────────────────────────
    {"t":"section","label":"(-) CUSTOS (CUSTO DIRETO DA OPERAÇÃO/FROTA)","bucket":"cst"},
    {"t":"group","bucket":"cst","cod":"02.02.01","label":"LICENCIAMENTO","contas":[
        ("02.02.01.001","EMPLACAMENTO"),("02.02.01.002","TRANSFERÊNCIA VEICULAR"),
        ("02.02.01.003","IPVA"),("02.02.01.005","TAXA DE LICENCIAMENTO"),
        ("02.02.01.006","DESPESAS COM CARTÓRIO"),("02.02.01.008","TAXA BOMBEIROS"),
        ("02.02.01.009","VISTORIA"),
    ]},
    {"t":"group","bucket":"cst","cod":"02.02.02","label":"HONORÁRIOS DESPACHANTE","contas":[
        ("02.02.02.001","HONORÁRIOS DESPACHANTE"),
    ]},
    {"t":"group","bucket":"cst","cod":"02.02.03","label":"SEGURO/ASSISTÊNCIA 24H","contas":[
        ("02.02.03.001","SEGURO TOTAL"),("02.02.03.003","RESERVA OPERACIONAL PARA SINISTROS"),
        ("02.02.03.004","GUINCHO"),("02.02.03.005","RASTREADOR VEICULAR"),
    ]},
    {"t":"group","bucket":"cst","cod":"02.02.04","label":"SUBLOCAÇÃO/TRANSPORTE","contas":[
        ("02.02.04.003","MULTAS NÃO REEMBOLSÁVEIS"),("02.02.04.004","TAXA DE FRETE"),
        ("02.02.04.005","GASOLINA"),("02.02.04.006","UBER OU APP DE TRANSPORTE"),
    ]},
    {"t":"group","bucket":"cst","cod":"02.02.05","label":"MANUTENÇÃO","contas":[
        ("02.02.05.001","MANUTENÇÃO PREVENTIVA"),("02.02.05.002","MANUTENÇÃO CORRETIVA"),
        ("02.02.05.006","COMPRA EQUIPAMENTO GNV"),("02.02.05.007","EQUIPAMENTOS (SENSOR, SIM, ETC)"),
        ("02.02.05.008","LAVAGEM VEICULAR NÃO REEMBOLSÁVEL"),("02.02.05.009","COMPRA DE PEÇAS"),
        ("02.02.05.010","COMPRA DE PNEUS"),
    ]},
    {"t":"group","bucket":"cst","cod":"02.02.06","label":"DESPESAS REEMBOLSÁVEIS","contas":[
        ("02.02.06.001","SAÍDA DE MULTA DE TRÂNSITO"),("02.02.06.002","SAÍDA DE SINISTRO"),
        ("02.02.06.006","COMBUSTÍVEL REEMBOLSÁVEL"),("02.02.06.010","REEMBOLSO DE CLIENTES"),
    ]},
    {"t":"group","bucket":"cst","cod":"02.02.07","label":"DESPESAS COM FROTA DE INVESTIDORES","contas":[
        ("02.02.07.01","MANUTENÇÕES — FROTA DE INVESTIDORES"),
        ("02.02.07.02","SINISTROS — FROTA DE INVESTIDORES"),
        ("02.02.07.03","DESPESAS OPERACIONAIS — FROTA DE INVESTIDORES"),
    ]},
    {"t":"subtotal","label":"MARGEM DE CONTRIBUIÇÃO","buckets":["rob","ded","cst"]},

    # ── SG&A ──────────────────────────────────────────────────────────
    {"t":"section","label":"(-) SG&A (DESPESAS GERAIS E ADMINISTRATIVAS)","bucket":"sga"},
    {"t":"group","bucket":"sga","cod":"02.03.01","label":"SALÁRIOS","contas":[
        ("02.03.01.001","SALÁRIO"),("02.03.01.002","ADTO SALARIAL"),
        ("02.03.01.003","FÉRIAS"),("02.03.01.004","13° SALÁRIO"),
        ("02.03.01.005","RESCISÃO"),("02.03.01.006","PRÊMIOS"),("02.03.01.007","COMISSÃO"),
    ]},
    {"t":"group","bucket":"sga","cod":"02.03.02","label":"BENEFÍCIOS","contas":[
        ("02.03.02.001","VT"),("02.03.02.003","VR"),("02.03.02.005","ASSISTÊNCIA MÉDICA"),
        ("02.03.02.006","P.C.M.S.O"),("02.03.02.007","TREINAMENTO"),("02.03.02.008","OUTROS BENEFÍCIOS"),
    ]},
    {"t":"group","bucket":"sga","cod":"02.03.03","label":"IMPOSTOS FOLHA","contas":[
        ("02.03.03.001","INSS"),("02.03.03.002","FGTS"),
    ]},
    {"t":"group","bucket":"sga","cod":"02.03.04","label":"PRÓ-LABORE","contas":[
        ("02.03.04.001","PRÓ-LABORE FOLHA"),("02.03.04.002","DISTRIBUIÇÃO DE LUCROS MENSAL"),
    ]},
    {"t":"group","bucket":"sga","cod":"02.04.01","label":"DESPESAS COMERCIAIS","contas":[
        ("02.04.01.001","MARKETING"),
    ]},
    {"t":"group","bucket":"sga","cod":"02.04.02","label":"OCUPAÇÃO","contas":[
        ("02.04.02.001","ALUGUEL DE IMÓVEIS"),("02.04.02.003","IPTU"),
        ("02.04.02.004","ÁGUA"),("02.04.02.005","LUZ"),("02.04.02.006","MANUTENÇÃO PREDIAL"),
    ]},
    {"t":"group","bucket":"sga","cod":"02.04.03","label":"SUPRIMENTOS","contas":[
        ("02.04.03.001","TELEFONE"),
    ]},
    {"t":"group","bucket":"sga","cod":"02.04.04","label":"DESPESAS ADMINISTRATIVAS","contas":[
        ("02.04.04.001","DESPESAS ADMINISTRATIVAS E DE ESCRITÓRIO"),
        ("02.04.04.004","TAXAS E DESPESAS LEGAIS"),
        ("02.04.04.005","OUTRAS DESPESAS"),("02.04.04.007","DESPESAS JURÍDICAS"),
    ]},
    {"t":"group","bucket":"sga","cod":"02.04.05","label":"SERVIÇOS PRESTADOS","contas":[
        ("02.04.05.001","HONORÁRIOS ADVOCATÍCIOS"),("02.04.05.003","SOFTWARES"),
        ("02.04.05.004","ÓRGÃOS DE PROTEÇÃO AO CRÉDITO"),("02.04.05.006","HONORÁRIOS DE CONSULTORIA"),
        ("02.04.05.007","SERVIÇOS CONTÁBEIS"),("02.04.05.008","SERVIÇOS DE LIMPEZA"),
        ("02.04.05.009","SERVIÇOS DE MANUTENÇÃO DE MÁQUINAS E EQUIPAMENTOS"),
    ]},
    {"t":"group","bucket":"sga","cod":"02.04.07","label":"OUTRAS SAÍDAS","contas":[
        ("02.04.07.002","OUTRAS SAÍDAS"),
    ]},
    {"t":"subtotal","label":"EBITDA","buckets":["rob","ded","cst","sga"]},
    {"t":"info",    "label":"DEPRECIAÇÃO"},
    {"t":"subtotal","label":"RESULTADO OPERACIONAL (EBIT)","buckets":["rob","ded","cst","sga"]},
    {"t":"pct",     "label":"MARGEM EBIT","buckets":["rob","ded","cst","sga"],"ref":"rob"},

    # ── RESULTADO FINANCEIRO ──────────────────────────────────────────
    {"t":"section","label":"(-) RESULTADO FINANCEIRO","bucket":"rfi"},
    {"t":"group","bucket":"rfi","cod":"02.04.06","label":"DESPESAS BANCÁRIAS","contas":[
        ("02.04.06.001","TARIFA BANCÁRIA"),("02.04.06.003","JUROS E MULTAS BANCÁRIAS PAGOS"),
        ("02.04.06.004","(-) DESCONTO PGTO BOLETOS"),("02.04.06.005","TAXA MAQUINETA"),
    ]},
    {"t":"group","bucket":"rfi","cod":"03.03.01_fin","label":"APLICAÇÃO FINANCEIRA","contas":[
        ("03.03.01.003","RENDIMENTO DE APLICAÇÕES"),
    ]},
    {"t":"group","bucket":"rfi","cod":"04.01.02_fin","label":"SAÍDAS DE MÚTUOS","contas":[
        ("04.01.02.002","PGTO JUROS SOBRE MÚTUOS"),
    ]},
    {"t":"group","bucket":"rfi","cod":"enc_fin","label":"ENCARGOS FINANCEIROS","contas":[
        ("03.01.03.002","CONSÓRCIO CONTEMPLADO JUROS"),("03.01.03.004","FINANCIAMENTO JUROS"),
    ]},

    # ── RESULTADOS NÃO OPERACIONAIS ────────────────────────────────────
    {"t":"section","label":"(-) RESULTADOS NÃO OPERACIONAIS","bucket":"rno"},
    {"t":"group","bucket":"rno","cod":"01.02.01","label":"OUTRAS ENTRADAS","contas":[
        ("01.02.01.002","DEPÓSITOS NÃO IDENTIFICADOS"),("01.02.01.003","OUTRAS ENTRADAS"),
        ("01.02.01.005","RECEBIMENTOS — FROTA INVESTIDORES"),
    ]},
    {"t":"group","bucket":"rno","cod":"03.03.02","label":"OUTROS INVESTIMENTOS","contas":[
        ("03.03.02.001","OUTROS INVESTIMENTOS"),
    ]},
    {"t":"subtotal","label":"LUCRO LÍQUIDO","buckets":["rob","ded","cst","sga","rfi","rno"]},

    # ── INVESTIMENTOS ─────────────────────────────────────────────────
    {"t":"section","label":"(-) INVESTIMENTOS","bucket":"inv"},
    {"t":"group","bucket":"inv","cod":"03.01.01","label":"VENDA DE VEÍCULOS","contas":[
        ("01.02.01.004","VENDA DE VEÍCULO"),
    ]},
    {"t":"group","bucket":"inv","cod":"03.01.02","label":"COMPRA DE VEÍCULOS","contas":[
        ("03.01.02.001","COMPRA DE VEÍCULOS À VISTA"),("03.01.02.002","ENTRADA COMPRA VEÍCULO"),
        ("03.01.02.003","ADIANTAMENTO DE CONSÓRCIO"),("03.01.03.005","CONSÓRCIO PARCELA NÃO CONTEMPLADA"),
    ]},
    {"t":"group","bucket":"inv","cod":"03.01.03_inv","label":"FINANCIAMENTOS VEICULARES","contas":[
        ("03.01.03.001","CONSÓRCIO CONTEMPLADO"),("03.01.03.003","PAGAMENTO DE FINANCIAMENTO"),
    ]},
    {"t":"group","bucket":"inv","cod":"03.02.01","label":"OUTROS IMOBILIZADOS","contas":[
        ("03.02.01.001","INSTALAÇÕES"),("03.02.01.002","COMPUTADORES E PERIFÉRICOS"),
        ("03.02.01.003","MÓVEIS E UTENSÍLIOS"),("03.02.01.004","SISTEMAS E SOFTWARES"),
        ("03.02.01.005","OUTRAS IMOBILIZAÇÕES"),
    ]},
    {"t":"group","bucket":"inv","cod":"02.04.08","label":"CONSTRUÇÃO DA OFICINA","contas":[
        ("02.04.08.01","COMPRA DE MATERIAL PARA OFICINA"),("02.04.08.02","PAGAMENTO DA MÃO DE OBRA"),
        ("02.04.08.03","ALUGUEL DE EQUIPAMENTOS"),
    ]},
    {"t":"group","bucket":"inv","cod":"03.03.01_inv","label":"APLICAÇÃO FINANCEIRA","contas":[
        ("03.03.01.001","APLICAÇÕES FINANCEIRAS"),("03.03.01.002","RESGATE DE APLICAÇÃO FINANCEIRA"),
    ]},

    # ── FINANCIAMENTOS ────────────────────────────────────────────────
    {"t":"section","label":"(-) FINANCIAMENTOS","bucket":"fin"},
    {"t":"group","bucket":"fin","cod":"04.01.01","label":"ENTRADAS DE MÚTUOS","contas":[
        ("04.01.01.001","ENTRADA DE MÚTUOS"),("04.01.01.002","ENTRADA CAUÇÃO"),
    ]},
    {"t":"group","bucket":"fin","cod":"04.01.02_fin2","label":"PGTO DE MÚTUOS","contas":[
        ("04.01.02.001","SAÍDA DE MÚTUOS"),("04.01.02.003","SAÍDA CAUÇÃO"),
    ]},
    {"t":"group","bucket":"fin","cod":"04.04.02","label":"REEMBOLSOS","contas":[
        ("04.04.02.01","ENTRADA DE REEMBOLSO"),("04.04.02.02","SAÍDA DE REEMBOLSO"),
    ]},

    # ── APORTE DE CAPITAL ─────────────────────────────────────────────
    {"t":"section","label":"(+) APORTE DE CAPITAL","bucket":"apt"},
    {"t":"account","bucket":"apt","cod":"04.04.01.003","label":"APORTE DE CAPITAL"},
    {"t":"subtotal","label":"FLUXO DE CAIXA ACIONISTA",
     "buckets":["rob","ded","cst","sga","rfi","rno","inv","fin","apt"]},

    # ── DISTRIBUIÇÃO ──────────────────────────────────────────────────
    {"t":"account","bucket":"dis","cod":"04.04.01.002","label":"DISTRIB RESULTADO"},
    {"t":"subtotal","label":"FLUXO DE CAIXA LIVRE",
     "buckets":["rob","ded","cst","sga","rfi","rno","inv","fin","apt","dis"]},
]


def _dre_montar_estrutura(lancamentos, ano, mes):
    """Retorna lista de 'rows' para renderizar o DRE na hierarquia do plano de contas."""
    from calendar import monthrange
    from collections import defaultdict

    d_ini = datetime(ano, mes, 1)
    d_fim = datetime(ano, mes, monthrange(ano, mes)[1], 23, 59, 59)

    # Aplica correções aceitas
    aceitas = _dre_correcoes_aceitas()
    corrigidos = []
    for l in lancamentos:
        chave = (l["codigo"], l["descricao"])
        if chave in aceitas:
            cod_novo, nat_nova = aceitas[chave]
            l = dict(l)
            l["codigo"] = cod_novo.split(" - ")[0].strip() if " - " in cod_novo else cod_novo
            l["natureza"] = nat_nova.split(" - ", 1)[1] if " - " in nat_nova else nat_nova
        corrigidos.append(l)

    filtrados = [l for l in corrigidos if d_ini <= l["dt"] <= d_fim]

    por_codigo = defaultdict(list)
    for l in filtrados:
        por_codigo[l["codigo"]].append(l)

    buckets = defaultdict(float)   # chave → soma líquida
    codigos_usados = set()

    def _somar_contas(contas_def, bucket_key):
        itens = []
        total = 0.0
        for cod, lbl in contas_def:
            ls = por_codigo.get(cod, [])
            codigos_usados.add(cod)
            val = sum(l["valor"] for l in ls)
            total += val
            itens.append({
                "codigo": cod, "label": lbl, "valor": val,
                "lancamentos": sorted(ls, key=lambda x: x["dt"]),
                "tem_dados": bool(ls),
            })
        buckets[bucket_key] += total
        return total, itens

    rows = []
    for bloco in _DRE_DISPLAY_PLAN:
        t = bloco["t"]

        if t == "section":
            rows.append({"tipo": "section", "label": bloco["label"]})

        elif t == "group":
            net, itens = _somar_contas(bloco["contas"], bloco["bucket"])
            rows.append({
                "tipo": "group", "cod": bloco["cod"], "label": bloco["label"],
                "net": net, "itens": itens,
            })

        elif t == "account":
            cod = bloco["cod"]
            ls  = por_codigo.get(cod, [])
            codigos_usados.add(cod)
            val = sum(l["valor"] for l in ls)
            if bloco.get("invert"):
                val = -val
            buckets[bloco["bucket"]] += val
            rows.append({
                "tipo": "account", "cod": cod, "label": bloco["label"],
                "valor": val,
                "lancamentos": sorted(ls, key=lambda x: x["dt"]),
            })

        elif t == "subtotal":
            valor = sum(buckets[k] for k in bloco["buckets"])
            rows.append({"tipo": "subtotal", "label": bloco["label"], "valor": valor})

        elif t == "info":
            rows.append({"tipo": "info", "label": bloco["label"]})

        elif t == "pct":
            ref = buckets.get(bloco.get("ref", "rob"), 0.0)
            num = sum(buckets[k] for k in bloco["buckets"])
            pct = num / ref * 100 if ref else 0.0
            rows.append({"tipo": "pct", "label": bloco["label"], "valor": pct})

    # Lançamentos não classificados em nenhuma conta do plano
    nao_class = {cod: ls for cod, ls in por_codigo.items()
                 if cod not in codigos_usados and any(l["valor"] != 0 for l in ls)}
    if nao_class:
        rows.append({"tipo": "section", "label": "⚠ NÃO CLASSIFICADOS"})
        for cod, ls in sorted(nao_class.items()):
            net = sum(l["valor"] for l in ls)
            rows.append({
                "tipo": "group", "cod": cod,
                "label": f"{cod} — {ls[0]['natureza'] or cod}",
                "net": net, "itens": [{
                    "codigo": cod, "label": ls[0]["natureza"] or cod, "valor": net,
                    "lancamentos": sorted(ls, key=lambda x: x["dt"]),
                }],
            })

    # Calcula % sobre ROB para cada linha com valor
    rob = next(
        (r["valor"] for r in rows
         if r["tipo"] == "subtotal" and "OPERACIONAL BRUTA" in r.get("label", "")),
        0.0
    )
    for r in rows:
        v = None
        if r["tipo"] == "subtotal":
            v = r.get("valor")
        elif r["tipo"] == "group":
            v = r.get("net")
        elif r["tipo"] in ("account",):
            v = r.get("valor")
        if v is not None and rob:
            r["pct_rob"] = v / rob * 100
        else:
            r["pct_rob"] = None

    return rows


def _dre_calcular(lancamentos):
    from collections import defaultdict
    code_val = defaultdict(float)
    for l in lancamentos:
        if l["codigo"]:
            code_val[l["codigo"]] += l["valor"]

    sections = []
    for sec_def in _DRE_LAYOUT:
        grupos = []
        sec_total = 0.0
        for grp_def in sec_def["grupos"]:
            sign = grp_def["sign"]
            itens = []
            grp_abs = 0.0
            for codigo, label in grp_def["itens"]:
                v = code_val.get(codigo, 0.0)
                grp_abs += v
                itens.append({"codigo": codigo, "label": label, "val": sign * v})
            grp_total = sign * grp_abs
            sec_total += grp_total
            grupos.append({
                "id": grp_def["id"], "label": grp_def["label"],
                "sign": sign, "total": grp_total, "itens": itens,
                "highlight": grp_def.get("highlight", False),
                "note": grp_def.get("note", ""),
            })
        sections.append({"id": sec_def["id"], "label": sec_def["label"],
                         "total": sec_total, "grupos": grupos})

    def _s(sid):
        for s in sections:
            if s["id"] == sid:
                return s["total"]
        return 0.0

    rb     = _s("rb")
    ded    = _s("ded")
    rl     = rb + ded
    custos = _s("custos")
    margem = rl + custos
    sga    = _s("sga")
    ebitda = margem + sga
    ebit   = ebitda        # depreciação = 0
    rfin   = _s("rfin")
    rnop   = _s("rnop")
    lucro_antes_ir = ebit + rfin + rnop
    ir_csll = _DRE_IR_CSLL
    ll     = lucro_antes_ir - ir_csll
    inv    = _s("inv")
    financ = _s("financ")
    aporte = _s("aporte")
    fluxo_ac  = ll + inv + financ + aporte
    distrib   = _s("distrib")
    fluxo_liv = fluxo_ac + distrib

    # Add %RL to each section (informational; template may display for L1 rows)
    for sec in sections:
        sec["pct"] = sec["total"] / rl if rl else 0.0

    return {
        "sections": sections,
        "receita_bruta": rb,  "deducoes": ded,
        "receita_liquida": rl,
        "custos": custos,
        "margem": margem,       "pct_margem": margem / rl if rl else 0,
        "sga": sga,
        "ebitda": ebitda,       "pct_ebitda": ebitda / rl if rl else 0,
        "depreciacao": 0.0,
        "ebit": ebit,           "pct_ebit": ebit / rl if rl else 0,
        "rfin": rfin,           "rnop": rnop,
        "lucro_antes_ir": lucro_antes_ir,
        "pct_lajir": lucro_antes_ir / rl if rl else 0,
        "ir_csll": ir_csll,
        "lucro_liquido": ll,    "pct_ll": ll / rl if rl else 0,
        "inv": inv,  "financ": financ,  "aporte": aporte,
        "fluxo_acionista": fluxo_ac,
        "distrib": distrib,
        "fluxo_livre": fluxo_liv,
    }


def _ler_lancamentos_jun_jul():
    """Lê o arquivo histórico de junho/julho 2025 (formato diferente do DRE principal)."""
    import openpyxl
    path = Path(__file__).resolve().parent / "planilhas" / "dados_junho_julho.xlsx"
    if not path.exists():
        return []
    registros = []
    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        for row in rows[1:]:  # pula cabeçalho
            if not row[0]:
                continue
            try:
                dt = datetime.strptime(str(row[2]), "%d/%m/%Y")
                natureza = str(row[3]) if row[3] else ""
                cod = natureza.split(" - ")[0].strip() if " - " in natureza else natureza.strip()
                valor_str = str(row[4]).replace("R$", "").replace("-", "").replace(".", "").replace(",", ".").strip()
                valor = float(valor_str)
                registros.append({"codigo": cod, "dt": dt, "valor": valor})
            except Exception:
                continue
    except Exception:
        pass
    return registros


def _calcular_indicadores_ativuz():
    """Calcula margens da Ativuz usando os últimos 12 meses de lançamentos."""
    from calendar import monthrange
    hoje = datetime.now(_BRT)
    mes_ini = hoje.month + 1
    ano_ini = hoje.year - 1
    if mes_ini > 12:
        mes_ini -= 12
        ano_ini += 1
    d_ini = datetime(ano_ini, mes_ini, 1)
    d_fim = datetime(hoje.year, hoje.month, monthrange(hoje.year, hoje.month)[1], 23, 59, 59)

    todos = _dre_ler_lancamentos() + _ler_lancamentos_jun_jul()
    filtrados = [l for l in todos if d_ini <= l["dt"] <= d_fim]
    if not filtrados:
        return None

    dre = _dre_calcular(filtrados)

    # Saldo devedor da carteira financeira (dívida líquida)
    saldo_devedor = 0.0
    try:
        sb = _supabase()
        if sb:
            rows = sb.table("financiamentos_contratos").select("*").execute().data or []
            hoje_d = hoje.date()
            for r in rows:
                try:
                    restante = _fin_calcular_restante(r, hoje_d)
                    saldo_devedor += restante * float(r["valor_parcela"])
                except Exception:
                    continue
    except Exception:
        pass

    # Depreciação anual da frota (vida útil 5 anos — Receita Federal veículos leves)
    depreciacao_anual = 0.0
    try:
        sb2 = _supabase()
        if sb2:
            res_frota = sb2.table("frota_veiculos").select("vl_aquisicao").eq("ativo", True).execute()
            for v in (res_frota.data or []):
                try:
                    depreciacao_anual += float(v.get("vl_aquisicao") or 0) / 5
                except (TypeError, ValueError):
                    continue
    except Exception:
        pass

    def _pct(v):
        return f"{v * 100:.2f}%".replace(".", ",")

    def _ratio(numerador, denominador):
        if not denominador:
            return "N/D"
        return f"{numerador / denominador:.2f}".replace(".", ",")

    rl     = dre["receita_liquida"]
    ebitda = dre["ebitda"]
    ebit   = ebitda - depreciacao_anual
    pct_ebit = (ebit / rl) if rl else 0

    return {
        "ticker": "ATIVUZ", "nome": "Ativuz", "erro": None, "is_ativuz": True,
        "pl":             "N/A",
        "pvp":            "N/A",
        "roe":            "N/D",
        "margem_bruta":   _pct(dre["pct_margem"]),
        "margem_ebitda":  _pct(dre["pct_ebitda"]),
        "margem_ebit":    _pct(pct_ebit),
        "margem_liquida": _pct(dre["pct_ll"]),
        "div_ebitda":     _ratio(saldo_devedor, ebitda),
        "div_ebit":       _ratio(saldo_devedor, ebit),
    }


# DRE temporariamente oculto (até segunda ordem).
# Para reativar: defina DRE_HABILITADO=1 no ambiente.
DRE_HABILITADO = _os.environ.get("DRE_HABILITADO", "0") == "1"


@app.context_processor
def _injetar_dre_habilitado():
    return {"dre_habilitado": DRE_HABILITADO}


@app.route("/dre")
def pagina_dre():
    if not DRE_HABILITADO:
        abort(404)
    hoje = datetime.now(_BRT)
    aba  = request.args.get("aba", "pagamento")  # "pagamento" | "referencia" | "ajustes"

    try:
        mes = int(request.args.get("mes", hoje.month))
        ano = int(request.args.get("ano", hoje.year))
        if not (1 <= mes <= 12):
            mes = hoje.month
    except (ValueError, TypeError):
        mes, ano = hoje.month, hoje.year

    pag   = _dre_ler_lancamentos("pagamento")
    ref   = _dre_ler_lancamentos("referencia")
    todos = pag + ref

    meses_pag = sorted({(l["dt"].year, l["dt"].month) for l in pag})
    meses_ref = sorted({(l["dt"].year, l["dt"].month) for l in ref})
    meses_disponiveis = sorted(set(meses_pag) | set(meses_ref))
    if not meses_disponiveis:
        meses_disponiveis = [(ano, mes)]

    rows_pag = _dre_montar_estrutura(pag, ano, mes)
    rows_ref = _dre_montar_estrutura(ref, ano, mes)

    # Mês anterior (automático)
    mes_ant = mes - 1 if mes > 1 else 12
    ano_ant = ano  if mes > 1 else ano - 1
    rows_pag_ant = _dre_montar_estrutura(pag, ano_ant, mes_ant)
    rows_ref_ant = _dre_montar_estrutura(ref, ano_ant, mes_ant)
    _MESES_PT = ["Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"]
    label_ant = f"{_MESES_PT[mes_ant-1]}/{ano_ant}"
    label_cur = f"{_MESES_PT[mes-1]}/{ano}"

    # Ajustes usam todos os lançamentos (sem filtro de período)
    ajustes = _dre_ajustes_natureza(todos)

    return render_template("dre.html",
        active="dre",
        aba=aba,
        mes=mes, ano=ano,
        meses_disponiveis=meses_disponiveis,
        rows_pag=rows_pag,
        rows_ref=rows_ref,
        rows_pag_ant=rows_pag_ant,
        rows_ref_ant=rows_ref_ant,
        label_ant=label_ant,
        label_cur=label_cur,
        ajustes=ajustes,
        sem_dados=len(todos) == 0,
    )


@app.route("/dre/api/aceitar-ajuste", methods=["POST"])
def dre_aceitar_ajuste():
    if not DRE_HABILITADO:
        abort(404)
    dados = request.get_json(force=True, silent=True) or {}
    codigo_atual  = str(dados.get("codigo_atual", "")).strip()
    descricao     = str(dados.get("descricao", "")).strip()
    codigo_novo   = str(dados.get("codigo_novo", "")).strip()
    natureza_nova = str(dados.get("natureza_nova", "")).strip()

    if not all([codigo_atual, descricao, codigo_novo]):
        return jsonify({"ok": False, "erro": "Dados incompletos"}), 400

    sb = _supabase()
    if not sb:
        return jsonify({"ok": False, "erro": "Supabase indisponível"}), 500

    try:
        existing = (sb.table("dre_correcoes")
                      .select("id")
                      .eq("codigo_atual", codigo_atual)
                      .eq("descricao", descricao)
                      .execute())
        payload = {"codigo_atual": codigo_atual, "descricao": descricao,
                   "codigo_novo": codigo_novo, "natureza_nova": natureza_nova}
        if existing.data:
            sb.table("dre_correcoes").update(payload).eq("id", existing.data[0]["id"]).execute()
        else:
            sb.table("dre_correcoes").insert(payload).execute()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)}), 500


@app.route("/dre/api/recalcular", methods=["POST"])
def dre_api_recalcular():
    if not DRE_HABILITADO:
        abort(404)
    from calendar import monthrange
    try:
        dados = request.get_json(force=True, silent=True) or {}
        lancamentos_raw = dados.get("lancamentos", [])
        mes = int(dados.get("mes", 1))
        ano = int(dados.get("ano", datetime.now(_BRT).year))
        acumulado = bool(dados.get("acumulado", False))
        if not (1 <= mes <= 12):
            return jsonify({"ok": False, "erro": "Mês inválido"}), 400

        # Parse novos lançamentos do arquivo importado
        novos = []
        for item in lancamentos_raw:
            try:
                dt = datetime.strptime(str(item.get("dt", "")), "%Y-%m-%d")
                valor = float(item.get("valor", 0))
                cod = str(item.get("codigo", "")).strip()
                num = str(item.get("num", "")).strip()
                if cod:
                    novos.append({"codigo": cod, "dt": dt, "valor": valor, "num": num})
            except Exception:
                continue

        # Mescla: novos têm prioridade; existentes preenchem o restante
        vistos = {(l["num"], l["codigo"]) for l in novos if l["num"]}
        existentes = _dre_ler_lancamentos() + _ler_lancamentos_jun_jul()
        for l in existentes:
            chave = (l.get("num", ""), l["codigo"])
            if chave[0] and chave in vistos:
                continue  # duplicata já presente no arquivo novo
            vistos.add(chave)
            novos.append(l)

        d_ini = datetime(ano, 1, 1) if acumulado else datetime(ano, mes, 1)
        d_fim = datetime(ano, mes, monthrange(ano, mes)[1], 23, 59, 59)
        if acumulado:
            d_ini_prev = datetime(ano - 1, 1, 1)
            d_fim_prev = datetime(ano - 1, mes, monthrange(ano - 1, mes)[1], 23, 59, 59)
        else:
            pm = mes - 1 if mes > 1 else 12
            py = ano if mes > 1 else ano - 1
            d_ini_prev = datetime(py, pm, 1)
            d_fim_prev = datetime(py, pm, monthrange(py, pm)[1], 23, 59, 59)

        def _f(d0, d1):
            return [l for l in novos if d0 <= l["dt"] <= d1]

        return jsonify({
            "ok": True,
            "dre":      _dre_calcular(_f(d_ini, d_fim)),
            "dre_prev": _dre_calcular(_f(d_ini_prev, d_fim_prev)),
        })
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)}), 500


# ── Financiamentos & Consórcios ───────────────────────────────────────────────

def _fin_calcular_restante(r, hoje):
    """
    Parcelas restantes de um contrato de financiamento.

    Assume parcelas mensais regulares terminando em data_vencimento. Se o
    contrato tiver data_parcela_1 / data_parcela_2 preenchidas (parcelas
    1 e 2 pagas fora do calendário mensal normal, ex.: entrada + carência),
    a contagem regular só é usada a partir da parcela 3 em diante.
    """
    from math import ceil
    parcelas_total = int(r["parcelas_total"])
    vcto_str = r.get("data_vencimento")
    if not vcto_str:
        return 0
    try:
        vcto = date.fromisoformat(str(vcto_str)[:10])
    except Exception:
        return 0

    dias = (vcto - hoje).days
    restante_regular = ceil(dias / 30.44) if dias > 0 else 0
    pagas_regular = max(0, parcelas_total - restante_regular)

    p1_str = r.get("data_parcela_1")
    p2_str = r.get("data_parcela_2")
    if p1_str and p2_str:
        try:
            p1 = date.fromisoformat(str(p1_str)[:10])
            p2 = date.fromisoformat(str(p2_str)[:10])
            if hoje < p1:
                pagas = 0
            elif hoje < p2:
                pagas = 1
            else:
                pagas = max(2, pagas_regular)
            return parcelas_total - min(pagas, parcelas_total)
        except Exception:
            pass

    return min(restante_regular, parcelas_total)


@app.route("/financiamentos")
def pagina_financiamentos():
    hoje = datetime.now(_BRT).date()

    sb   = _supabase()
    rows = sb.table("financiamentos_contratos").select("*").order("created_at").execute().data or []

    contratos = []
    vendidos  = []
    for r in rows:
        parcelas = int(r["parcelas_total"])
        parcela  = float(r["valor_parcela"])
        entrada  = float(r.get("valor_entrada") or 0)
        restante = _fin_calcular_restante(r, hoje)
        pagas    = parcelas - restante

        item = {
            "id":              r["id"],
            "operacao":        r["operacao"],
            "contrato":        r.get("contrato") or "",
            "placa":           r.get("placa") or "",
            "data_vencimento": str(r.get("data_vencimento") or "")[:10] or None,
            "restante":        restante,
            "parcelas_total":  parcelas,
            "valor_parcela":   parcela,
            "total_pago":      pagas * parcela + entrada,
            "total":           parcelas * parcela + entrada,
            "devedor":         restante * parcela,
            "c_prazo":         min(restante, 12) * parcela,
            "l_prazo":         max(restante - 12, 0) * parcela,
            "pct_quitado":     pagas / parcelas if parcelas else 0,
            "quitado":         restante == 0,
            "tipo":            r.get("tipo") or "financiamento",
            "valor_resgate":   float(r["valor_resgate"]) if r.get("valor_resgate") is not None else None,
            "data_resgate":    str(r.get("data_resgate") or "")[:10] or None,
        }

        if r.get("vendido"):
            vendidos.append(item)
        else:
            contratos.append(item)

    contratos.sort(key=lambda x: x["pct_quitado"], reverse=True)
    vendidos.sort(key=lambda x: x["placa"])

    ativos   = [c for c in contratos if not c["quitado"]]
    quitados = [c for c in contratos if     c["quitado"]]

    soma_devedor    = sum(c["devedor"]       for c in ativos)
    soma_total_pago = sum(c["total_pago"]    for c in contratos)
    soma_c_prazo    = sum(c["c_prazo"]       for c in ativos)
    soma_l_prazo    = sum(c["l_prazo"]       for c in ativos)
    soma_mensal     = sum(c["valor_parcela"] for c in ativos)
    tempo_medio     = sum(c["restante"]      for c in ativos) / len(ativos) if ativos else 0

    ativos_vcto = [c for c in ativos if c["data_vencimento"]]
    mais_perto  = min(ativos_vcto, key=lambda x: x["data_vencimento"])["operacao"] if ativos_vcto else "—"

    cards = {
        "saldo_devedor": soma_devedor,
        "total_pago":    soma_total_pago,
        "curto_prazo":   soma_c_prazo,
        "longo_prazo":   soma_l_prazo,
        "valor_mensal":  soma_mensal,
        "tempo_medio":   round(tempo_medio, 1),
        "mais_perto":    mais_perto,
        "r_quitados":    sum(c["total_pago"] for c in quitados),
        "pct_cp":        soma_c_prazo / soma_devedor if soma_devedor else 0,
        "pct_lp":        soma_l_prazo / soma_devedor if soma_devedor else 0,
    }

    return render_template("financiamentos.html",
        active="financiamentos",
        contratos=contratos,
        vendidos=vendidos,
        cards=cards,
    )


@app.route("/api/financiamentos/<uuid:contrato_id>/vendido", methods=["PUT"])
def api_financiamentos_marcar_vendido(contrato_id):
    sb = _supabase()
    if not sb:
        return jsonify({"ok": False, "erro": "Sem conexão com o banco."}), 500
    vendido = bool((request.get_json(silent=True) or {}).get("vendido", True))
    try:
        sb.table("financiamentos_contratos").update(
            {"vendido": vendido}
        ).eq("id", str(contrato_id)).execute()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)}), 500


@app.route("/api/financiamentos/<uuid:contrato_id>/resgate", methods=["PUT"])
def api_financiamentos_registrar_resgate(contrato_id):
    """Registra (ou limpa) o valor e a data do resgate de um consórcio —
    o valor que fica pendente pra sacar quando o grupo/cota é encerrado."""
    sb = _supabase()
    if not sb:
        return jsonify({"ok": False, "erro": "Sem conexão com o banco."}), 500
    body = request.get_json(silent=True) or {}
    valor_raw = body.get("valor_resgate")
    data_raw  = (body.get("data_resgate") or "").strip()
    try:
        valor = float(valor_raw) if valor_raw not in (None, "") else None
    except (TypeError, ValueError):
        return jsonify({"ok": False, "erro": "Valor de resgate inválido."}), 400
    try:
        sb.table("financiamentos_contratos").update({
            "valor_resgate": valor,
            "data_resgate":  data_raw or None,
        }).eq("id", str(contrato_id)).execute()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)}), 500


# ── Capital Investido ─────────────────────────────────────────────────────────

_CI_CSV_URL = (
    "https://docs.google.com/spreadsheets/d/e/"
    "2PACX-1vRZFq7H45YwN5Sbc9yZbSc9HGcTOl99X2jC2TVYVT828yCpilVXhzT55-W3Ma5ctQ"
    "/pub?gid=391372590&single=true&output=csv"
)
_CI_CACHE = {"day": None, "text": ""}


def _ci_fetch_csv():
    import requests as _req
    today = datetime.now(_BRT).strftime("%Y-%m-%d")
    if _CI_CACHE["day"] == today and _CI_CACHE["text"]:
        return _CI_CACHE["text"], None
    try:
        resp = _req.get(_CI_CSV_URL, timeout=15,
                        headers={"User-Agent": "Mozilla/5.0"},
                        allow_redirects=True)
        resp.raise_for_status()
        text = resp.content.decode("utf-8")
        _CI_CACHE["day"]  = today
        _CI_CACHE["text"] = text
        return text, None
    except Exception as e:
        return "", str(e)


def _fin_total_pago():
    try:
        sb   = _supabase()
        rows = sb.table("financiamentos_contratos").select("*").execute().data or []
        hoje = datetime.now(_BRT).date()
        total = 0.0
        for r in rows:
            if r.get("vendido"):
                continue
            parcelas = int(r["parcelas_total"])
            parcela  = float(r["valor_parcela"])
            entrada  = float(r.get("valor_entrada") or 0)
            restante = _fin_calcular_restante(r, hoje)
            pagas    = parcelas - restante
            total += pagas * parcela + entrada
        return total
    except Exception:
        return 0.0


@app.route("/capital-investido")
def pagina_capital_investido():
    csv_text, csv_error = _ci_fetch_csv()
    total_pago = _fin_total_pago()
    aportes_extra, aportes_fora = [], []
    try:
        sb   = _supabase()
        cols = "data, investidor, descricao, banco_destino, valor, computar"
        try:
            res = sb.table("capital_aportes").select(cols).order("data").execute()
        except Exception:
            # coluna `computar` ainda não criada — tudo entra no cálculo
            res = sb.table("capital_aportes").select(
                "data, investidor, descricao, banco_destino, valor"
            ).order("data").execute()
        for r in (res.data or []):
            item = {
                "data":          str(r["data"]),
                "investidor":    r["investidor"],
                "descricao":     r.get("descricao") or "",
                "banco_destino": r.get("banco_destino") or "",
                "valor":         float(r["valor"]),
            }
            if r.get("computar", True):
                aportes_extra.append(item)
            else:
                aportes_fora.append(item)
    except Exception:
        aportes_extra, aportes_fora = [], []
    try:
        tir = _capital_tir()
    except Exception as e:
        tir = {"ok": False, "erro": str(e)}
    return render_template("capital_investido.html",
        active="capital_investido",
        tir=tir,
        csv_text=csv_text,
        csv_error=csv_error,
        total_pago=total_pago,
        aportes_extra=aportes_extra,
        aportes_fora=aportes_fora,
    )


def _capital_aportes_todos():
    """Une os aportes das duas fontes — o CSV publicado da planilha APORTES e a
    tabela capital_aportes no Supabase — no formato que o cálculo de TIR espera.
    Registros marcados com computar=false ficam de fora."""
    from services.tir import parse_csv_aportes

    csv_text, _ = _ci_fetch_csv()
    linhas = parse_csv_aportes(csv_text)

    try:
        sb = _supabase()
        try:
            res = sb.table("capital_aportes").select(
                "data, investidor, descricao, banco_destino, valor, computar"
            ).order("data").execute()
        except Exception:
            res = sb.table("capital_aportes").select(
                "data, investidor, descricao, banco_destino, valor"
            ).order("data").execute()
        for r in (res.data or []):
            if not r.get("computar", True):
                continue
            linhas.append({
                "data":          str(r["data"]),
                "investidor":    r["investidor"],
                "descricao":     r.get("descricao") or "",
                "banco_destino": r.get("banco_destino") or "",
                "valor":         float(r["valor"]),
            })
    except Exception:
        pass

    return linhas


def _capital_tir(valor_veiculos=None):
    """Calcula a TIR por sócio. Sem valor informado, usa a FIPE atual da frota."""
    from services.tir import calcular_tir_por_socio

    origem = "manual"
    if valor_veiculos is None:
        origem = "fipe"
        try:
            valor_veiculos = _frota_valor_fipe_total()
        except Exception:
            valor_veiculos = None
    if valor_veiculos is None:
        return {"ok": False,
                "erro": "Sem valor de veículos: a frota não tem FIPE do mês "
                        "atual nem do anterior. Informe um valor manualmente."}

    dados = calcular_tir_por_socio(
        _capital_aportes_todos(),
        float(valor_veiculos),
        hoje=datetime.now(_BRT).date(),
    )
    dados["ok"] = True
    dados["origem_valor"] = origem
    return dados


@app.route("/api/capital/tir")
def api_capital_tir():
    """TIR (XIRR) por sócio. ?valor_veiculos=123456.78 sobrescreve a FIPE."""
    bruto = (request.args.get("valor_veiculos") or "").strip()
    valor = None
    if bruto:
        from services.tir import parse_brl
        valor = parse_brl(bruto)
        if valor is None:
            return jsonify({"ok": False, "erro": "Valor de veículos inválido."}), 400
    dados = _capital_tir(valor)
    return jsonify(dados), (200 if dados.get("ok") else 422)


@app.route("/api/capital/aportes", methods=["POST"])
def api_capital_aportes():
    body       = request.get_json(silent=True) or {}
    data       = (body.get("data")       or "").strip()
    investidor = (body.get("investidor") or "").strip()
    descricao  = (body.get("descricao")  or "").strip()
    banco_dest = (body.get("banco_destino") or "").strip()
    try:
        valor = float(str(body.get("valor", "")).replace(",", "."))
    except (ValueError, TypeError):
        return jsonify({"ok": False, "erro": "Valor inválido"}), 400
    if not data or not investidor:
        return jsonify({"ok": False, "erro": "Data e investidor são obrigatórios"}), 400
    sb = _supabase()
    if sb is None:
        return jsonify({"ok": False, "erro": "Supabase não configurado"}), 500
    registro = {
        "data":          data,
        "investidor":    investidor,
        "descricao":     descricao,
        "banco_destino": banco_dest,
        "valor":         valor,
    }
    if not body.get("computar", True):
        registro["computar"] = False
    sb.table("capital_aportes").insert(registro).execute()
    return jsonify({"ok": True, "data": data, "investidor": investidor, "valor": valor})


# ── Frota ─────────────────────────────────────────────────────────────────────

def _frota_valor_fipe_total(veiculos=None):
    """Soma o valor FIPE atual da frota, preferindo o valor manual do mês
    corrente, depois a FIPE do mês corrente e por fim a do mês anterior.
    Retorna None quando nenhum veículo tem valor conhecido."""
    if veiculos is None:
        veiculos, _, _ = _ler_frota_dados()
    curr_key, curr_label, prev_key, _ = _frota_mes_atual()
    manual = _frota_ler_manual()  # {placa: {mes_ref_label: {valor, ...}}}
    total, achou = 0.0, False
    for v in veiculos:
        mc = (manual.get(v.get("placa", "")) or {}).get(curr_label)
        if mc and mc.get("valor") is not None:
            total += float(mc["valor"]); achou = True
        elif v.get(curr_key) is not None:
            total += float(v[curr_key]); achou = True
        elif v.get(prev_key) is not None:
            total += float(v[prev_key]); achou = True
    return total if achou else None


def _ler_frota_dados():
    """
    Lê veículos e histórico FIPE do Supabase (frota_veiculos + frota_fipe_historico).
    Retorna (veiculos, codigos, erro) com a mesma estrutura anterior para compatibilidade
    com o template frota.html.
    """
    try:
        sb = _supabase()
        if sb is None:
            return [], [], "Supabase não configurado."

        res_v = sb.table("frota_veiculos").select(
            "modelo, placa, ano_modelo, cod_fipe, dt_aquisicao, vl_aquisicao"
        ).eq("ativo", True).execute()

        res_h = sb.table("frota_fipe_historico").select(
            "placa, mes_ref, valor"
        ).eq("fonte", "planilha").execute()

        # Pivot histórico: {placa: {python_key: valor}}
        # 'JAN/25' → 'jan25', 'MAI/26' → 'mai26'
        hist: dict[str, dict[str, float]] = {}
        for row in (res_h.data or []):
            key = row["mes_ref"].replace("/", "").lower()
            hist.setdefault(row["placa"], {})[key] = float(row["valor"])

        _EXCLUIR_PLACAS = {"QGO-2H58"}
        veiculos = []
        for v in (res_v.data or []):
            placa = v["placa"]
            if placa.upper() in _EXCLUIR_PLACAS:
                continue
            meses = hist.get(placa, {})
            dt = v.get("dt_aquisicao") or ""
            try:
                dt = datetime.strptime(dt, "%Y-%m-%d").strftime("%d/%m/%Y")
            except Exception:
                pass
            veiculos.append({
                "modelo":       v["modelo"],
                "placa":        placa,
                "ano_modelo":   v["ano_modelo"],
                "cod_fipe":     v["cod_fipe"],
                "dt_aquisicao": dt,
                "vl_aquisicao": float(v["vl_aquisicao"]) if v["vl_aquisicao"] is not None else None,
                **{k: meses.get(k) for k in (
                    "jan25","fev25","mar25","abr25","mai25","jun25",
                    "jul25","ago25","set25","out25","nov25","dez25",
                    "jan26","fev26","mar26","abr26","mai26","jun26",
                    "jul26","ago26","set26","out26","nov26","dez26",
                )},
            })

        # Derive codigos from vehicles (used by template for CODIGOS JS var)
        seen: dict[tuple, dict] = {}
        for v in veiculos:
            key = (v["cod_fipe"], v["ano_modelo"])
            if key not in seen:
                seen[key] = {"cod_fipe": v["cod_fipe"], "modelo": v["modelo"],
                             "ano_modelo": v["ano_modelo"], "qtd": 0}
            seen[key]["qtd"] += 1
        codigos = list(seen.values())

        return veiculos, codigos, None
    except Exception as e:
        import traceback; traceback.print_exc()
        return [], [], str(e)


def _frota_mes_atual():
    """Retorna (curr_key, curr_label, prev_key, prev_label) baseado na data do sistema."""
    MESES = ['jan','fev','mar','abr','mai','jun','jul','ago','set','out','nov','dez']
    hoje  = datetime.now(_BRT)
    ano, mes = hoje.year, hoje.month
    yy        = str(ano)[2:]
    curr_key   = MESES[mes - 1] + yy
    curr_label = MESES[mes - 1].upper() + '/' + yy
    prev_mes  = mes - 1 if mes > 1 else 12
    prev_ano  = ano if mes > 1 else ano - 1
    prev_yy   = str(prev_ano)[2:]
    prev_key   = MESES[prev_mes - 1] + prev_yy
    prev_label = MESES[prev_mes - 1].upper() + '/' + prev_yy
    return curr_key, curr_label, prev_key, prev_label


def _frota_ler_manual():
    """Retorna {placa: {mes_ref: {valor, atualizado_em}}} lido de frota_fipe_historico."""
    try:
        sb = _supabase()
        if sb is None:
            return {}
        res = sb.table("frota_fipe_historico").select(
            "placa, mes_ref, valor, atualizado_em"
        ).eq("fonte", "manual").execute()
        out = {}
        for row in (res.data or []):
            placa  = row["placa"]
            ref    = row["mes_ref"]
            dt_str = row.get("atualizado_em") or ""
            try:
                dt_str = datetime.strptime(dt_str, "%Y-%m-%d").strftime("%d/%m/%Y")
            except Exception:
                pass
            out.setdefault(placa, {})[ref] = {
                "valor":         float(row["valor"]),
                "atualizado_em": dt_str,
            }
        return out
    except Exception:
        import traceback; traceback.print_exc()
        return {}


def _frota_salvar_manual(placa, valor, ref):
    """Upsert de (placa, mes_ref) em frota_fipe_historico com fonte='manual'."""
    sb = _supabase()
    if sb is None:
        return
    sb.table("frota_fipe_historico").upsert({
        "placa":         placa,
        "mes_ref":       ref,
        "valor":         valor,
        "fonte":         "manual",
        "atualizado_em": datetime.now(_BRT).strftime("%Y-%m-%d"),
    }, on_conflict="placa,mes_ref").execute()


_SOB_ADM_FIPE_VALORES = {
    "005540-9": 800.0,
    "095010-6": 1200.0,
}
_SOB_ADM_PLACA_VALORES = {}
_SOB_ADM_TAXA = 0.15
_SOB_ADM_PLACA_EXTRA = ""


def _ler_sob_administracao():
    """Lê DADOS_CLIENTES_CONS.xlsx: veículos cuja unidade não seja Ativuz/AZ ou placa especial."""
    import openpyxl
    xlsx_path = _clientes_cons_xlsx_path()
    if not xlsx_path.exists():
        return [], None

    try:
        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not rows:
            return [], None

        header = rows[0]

        def _norm(s):
            s = unicodedata.normalize("NFD", str(s or "").lower())
            return "".join(c for c in s if unicodedata.category(c) != "Mn")

        def _ci(kw):
            nk = _norm(kw)
            return next((i for i, h in enumerate(header) if nk in _norm(str(h or ""))), None)

        def _ci_exact(name):
            n = _norm(name)
            return next((i for i, h in enumerate(header) if _norm(str(h or "")) == n), None)

        i_placa    = _ci("placa")
        i_mod      = _ci_exact("modelo") if _ci_exact("modelo") is not None else _ci("modelo")
        i_marca    = _ci("marca")
        i_prop     = _ci("unidade do ve")
        i_loc      = _ci("razao social cliente") or _ci("razao social") or _ci("cliente")
        i_tipo     = _ci("tipo de contrato")
        i_ini      = _ci("inicio do contrato") or _ci("inicio de contrato") or _ci("inicio")
        i_fim      = _ci("termino do contrato") or _ci("termino")
        i_anomod   = _ci("ano modelo")

        def _v(row, i):
            if i is None or i >= len(row): return ""
            v = row[i]
            if v is None: return ""
            if hasattr(v, "strftime"): return v.strftime("%d/%m/%Y")
            return str(v).strip()

        _EXTRA = _SOB_ADM_PLACA_EXTRA.upper().replace("-", "").replace(" ", "")

        hoje = date.today()
        veiculos = []
        for row in rows[1:]:
            placa = _v(row, i_placa)
            if not placa:
                continue

            placa_id = placa.upper().replace("-", "").replace(" ", "")
            marca    = _norm(_v(row, i_marca))
            modelo   = _norm(_v(row, i_mod))

            # inclui: Polo, BYD ou placa especial
            eh_polo  = "polo" in modelo
            eh_byd   = "byd" in marca
            eh_extra = bool(_EXTRA) and placa_id == _EXTRA
            if not (eh_polo or eh_byd or eh_extra):
                continue

            unid = "GC AUTOELÉTRICA" if eh_extra else _v(row, i_prop)

            if eh_polo:
                valor_s = 800.0
            elif eh_byd:
                valor_s = 1200.0
            else:
                valor_s = _SOB_ADM_PLACA_VALORES.get(placa.upper())
            taxa_s  = round(valor_s * _SOB_ADM_TAXA, 2) if valor_s else None

            ini_raw = row[i_ini] if i_ini is not None and i_ini < len(row) else None
            ini_date = None
            if ini_raw:
                if isinstance(ini_raw, datetime): ini_date = ini_raw.date()
                elif isinstance(ini_raw, date):   ini_date = ini_raw
                else:
                    for fmt in ["%d/%m/%Y", "%Y-%m-%d"]:
                        try: ini_date = datetime.strptime(str(ini_raw)[:10], fmt).date(); break
                        except ValueError: pass

            dias_ativos = (hoje - ini_date).days if ini_date else 0
            receita_acum = round(dias_ativos * (taxa_s / 7), 2) if taxa_s and dias_ativos > 0 else 0.0

            ini_fmt = ini_date.strftime("%d/%m/%Y") if ini_date else ""

            veiculos.append({
                "placa":        placa,
                "montadora":    _v(row, i_marca),
                "modelo":       _v(row, i_mod),
                "fipe":         "",
                "ano_fab":      _v(row, i_anomod),
                "ano_mod":      _v(row, i_anomod),
                "proprietario": unid,
                "locatario":    _v(row, i_loc),
                "tipo_contrato":_v(row, i_tipo),
                "km":           "",
                "inicio":       ini_fmt,
                "termino":      _v(row, i_fim),
                "situacao":     "EM ANDAMENTO",
                "valor_semanal":valor_s,
                "taxa_semanal": taxa_s,
                "dias_ativos":  dias_ativos,
                "receita_acum": receita_acum,
            })

        veiculos.sort(key=lambda v: (v["proprietario"], v["inicio"]))
        return veiculos, None

    except Exception as e:
        import traceback; traceback.print_exc()
        return [], str(e)


def _gerar_segundas(ini, fim):
    """Retorna todas as segundas-feiras em [ini, fim] inclusive."""
    from datetime import timedelta
    days = (7 - ini.weekday()) % 7   # 0 se já é segunda
    cur = ini + timedelta(days=days)
    result = []
    while cur <= fim:
        result.append(cur)
        cur += timedelta(weeks=1)
    return result


@app.route("/api/sob-adm/recebimentos")
def api_sob_adm_recebimentos():
    """
    Retorna cada veículo com seu calendário de segundas-feiras (passadas + futuras)
    e o status de recebimento de cada uma.

    Supabase — tabela necessária (execute uma vez):
      CREATE TABLE sob_adm_recebimentos (
        id uuid DEFAULT gen_random_uuid() PRIMARY KEY,
        placa text NOT NULL,
        data_semana date NOT NULL,
        taxa_valor numeric(10,2) NOT NULL DEFAULT 0,
        recebido boolean DEFAULT false,
        created_at timestamptz DEFAULT now(),
        UNIQUE(placa, data_semana)
      );
    """
    from datetime import timedelta
    sob_adm, erro = _ler_sob_administracao()
    if erro:
        return jsonify({"ok": False, "erro": erro})

    # Carrega registros do Supabase
    recebidos = {}   # (placa, "YYYY-MM-DD") -> bool
    sb = _supabase()
    if sb:
        try:
            res = sb.table("sob_adm_recebimentos").select("placa,data_semana,recebido").execute()
            for r in (res.data or []):
                ds = str(r["data_semana"])[:10]
                recebidos[(r["placa"], ds)] = bool(r.get("recebido", False))
        except Exception:
            pass

    hoje = date.today()
    result = []

    for v in sob_adm:
        if not v["inicio"] or not v["taxa_semanal"]:
            continue
        try:
            ini = datetime.strptime(v["inicio"], "%d/%m/%Y").date()
        except ValueError:
            continue

        fim_contrato = None
        if v.get("termino"):
            for fmt in ["%d/%m/%Y", "%Y-%m-%d"]:
                try:
                    fim_contrato = datetime.strptime(str(v["termino"])[:10], fmt).date()
                    break
                except ValueError:
                    pass

        ate_hoje = min(hoje, fim_contrato) if fim_contrato else hoje
        segundas_passadas = _gerar_segundas(ini, ate_hoje)
        segundas_futuras  = _gerar_segundas(hoje + timedelta(days=1), fim_contrato) if fim_contrato and fim_contrato > hoje else []

        semanas = []
        for d in segundas_passadas:
            ds = d.isoformat()
            semanas.append({"data": ds, "recebido": recebidos.get((v["placa"], ds), False), "passada": True})
        for d in segundas_futuras:
            ds = d.isoformat()
            semanas.append({"data": ds, "recebido": False, "passada": False})

        taxa = v["taxa_semanal"]
        rec  = round(sum(taxa for s in semanas if s["passada"] and s["recebido"]), 2)
        pend = round(sum(taxa for s in semanas if s["passada"] and not s["recebido"]), 2)
        fut  = round(sum(taxa for s in semanas if not s["passada"]), 2)

        result.append({
            "placa":            v["placa"],
            "modelo":           v["modelo"],
            "proprietario":     v["proprietario"],
            "locatario":        v["locatario"],
            "valor_semanal":    v["valor_semanal"],
            "taxa_semanal":     taxa,
            "inicio":           v["inicio"],
            "termino":          v.get("termino") or None,
            "semanas":          semanas,
            "recebido_total":   rec,
            "pendente_total":   pend,
            "projetado_futuro": fut,
        })

    return jsonify({"ok": True, "veiculos": result})


@app.route("/api/sob-adm/recebimentos/toggle", methods=["POST"])
def api_sob_adm_toggle():
    dados = request.get_json(force=True, silent=True) or {}
    placa       = str(dados.get("placa", "")).strip()
    data_semana = str(dados.get("data_semana", "")).strip()

    if not placa or not data_semana:
        return jsonify({"ok": False, "erro": "Parâmetros inválidos"}), 400

    sob_adm, _ = _ler_sob_administracao()
    v = next((x for x in sob_adm if x["placa"] == placa), None)
    taxa = float(v["taxa_semanal"]) if v and v.get("taxa_semanal") else 0.0

    sb = _supabase()
    if not sb:
        return jsonify({"ok": False, "erro": "Banco indisponível"}), 503

    try:
        res = sb.table("sob_adm_recebimentos").select("id,recebido").eq("placa", placa).eq("data_semana", data_semana).execute()
        if res.data:
            novo = not res.data[0]["recebido"]
            sb.table("sob_adm_recebimentos").update({"recebido": novo}).eq("id", res.data[0]["id"]).execute()
        else:
            novo = True
            sb.table("sob_adm_recebimentos").insert({"placa": placa, "data_semana": data_semana, "taxa_valor": taxa, "recebido": True}).execute()
        return jsonify({"ok": True, "recebido": novo})
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)}), 500



_CONTRATOS_XLSX = Path(__file__).parent / "planilhas" / "contratos_locacao.xlsx"


def _sync_contratos_supabase(contratos: list):
    """Upsert de todos os contratos na tabela contratos_frota do Supabase."""
    if not contratos:
        return
    try:
        sb = _supabase()
        if not sb:
            return
        rows = []
        from datetime import datetime as _dt
        for c in contratos:
            rows.append({
                "contrato_comercial": c["contrato_comercial"],
                "contrato_locacao":   c["contrato_locacao"],
                "cliente":            c["cliente"],
                "placa":              c["placa"],
                "modelo":             c["modelo"],
                "grupo":              c["grupo"],
                "tipo_contrato":      c["tipo_contrato"],
                "tipo_pessoa":        c["tipo_pessoa"],
                "situacao":           c["situacao"],
                "sit_faturamento":    c["sit_faturamento"],
                "periodo":            c["periodo"],
                "inicio":             c["inicio"],
                "termino_previsto":   c["termino_previsto"],
                "valor_locacao":      c["valor_locacao"],
                "valor_inicial":      c["valor_inicial"],
                "gasto_total":        c["gasto_total"],
                "gasto_sinistros":    c["gasto_sinistros"],
                "gasto_manutencao":   c["gasto_manutencao"],
                "km":                 c["km"],
                "unidade_fat":        c["unidade_fat"],
                "atualizado_em":      _dt.utcnow().isoformat(),
            })
        sb.table("contratos_frota").upsert(rows, on_conflict="contrato_comercial").execute()
    except Exception as e:
        print(f"[sync_contratos] erro: {e}")


def _sync_contas_supabase():
    """Lê CONTAS-A-RECEBER.xlsx e upsert em contas_receber_frota no Supabase."""
    try:
        sb = _supabase()
        if not sb:
            return
        import openpyxl
        from datetime import datetime as _dt, date as _date

        _base     = Path(__file__).parent / "planilhas"
        xlsx_path = _base / "CONTAS-A-RECEBER.xlsx"
        if not xlsx_path.exists():
            return

        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()

        header_idx = 0
        for ri, row in enumerate(all_rows[:10]):
            nh_row = [_nh(str(c or "")) for c in row]
            if sum(1 for t in ["receber de", "vencimento", "valor"]
                   if any(t in n for n in nh_row)) >= 2:
                header_idx = ri
                break

        header    = all_rows[header_idx]
        data_rows = all_rows[header_idx + 1:]

        def _ci(keyword):
            nk = _nh(keyword)
            return next((i for i, h in enumerate(header)
                         if h is not None and nk in _nh(str(h))), None)

        i_nome  = _ci("receber de (fantasia)") or _ci("receber de")
        i_valor = _ci("valor previsto") or _ci("valor")
        i_venc  = _ci("data de vencimento") or _ci("vencimento")
        i_sit   = _ci("situacao (data de vencimento)") or _ci("situacao")
        i_tipo  = _ci("tipo de fatura") or _ci("tipo")
        i_doc   = _ci("numero do documento") or _ci("documento")
        i_unid  = _ci("unidade")
        i_comp  = _ci("data de competencia") or _ci("competencia")

        hoje = _date.today()
        seen = {}
        for row in data_rows:
            def _get(idx):
                return row[idx] if idx is not None and idx < len(row) else None

            nome_raw = _get(i_nome)
            if not nome_raw:
                continue
            nome = str(nome_raw).strip()
            if not nome:
                continue

            venc_raw = _get(i_venc)
            venc_date = None
            if venc_raw:
                if isinstance(venc_raw, _dt):
                    venc_date = venc_raw.date()
                elif isinstance(venc_raw, _date):
                    venc_date = venc_raw
                else:
                    venc_str = str(venc_raw).strip()
                    for fmt in ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"]:
                        try:
                            venc_date = _dt.strptime(venc_str, fmt).date()
                            break
                        except (ValueError, TypeError):
                            pass
            if venc_date is None:
                continue

            comp_raw  = _get(i_comp)
            comp_date = None
            if comp_raw:
                if isinstance(comp_raw, _dt):
                    comp_date = comp_raw.date()
                elif isinstance(comp_raw, _date):
                    comp_date = comp_raw

            num_doc     = str(_get(i_doc)   or "").strip()
            unidade     = str(_get(i_unid)  or "").strip()
            sit_raw     = _get(i_sit)
            tipo_raw    = _get(i_tipo)
            situacao    = str(sit_raw  or "").strip()
            tipo_fatura = str(tipo_raw or "").strip()
            valor       = _parse_valor_excel(_get(i_valor))

            for bad in ("None", "nan", ""):
                if num_doc  == bad: num_doc  = ""
                if unidade  == bad: unidade  = ""

            # positivo = dias até vencer (futuro), negativo = dias em atraso
            dias_vencimento = (venc_date - hoje).days if venc_date else None

            faixa = ""
            if dias_vencimento is not None:
                d = dias_vencimento
                if d > 0:    faixa = "A vencer"
                elif d == 0: faixa = "Vence hoje"
                elif d >= -7: faixa = "1-7 dias"
                elif d >= -15: faixa = "8-15 dias"
                elif d >= -30: faixa = "16-30 dias"
                else: faixa = "Mais de 30 dias"

            row_id = f"{num_doc}|{venc_date.isoformat()}" if num_doc else f"{nome}|{venc_date.isoformat()}"
            seen[row_id] = {
                "id":               row_id,
                "numero_documento": num_doc,
                "cliente":          nome,
                "unidade":          unidade,
                "data_vencimento":  venc_date.isoformat(),
                "data_competencia": comp_date.isoformat() if comp_date else None,
                "valor":            valor,
                "situacao":         situacao,
                "tipo_fatura":      tipo_fatura,
                "dias_vencimento":  dias_vencimento,
                "faixa_vencimento": faixa,
                "atualizado_em":    _dt.utcnow().isoformat(),
            }

        rows = list(seen.values())
        if rows:
            sb.table("contas_receber_frota").upsert(rows, on_conflict="id").execute()
    except Exception as e:
        print(f"[sync_contas] erro: {e}")


def _ler_contratos():
    import openpyxl
    if not _CONTRATOS_XLSX.exists():
        return [], f"Arquivo não encontrado: {_CONTRATOS_XLSX.name}"
    try:
        wb = openpyxl.load_workbook(str(_CONTRATOS_XLSX), data_only=True)
        ws = wb["Relatório"]
        today = date.today()

        def _fmt(v):
            if isinstance(v, datetime): return v.strftime('%d/%m/%Y')
            if isinstance(v, date):     return v.strftime('%d/%m/%Y')
            return str(v) if v else ''

        rows = []
        for row in ws.iter_rows(min_row=6, max_row=ws.max_row, values_only=True):
            if not row[9]:  # Contrato Comercial
                continue
            termino_raw = row[52]  # Término Previsto
            dias_vencer = None
            if isinstance(termino_raw, datetime):
                dias_vencer = (termino_raw.date() - today).days
            rows.append({
                'contrato_comercial': str(row[9]  or ''),   # Contrato Comercial
                'contrato_locacao':   str(row[12] or ''),   # Contrato de Locação
                'periodo':            str(row[41] or ''),   # Período (meses)
                'cliente':            str(row[6]  or ''),   # Cliente
                'unidade_fat':        str(row[54] or ''),   # Unidade de faturamento
                'valor_locacao':      float(row[57] or 0),  # Valor de locação vigente
                'tipo_pessoa':        str(row[51] or ''),   # Tipo de cliente
                'gasto_total':        float(row[27] or 0),  # Gasto Total
                'gasto_sinistros':    float(row[26] or 0),  # Gasto Sinistros
                'gasto_manutencao':   float(row[25] or 0),  # Gasto Manutenção
                'inicio':             _fmt(row[30]),         # Início de Contrato
                'termino_previsto':   _fmt(row[52]),         # Término Previsto
                'situacao':           str(row[46] or ''),   # Situação
                'placa':              str(row[59] or ''),   # Veículo Atual
                'modelo':             str(row[36] or ''),   # Modelo
                'km':                 int(row[31]  or 0),   # Km confirmado
                'grupo':              str(row[28] or ''),   # Grupo
                'tipo_contrato':      str(row[50] or ''),   # Tipo de Contrato
                'sit_faturamento':    str(row[47] or ''),   # Situação de Faturamento
                'valor_inicial':      float(row[58] or 0),  # Valor inicial de locação
                'dias_vencer':        dias_vencer,
            })
        wb.close()
        return rows, None
    except Exception as ex:
        return [], str(ex)


@app.route("/insights/contratos")
def pagina_contratos():
    contratos, erro = _ler_contratos()
    _sync_contratos_supabase(contratos)
    ativos = [c for c in contratos if c['situacao'] == 'EM ANDAMENTO']

    receita_mes    = sum(c['valor_locacao'] for c in ativos)
    gasto_acum     = sum(c['gasto_total']   for c in contratos)
    a_vencer_30    = sum(1 for c in ativos
                         if c['dias_vencer'] is not None and 0 <= c['dias_vencer'] <= 30)

    tipos_count = {}
    for c in contratos:
        t = c['tipo_contrato'] or 'Não informado'
        tipos_count[t] = tipos_count.get(t, 0) + 1

    clientes_rec = {}
    for c in ativos:
        clientes_rec[c['cliente']] = clientes_rec.get(c['cliente'], 0) + c['valor_locacao']
    top10 = sorted(clientes_rec.items(), key=lambda x: -x[1])[:10]

    return render_template("contratos.html",
        active="contratos",
        contratos=contratos,
        erro=erro,
        kpi_ativos=len(ativos),
        kpi_receita=receita_mes,
        kpi_vencer30=a_vencer_30,
        kpi_gasto=gasto_acum,
        tipos_count=tipos_count,
        top10_clientes=top10,
    )


@app.route("/insights/frota")
def pagina_frota():
    veiculos, codigos, erro = _ler_frota_dados()
    manual = _frota_ler_manual()
    curr_key, curr_label, prev_key, prev_label = _frota_mes_atual()
    sob_adm, sob_adm_erro = _ler_sob_administracao()
    return render_template("frota.html",
        active="frota",
        veiculos=veiculos,
        codigos=codigos,
        manual=manual,
        erro=erro,
        curr_mes_key=curr_key,
        curr_mes_label=curr_label,
        prev_mes_key=prev_key,
        prev_mes_label=prev_label,
        sob_adm=sob_adm,
        sob_adm_erro=sob_adm_erro,
    )


@app.route("/api/frota/manual", methods=["POST"])
def api_frota_manual():
    body  = request.get_json(silent=True) or {}
    placa = (body.get("placa") or "").strip().upper()
    ref   = (body.get("ref")   or "").strip()
    try:
        valor = float(str(body.get("valor", "")).replace(",", "."))
    except (ValueError, TypeError):
        return jsonify({"ok": False, "erro": "Valor inválido"}), 400
    if not placa:
        return jsonify({"ok": False, "erro": "Placa obrigatória"}), 400
    _frota_salvar_manual(placa, valor, ref)
    return jsonify({"ok": True, "placa": placa, "valor": valor, "ref": ref})


@app.route("/api/frota/manual/batch", methods=["POST"])
def api_frota_manual_batch():
    """Upsert em bulk por combinação (cod_fipe + ano_modelo) em frota_fipe_historico."""
    body     = request.get_json(silent=True) or {}
    entradas = body.get("entradas") or []
    if not entradas:
        return jsonify({"ok": False, "erro": "Nenhuma entrada"}), 400

    sb = _supabase()
    if sb is None:
        return jsonify({"ok": False, "erro": "Supabase não configurado"}), 500

    res_v = sb.table("frota_veiculos").select(
        "placa, cod_fipe, ano_modelo"
    ).eq("ativo", True).execute()
    veiculos = res_v.data or []
    agora = datetime.now(_BRT).strftime("%Y-%m-%d")

    rows = []
    for entrada in entradas:
        cod     = (entrada.get("cod_fipe")   or "").strip()
        ano_mod = (entrada.get("ano_modelo") or "").strip()
        ref     = (entrada.get("ref")        or "").strip()
        try:
            valor = float(str(entrada.get("valor", "")).replace(",", "."))
        except (ValueError, TypeError):
            continue
        for v in veiculos:
            if v["cod_fipe"] == cod and ((not ano_mod) or v["ano_modelo"] == ano_mod):
                rows.append({
                    "placa":         v["placa"],
                    "mes_ref":       ref,
                    "valor":         valor,
                    "fonte":         "manual",
                    "atualizado_em": agora,
                })

    if rows:
        sb.table("frota_fipe_historico").upsert(rows, on_conflict="placa,mes_ref").execute()

    return jsonify({"ok": True, "atualizados": len(rows)})


# ── Carteira Judicializada ────────────────────────────────────────────────────

@app.route("/api/carteira-judicializada", methods=["GET"])
def api_carteira_judicializada_listar():
    sb = _supabase()
    if sb is None:
        return jsonify({"ok": False, "erro": "Supabase não configurado"}), 503
    try:
        res = sb.table("carteira_judicializada").select("*").order("criado_em").execute()
        return jsonify({"ok": True, "data": res.data or []})
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)}), 500


@app.route("/api/carteira-judicializada", methods=["POST"])
def api_carteira_judicializada_inserir():
    body = request.get_json(silent=True) or {}
    cliente = (body.get("cliente") or "").strip()
    if not cliente:
        return jsonify({"ok": False, "erro": "Cliente obrigatório"}), 400
    sb = _supabase()
    if sb is None:
        return jsonify({"ok": False, "erro": "Supabase não configurado"}), 503
    try:
        res = sb.table("carteira_judicializada").insert({
            "cliente":          cliente,
            "cpf_cnpj":         (body.get("cpf_cnpj")        or "").strip(),
            "avalista":         (body.get("avalista")         or "").strip(),
            "cpf_avalista":     (body.get("cpf_avalista")     or "").strip(),
            "inicio_divida":    body.get("inicio_divida")     or None,
            "valor_atual":      float(body.get("valor_atual") or 0),
            "status":           body.get("status")            or "Ajuizado",
            "num_processo":     (body.get("num_processo")     or "").strip(),
            "proximo_prazo":    body.get("proximo_prazo")     or None,
            "descricao_prazo":  (body.get("descricao_prazo")  or "").strip(),
        }).execute()
        return jsonify({"ok": True, "data": res.data[0] if res.data else {}})
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)}), 500


@app.route("/api/carteira-judicializada/<uuid:registro_id>", methods=["PUT"])
def api_carteira_judicializada_atualizar(registro_id):
    body = request.get_json(silent=True) or {}
    cliente = (body.get("cliente") or "").strip()
    if not cliente:
        return jsonify({"ok": False, "erro": "Cliente obrigatório"}), 400
    sb = _supabase()
    if sb is None:
        return jsonify({"ok": False, "erro": "Supabase não configurado"}), 503
    try:
        res = sb.table("carteira_judicializada").update({
            "cliente":          cliente,
            "cpf_cnpj":         (body.get("cpf_cnpj")        or "").strip(),
            "avalista":         (body.get("avalista")         or "").strip(),
            "cpf_avalista":     (body.get("cpf_avalista")     or "").strip(),
            "inicio_divida":    body.get("inicio_divida")     or None,
            "valor_atual":      float(body.get("valor_atual") or 0),
            "status":           body.get("status")            or "Ajuizado",
            "num_processo":     (body.get("num_processo")     or "").strip(),
            "proximo_prazo":    body.get("proximo_prazo")     or None,
            "descricao_prazo":  (body.get("descricao_prazo")  or "").strip(),
            "atualizado_em":    "now()",
        }).eq("id", str(registro_id)).execute()
        return jsonify({"ok": True, "data": res.data[0] if res.data else {}})
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)}), 500


# ── Acordo Judicializado ─────────────────────────────────────────────────────

@app.route("/api/carteira-judicializada/<uuid:registro_id>/acordo", methods=["PUT"])
def api_jud_acordo_salvar(registro_id):
    body = request.get_json(silent=True) or {}
    sb = _supabase()
    if sb is None:
        return jsonify({"ok": False, "erro": "Supabase não configurado"}), 503
    try:
        sb.table("carteira_judicializada").update({
            "acordo_dados": body.get("acordo_dados"),
        }).eq("id", str(registro_id)).execute()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)}), 500


@app.route("/api/carteira-judicializada/<uuid:registro_id>", methods=["DELETE"])
def api_jud_excluir(registro_id):
    sb = _supabase()
    if sb is None:
        return jsonify({"ok": False, "erro": "Supabase não configurado"}), 503
    try:
        sb.table("carteira_judicializada").delete().eq("id", str(registro_id)).execute()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)}), 500


# ── Google Calendar — Carteira Judicializada ─────────────────────────────────

@app.route("/api/google-calendar/status")
def api_gcal_status():
    try:
        from services.google_calendar import is_authorized, has_credentials
        tem_creds = has_credentials()
        autorizado = is_authorized() if tem_creds else False
        return jsonify({"ok": True, "autorizado": autorizado, "tem_credenciais": tem_creds})
    except ImportError:
        return jsonify({"ok": True, "autorizado": False, "tem_credenciais": False})

@app.route("/api/google-calendar/auth")
def api_gcal_auth():
    from services.google_calendar import get_auth_url, CREDENTIALS_FILE
    if not CREDENTIALS_FILE.exists():
        return jsonify({"ok": False, "erro": "Arquivo google_credentials.json não encontrado"}), 400
    redirect_uri = _os.environ.get("GCAL_REDIRECT_URI", "http://localhost:5000/oauth2callback")
    url, state = get_auth_url(redirect_uri)
    session['gcal_state'] = state
    return jsonify({"ok": True, "url": url})

@app.route("/oauth2callback")
def oauth2callback():
    from services.google_calendar import exchange_code
    code  = request.args.get('code')
    state = request.args.get('state')
    redirect_uri = _os.environ.get("GCAL_REDIRECT_URI", "http://localhost:5000/oauth2callback")
    try:
        exchange_code(code, state, redirect_uri)
        return "<script>window.close(); window.opener && window.opener.location.reload();</script><p>Autorizado! Pode fechar esta aba.</p>"
    except Exception as e:
        return f"<p>Erro: {e}</p>", 400

@app.route("/api/carteira-judicializada/<uuid:registro_id>/sync-calendar", methods=["POST"])
def api_jud_sync_calendar(registro_id):
    from services.google_calendar import criar_eventos_parcelas, is_authorized
    if not is_authorized():
        return jsonify({"ok": False, "erro": "Google Calendar não autorizado"}), 403
    sb = _supabase()
    if sb is None:
        return jsonify({"ok": False, "erro": "Supabase não configurado"}), 503
    try:
        res = sb.table("carteira_judicializada").select("*").eq("id", str(registro_id)).single().execute()
        registro = res.data
        if not registro:
            return jsonify({"ok": False, "erro": "Registro não encontrado"}), 404
        ok, resultado = criar_eventos_parcelas(registro)
        if not ok:
            return jsonify({"ok": False, "erro": resultado}), 500
        return jsonify({"ok": True, "criados": resultado})
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)}), 500

@app.route("/api/carteira-judicializada/sync-calendar-todos", methods=["POST"])
def api_jud_sync_calendar_todos():
    from services.google_calendar import criar_eventos_parcelas, is_authorized
    if not is_authorized():
        return jsonify({"ok": False, "erro": "Google Calendar não autorizado"}), 403
    sb = _supabase()
    if sb is None:
        return jsonify({"ok": False, "erro": "Supabase não configurado"}), 503
    try:
        res = sb.table("carteira_judicializada").select("*").eq("status", "Acordo").execute()
        registros = res.data or []
        total_criados = 0
        for registro in registros:
            ok, resultado = criar_eventos_parcelas(registro)
            if ok:
                total_criados += resultado
        return jsonify({"ok": True, "criados": total_criados, "acordos": len(registros)})
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)}), 500


# ── Checklist Judicializada ───────────────────────────────────────────────────

@app.route("/api/jud-checklist/<uuid:registro_id>", methods=["GET"])
def api_jud_checklist_get(registro_id):
    sb = _supabase()
    if sb is None:
        return jsonify({"ok": False, "erro": "Supabase não configurado"}), 503
    try:
        key = "JUD-" + str(registro_id)
        res = sb.table("checklist_contratos").select("*").eq("contrato", key).execute()
        if res.data:
            contrato_id = res.data[0]["id"]
        else:
            ins = sb.table("checklist_contratos").insert({"contrato": key, "placa": "", "cliente": "", "unidade": ""}).execute()
            contrato_id = ins.data[0]["id"]
        itens = sb.table("checklist_itens").select("*").eq("contrato_id", contrato_id).order("created_at").execute().data or []
        return jsonify({"ok": True, "contrato_id": contrato_id,
                        "itens": [{"id": i["id"], "nome": i["nome"], "marcado": i["marcado"]} for i in itens]})
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)}), 500


@app.route("/api/jud-checklist/<uuid:registro_id>/item", methods=["POST"])
def api_jud_checklist_add_item(registro_id):
    body = request.get_json(silent=True) or {}
    contrato_id = body.get("contrato_id")
    nome = (body.get("nome") or "").strip()
    if not contrato_id or not nome:
        return jsonify({"ok": False, "erro": "contrato_id e nome são obrigatórios"}), 400
    sb = _supabase()
    if sb is None:
        return jsonify({"ok": False, "erro": "Supabase não configurado"}), 503
    try:
        res = sb.table("checklist_itens").insert({"contrato_id": contrato_id, "nome": nome, "marcado": False}).execute()
        item = res.data[0]
        return jsonify({"ok": True, "item": {"id": item["id"], "nome": item["nome"], "marcado": item["marcado"]}})
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)}), 500


@app.route("/api/jud-checklist/item/<uuid:item_id>", methods=["PUT"])
def api_jud_checklist_toggle(item_id):
    body = request.get_json(silent=True) or {}
    sb = _supabase()
    if sb is None:
        return jsonify({"ok": False, "erro": "Supabase não configurado"}), 503
    try:
        sb.table("checklist_itens").update({"marcado": bool(body.get("marcado", False))}).eq("id", str(item_id)).execute()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)}), 500


@app.route("/api/jud-checklist/item/<uuid:item_id>", methods=["DELETE"])
def api_jud_checklist_delete(item_id):
    sb = _supabase()
    if sb is None:
        return jsonify({"ok": False, "erro": "Supabase não configurado"}), 503
    try:
        sb.table("checklist_itens").delete().eq("id", str(item_id)).execute()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)}), 500


# ── Checklist ─────────────────────────────────────────────────────────────────

def _clientes_cons_xlsx_path():
    base = Path(__file__).resolve().parent
    return base / "planilhas" / "DADOS_CLIENTES_CONS.xlsx"


_IMAGEM_MAP = [
    ("GOL",     "gol_sf.png"),
    ("VOYAGE",  "voyage.png"),
    ("POLO",    "polo.png"),
    ("DOLPHIN", "byd.png"),
    ("SANDERO", "sandero.png"),
    ("CG",      "CG.png"),
    ("NXR",     "nxr.png"),
]
_BLEND_MULTIPLY = {"gol_sf.png", "voyage.png", "polo.png", "byd.png", "CG.png"}


def _imagem_veiculo(modelo):
    m = (modelo or "").upper()
    for keyword, fname in _IMAGEM_MAP:
        if keyword in m:
            return fname
    return None


def _ler_veiculos():
    import openpyxl
    xlsx_path = _clientes_cons_xlsx_path()
    if not xlsx_path.exists():
        return [], "Planilha não encontrada em planilhas/DADOS_CLIENTES_CONS.xlsx."

    try:
        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if len(rows) < 1:
            return [], "Planilha sem dados."

        header_row = rows[0]

        def _norm(s):
            s = unicodedata.normalize("NFD", str(s or "").lower())
            return "".join(c for c in s if unicodedata.category(c) != "Mn")

        headers_norm = [_norm(h) for h in header_row]

        def _ci(keyword):
            kn = _norm(keyword)
            return next((i for i, h in enumerate(headers_norm) if kn in h), None)

        def _ci_exact(name):
            n = _norm(name)
            return next((i for i, h in enumerate(headers_norm) if h == n), None)

        i_placa    = _ci("placa")
        i_modelo   = _ci_exact("modelo") if _ci_exact("modelo") is not None else _ci("modelo")
        i_cliente  = _ci("razao social cliente") or _ci("razao social") or _ci("cliente")
        i_contrato = _ci("contrato de locacao") or _ci("contrato")
        i_unidade  = _ci("unidade do veiculo") or _ci("unidade")
        i_inicio   = _ci("inicio de contrato") or _ci("inicio")
        i_termino  = _ci("termino de contrato") or _ci("termino")
        i_tipo     = _ci("tipo de contrato")

        def _v(row, i):
            if i is None or i >= len(row):
                return ""
            v = row[i]
            if v is None:
                return ""
            if hasattr(v, "strftime"):
                return v.strftime("%d/%m/%Y")
            return str(v).strip()

        seen_placas = {}
        for row in rows[1:]:
            placa = _v(row, i_placa)
            if not placa:
                continue
            modelo = _v(row, i_modelo)
            img    = _imagem_veiculo(modelo)
            seen_placas[placa] = {
                "placa":    placa,
                "modelo":   modelo,
                "cliente":  _v(row, i_cliente),
                "contrato": _v(row, i_contrato),
                "unidade":  _v(row, i_unidade),
                "inicio":   _v(row, i_inicio),
                "termino":  _v(row, i_termino),
                "tipo":     _v(row, i_tipo),
                "imagem":   img,
                "blend":    img in _BLEND_MULTIPLY if img else False,
            }
        veiculos = sorted(seen_placas.values(), key=lambda v: v["cliente"].lower())
        return veiculos, None

    except Exception as e:
        import traceback; traceback.print_exc()
        return [], str(e)


@app.route("/checklist")
def pagina_checklist():
    veiculos, erro_leitura = _ler_veiculos()

    badge_data = {}
    sb = _supabase()
    if sb:
        try:
            contratos_res = sb.table("checklist_contratos").select("id, contrato").execute()
            if contratos_res.data:
                ids_map = {r["id"]: r["contrato"] for r in contratos_res.data}
                itens_res = sb.table("checklist_itens").select("contrato_id, marcado").execute()
                for item in (itens_res.data or []):
                    cid  = item["contrato_id"]
                    cnum = ids_map.get(cid)
                    if cnum:
                        if cnum not in badge_data:
                            badge_data[cnum] = {"total": 0, "marcados": 0}
                        badge_data[cnum]["total"] += 1
                        if item["marcado"]:
                            badge_data[cnum]["marcados"] += 1
        except Exception:
            pass

    return render_template("checklist.html",
                           active="frota",
                           veiculos=veiculos,
                           badge_data=badge_data,
                           erro_leitura=erro_leitura)


@app.route("/api/checklist/contrato")
def api_checklist_get():
    contrato = request.args.get("contrato", "").strip()
    placa    = request.args.get("placa", "").strip()
    cliente  = request.args.get("cliente", "").strip()
    unidade  = request.args.get("unidade", "").strip()

    if not contrato:
        return jsonify({"error": "Contrato obrigatório"}), 400

    sb = _supabase()
    if not sb:
        return jsonify({"error": "Supabase indisponível"}), 503

    ITENS_PADRAO = ["INDICAÇÃO DE CONDUTOR", "CONTRATO", "NOTA PROMISSÓRIA", "CHAVE RESERVA"]

    res = sb.table("checklist_contratos").select("*").eq("contrato", contrato).execute()

    if res.data:
        rec        = res.data[0]
        contrato_id = rec["id"]
        itens_res  = sb.table("checklist_itens").select("*").eq("contrato_id", contrato_id).order("created_at").execute()
        itens = itens_res.data or []
        if not itens:
            for nome in ITENS_PADRAO:
                sb.table("checklist_itens").insert({"contrato_id": contrato_id, "nome": nome, "marcado": False}).execute()
            itens = sb.table("checklist_itens").select("*").eq("contrato_id", contrato_id).order("created_at").execute().data or []
    else:
        ins = sb.table("checklist_contratos").insert({
            "contrato": contrato, "placa": placa, "cliente": cliente, "unidade": unidade,
        }).execute()
        rec         = ins.data[0]
        contrato_id = rec["id"]
        for nome in ITENS_PADRAO:
            sb.table("checklist_itens").insert({"contrato_id": contrato_id, "nome": nome, "marcado": False}).execute()
        itens = sb.table("checklist_itens").select("*").eq("contrato_id", contrato_id).order("created_at").execute().data or []

    return jsonify({
        "contrato_id": contrato_id,
        "contrato":    contrato,
        "placa":       rec.get("placa", placa),
        "cliente":     rec.get("cliente", cliente),
        "unidade":     rec.get("unidade", unidade),
        "itens": [{"id": i["id"], "nome": i["nome"], "marcado": i["marcado"]} for i in itens],
    })


@app.route("/api/checklist/salvar", methods=["POST"])
def api_checklist_salvar():
    data        = request.get_json(force=True)
    contrato_id = data.get("contrato_id")
    itens       = data.get("itens", [])

    if not contrato_id:
        return jsonify({"error": "contrato_id obrigatório"}), 400

    sb = _supabase()
    if not sb:
        return jsonify({"error": "Supabase indisponível"}), 503

    try:
        for item in itens:
            if item.get("is_new"):
                sb.table("checklist_itens").insert({
                    "contrato_id": contrato_id,
                    "nome":        item["nome"],
                    "marcado":     item.get("marcado", False),
                }).execute()
            else:
                sb.table("checklist_itens").update({
                    "marcado": bool(item.get("marcado", False))
                }).eq("id", item["id"]).execute()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500



@app.route("/api/inad/obs", methods=["PUT"])
def api_inad_obs_salvar():
    data  = request.get_json(force=True)
    chave = (data.get("chave") or "").strip()
    texto = (data.get("texto") or "").strip()[:500]
    if not chave:
        return jsonify({"error": "chave obrigatória"}), 400
    sb = _supabase()
    if not sb:
        return jsonify({"error": "Supabase indisponível"}), 503
    try:
        existing = sb.table("inad_observacoes").select("id").eq("chave", chave).execute()
        if existing.data:
            sb.table("inad_observacoes").update({"texto": texto}).eq("chave", chave).execute()
        else:
            sb.table("inad_observacoes").insert({"chave": chave, "texto": texto}).execute()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/checklist/item/<uuid:item_id>", methods=["PUT"])
def api_checklist_toggle(item_id):
    data    = request.get_json(force=True)
    marcado = bool(data.get("marcado", False))
    sb = _supabase()
    if not sb:
        return jsonify({"error": "Supabase indisponível"}), 503
    try:
        sb.table("checklist_itens").update({"marcado": marcado}).eq("id", str(item_id)).execute()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/checklist/item", methods=["POST"])
def api_checklist_add_item():
    data        = request.get_json(force=True)
    contrato_id = data.get("contrato_id")
    nome        = (data.get("nome") or "").strip().upper()[:80]
    if not contrato_id or not nome:
        return jsonify({"error": "contrato_id e nome obrigatórios"}), 400
    sb = _supabase()
    if not sb:
        return jsonify({"error": "Supabase indisponível"}), 503
    try:
        res  = sb.table("checklist_itens").insert({
            "contrato_id": contrato_id, "nome": nome, "marcado": False
        }).execute()
        item = res.data[0]
        return jsonify({"ok": True, "id": item["id"], "nome": item["nome"], "marcado": item["marcado"]})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/checklist/item/<uuid:item_id>", methods=["DELETE"])
def api_checklist_delete_item(item_id):
    sb = _supabase()
    if not sb:
        return jsonify({"error": "Supabase indisponível"}), 503
    try:
        sb.table("checklist_itens").delete().eq("id", str(item_id)).execute()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Hodômetros semanais ───────────────────────────────────────────────────────
# Motoristas enviam foto do hodômetro toda segunda-feira. O que gravamos é a
# leitura absoluta do painel; o rodado da semana é a diferença entre leituras
# consecutivas. Cada semana é fechada em si (franquia não acumula), e o que
# passar da franquia é cobrado por km.

_HOD_FRANQUIA_PADRAO = 1500.0
_HOD_VALOR_KM_PADRAO = 0.50

# Franquia diferenciada por modelo: os BYD são contratados com 2.000 km/semana.
# A regra é do modelo, não da placa — carro novo do mesmo tipo já entra certo,
# sem cadastro manual. Uma linha em hodometro_config ainda vence isto, para o
# caso de um contrato específico fugir da regra.
_HOD_FRANQUIA_POR_MODELO = [
    ("DOLPHIN", 2000.0),
    ("BYD",     2000.0),
]


# Contratos de leitura mensal (terceirização de frota e carro por assinatura):
# o hodômetro é registrado uma vez por mês, em qualquer dia, com franquia de
# 4.000 km no mês em vez da franquia semanal.
_HOD_FRANQUIA_MENSAL = 4000.0
_HOD_TIPOS_MENSAIS   = ("TERCEIR", "ASSINATURA")


def _hod_e_mensal(tipo_contrato):
    t = (tipo_contrato or "").upper()
    return any(chave in t for chave in _HOD_TIPOS_MENSAIS)


# Cor de fundo da linha por modelo, para reconhecer o veículo de relance na
# grade. A chave vira uma classe CSS (hod-row-<chave>).
_HOD_COR_POR_MODELO = [
    ("DOLPHIN", "byd"),
    ("BYD",     "byd"),
    ("POLO",    "polo"),
]


def _hod_destaque_do_modelo(modelo):
    m = (modelo or "").upper()
    for chave, classe in _HOD_COR_POR_MODELO:
        if chave in m:
            return classe
    return None


def _hod_franquia_do_modelo(modelo):
    m = (modelo or "").upper()
    for chave, franquia in _HOD_FRANQUIA_POR_MODELO:
        if chave in m:
            return franquia
    return _HOD_FRANQUIA_PADRAO


def _hod_franquia_do_veiculo(veiculo, cfg):
    """
    Precedência: cadastro manual da placa > terceirização (mensal) > modelo.
    Devolve (franquia, mensal).
    """
    mensal = _hod_e_mensal(veiculo.get("tipo"))
    if cfg.get("franquia_km"):
        return float(cfg["franquia_km"]), mensal
    if mensal:
        return _HOD_FRANQUIA_MENSAL, True
    return _hod_franquia_do_modelo(veiculo.get("modelo")), False

_MESES_PT = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun",
             "Jul", "Ago", "Set", "Out", "Nov", "Dez"]


def _hod_segundas_do_mes(ano, mes):
    """Todas as segundas-feiras de um mês, em ordem."""
    import calendar as _cal
    ultimo_dia = _cal.monthrange(ano, mes)[1]
    d = date(ano, mes, 1)
    d += timedelta(days=(7 - d.weekday()) % 7)   # anda até a 1ª segunda
    segundas = []
    while d.day <= ultimo_dia and d.month == mes:
        segundas.append(d)
        d += timedelta(days=7)
    return segundas


def _hod_calendario(hoje=None, corte=None):
    """
    Monta os grupos de colunas da planilha.

    `corte` é a primeira segunda-feira que faz sentido exibir — normalmente a
    segunda do cadastro inicial. Segundas anteriores a ela nunca aparecem: são
    semanas em que o controle ainda não existia e ficariam para sempre vazias.

    Regra de visibilidade pedida pela operação: o mês corrente fica aberto; ao
    alcançar a última segunda dele, o mês seguinte já abre junto. Meses
    encerrados ficam recolhidos, exibindo só a última segunda — o front deixa
    expandir sob demanda.
    """
    hoje  = hoje or date.today()
    atual = date(hoje.year, hoje.month, 1)
    corte = corte or (hoje - timedelta(days=hoje.weekday()))

    segundas_atual = _hod_segundas_do_mes(atual.year, atual.month)
    abrir_proximo  = bool(segundas_atual) and hoje >= segundas_atual[-1]

    inicio = date(corte.year, corte.month, 1)
    if inicio > atual:
        inicio = atual

    fim = atual
    if abrir_proximo:
        fim = date(atual.year + (atual.month == 12), (atual.month % 12) + 1, 1)

    grupos, cursor = [], inicio
    while cursor <= fim:
        segundas = [d for d in _hod_segundas_do_mes(cursor.year, cursor.month) if d >= corte]
        if segundas:
            aberto = cursor >= atual
            grupos.append({
                "ym":       cursor.strftime("%Y-%m"),
                "label":    f"{_MESES_PT[cursor.month - 1]}/{cursor.year}",
                "aberto":   aberto,
                "segundas": [{"iso": d.isoformat(), "label": d.strftime("%d/%m")} for d in segundas],
                "resumo":   segundas[-1].isoformat(),   # coluna mostrada quando recolhido
            })
        cursor = date(cursor.year + (cursor.month == 12), (cursor.month % 12) + 1, 1)

    return grupos


def _hod_dinheiro(valor):
    """
    Arredonda em reais com meio-para-cima. O round() nativo usa arredondamento
    bancário (100.275 -> 100.27), o que não corresponde ao que se cobra.
    """
    from decimal import Decimal as _D, ROUND_HALF_UP as _HU
    return float(_D(str(valor)).quantize(_D("0.01"), rounding=_HU))


def _hod_parse_km(valor):
    """
    Aceita o número no formato que o usuário digitar: '123456.78', '123456,78'
    ou '123.456,78'. Havendo vírgula, ela é a decimal e o ponto é milhar.
    """
    s = str(valor).strip()
    if not s:
        raise ValueError("vazio")
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    return float(s)


def _hod_calcular(leituras, franquia, valor_km, mensal=False):
    """
    leituras: {data_iso: km}. Devolve {data_iso: {...}} com o cálculo de cada
    período, comparando cada leitura com a imediatamente anterior registrada.
    Em contratos mensais o período é o mês; nos demais, a semana.

    status:
      base      primeira leitura da placa — marco zero, não gera cobrança
      ok        rodou dentro da franquia
      excedente rodou acima da franquia
      lacuna    pulou um período; sinalizado e não cobrado
      erro      leitura menor que a anterior (troca de painel ou digitação)
    """
    datas = sorted(leituras.keys())
    out, anterior = {}, None

    for iso in datas:
        km = float(leituras[iso])
        info = {"km": km, "rodado": None, "excedente": 0.0, "valor": 0.0, "status": "base"}

        if anterior is not None:
            d_ant  = date.fromisoformat(anterior)
            d_atu  = date.fromisoformat(iso)
            rodado = km - float(leituras[anterior])
            info["rodado"] = rodado

            if mensal:
                # Meses de calendário: a data gravada é a última segunda do mês,
                # então meses consecutivos distam exatamente 1.
                pulou = (d_atu.year * 12 + d_atu.month) - (d_ant.year * 12 + d_ant.month) > 1
            else:
                pulou = (d_atu - d_ant).days / 7.0 > 1.01

            if rodado < 0:
                info["status"] = "erro"
            elif pulou:
                info["status"] = "lacuna"
            else:
                # Arredonda o excedente antes de multiplicar: a subtração de
                # leituras carrega ruído binário (1700.5499999...) que, levado
                # direto ao produto, derruba um centavo da cobrança.
                excedente = round(max(0.0, rodado - franquia), 2)
                info["excedente"] = excedente
                info["valor"]     = _hod_dinheiro(excedente * valor_km)
                info["status"]    = "excedente" if excedente > 0 else "ok"

        out[iso] = info
        anterior = iso

    return out


@app.route("/api/hodometros")
def api_hodometros_grid():
    veiculos, erro = _ler_veiculos()
    if erro:
        return jsonify({"error": erro}), 500

    sb = _supabase()
    if not sb:
        return jsonify({"error": "Supabase indisponível"}), 503

    try:
        regs = sb.table("hodometros").select("placa, data_segunda, km").execute().data or []
        cfgs = sb.table("hodometro_config").select("*").execute().data or []
        # Carro vendido sai da frota e não tem mais hodômetro a controlar. A
        # fonte da verdade é o flag da página de financiamentos, para não haver
        # lista de placas duplicada aqui.
        vend = sb.table("financiamentos_contratos").select("placa, vendido").execute().data or []
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    vendidas = {(r.get("placa") or "").strip().upper() for r in vend if r.get("vendido")}
    vendidas.discard("")

    cfg_map = {c["placa"]: c for c in cfgs}

    por_placa = {}
    for r in regs:
        por_placa.setdefault(r["placa"], {})[str(r["data_segunda"])] = float(r["km"])

    # A grade começa na leitura mais antiga já registrada; enquanto não houver
    # nenhuma, começa na segunda desta semana — assim as semanas anteriores ao
    # início do controle não poluem a tela com colunas eternamente vazias.
    # `voltar` abre semanas anteriores sob demanda: sem isso não haveria onde
    # digitar a leitura de base da primeira semana, já que a coluna só existiria
    # depois que o dado existisse.
    voltar      = max(0, min(request.args.get("voltar", 0, type=int) or 0, 104))
    hoje        = date.today()
    corte       = hoje - timedelta(days=hoje.weekday()) - timedelta(weeks=voltar)
    todas_datas = [d for m in por_placa.values() for d in m]
    if todas_datas:
        corte = min(corte, date.fromisoformat(min(todas_datas)))

    grupos = _hod_calendario(hoje=hoje, corte=corte)

    linhas = []
    for v in veiculos:
        placa = v["placa"]
        if placa.strip().upper() in vendidas:
            continue
        cfg               = cfg_map.get(placa, {})
        franquia, mensal  = _hod_franquia_do_veiculo(v, cfg)
        valor_km          = float(cfg.get("valor_km_extra") or _HOD_VALOR_KM_PADRAO)

        celulas = _hod_calcular(por_placa.get(placa, {}), franquia, valor_km, mensal)
        linhas.append({
            "placa":     placa,
            "modelo":    v["modelo"],
            "cliente":   v["cliente"],     # é também o nome exibido no hover
            "contrato":  v["contrato"],
            "unidade":   v["unidade"],
            "franquia":  franquia,
            "valor_km":  valor_km,
            "mensal":    mensal,
            "destaque":  _hod_destaque_do_modelo(v["modelo"]),
            "celulas":   celulas,
            "total":     _hod_dinheiro(sum(c["valor"] for c in celulas.values())),
        })

    return jsonify({
        "grupos": grupos,
        "linhas": linhas,
        "total_geral": _hod_dinheiro(sum(l["total"] for l in linhas)),
    })


@app.route("/api/hodometros", methods=["POST"])
def api_hodometros_salvar():
    body  = request.get_json(force=True) or {}
    placa = (body.get("placa") or "").strip().upper()
    data  = (body.get("data") or "").strip()
    km_in = body.get("km")

    if not placa or not data:
        return jsonify({"error": "placa e data são obrigatórios"}), 400
    try:
        d = date.fromisoformat(data)
    except ValueError:
        return jsonify({"error": "data inválida (use AAAA-MM-DD)"}), 400
    if d.weekday() != 0:
        return jsonify({"error": "a data precisa ser uma segunda-feira"}), 400

    sb = _supabase()
    if not sb:
        return jsonify({"error": "Supabase indisponível"}), 503

    # Campo esvaziado apaga a leitura — é como o usuário desfaz um lançamento.
    if km_in is None or str(km_in).strip() == "":
        try:
            sb.table("hodometros").delete().eq("placa", placa).eq("data_segunda", data).execute()
            return jsonify({"ok": True, "removido": True})
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    try:
        km = round(_hod_parse_km(km_in), 2)
    except ValueError:
        return jsonify({"error": "km inválido"}), 400
    if km < 0:
        return jsonify({"error": "km não pode ser negativo"}), 400

    try:
        sb.table("hodometros").upsert(
            {"placa": placa, "data_segunda": data, "km": km},
            on_conflict="placa,data_segunda",
        ).execute()
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    return jsonify({"ok": True, "km": km})


@app.route("/api/hodometros/placa/<placa>")
def api_hodometros_placa(placa):
    """Histórico de uma placa só — alimenta o bloco dentro do modal do contrato."""
    placa = placa.strip().upper()
    sb = _supabase()
    if not sb:
        return jsonify({"error": "Supabase indisponível"}), 503

    try:
        regs = sb.table("hodometros").select("data_segunda, km") \
                 .eq("placa", placa).order("data_segunda").execute().data or []
        cfg = sb.table("hodometro_config").select("*").eq("placa", placa).execute().data or []
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    veiculos, _ = _ler_veiculos()
    veiculo = next((v for v in veiculos if v["placa"].strip().upper() == placa), {})

    c                 = cfg[0] if cfg else {}
    franquia, mensal  = _hod_franquia_do_veiculo(veiculo, c)
    valor_km          = float(c.get("valor_km_extra") or _HOD_VALOR_KM_PADRAO)

    leituras = {str(r["data_segunda"]): float(r["km"]) for r in regs}
    celulas  = _hod_calcular(leituras, franquia, valor_km, mensal)

    historico = [dict(data=iso, **celulas[iso]) for iso in sorted(celulas, reverse=True)]
    return jsonify({
        "placa":     placa,
        "franquia":  franquia,
        "valor_km":  valor_km,
        "mensal":    mensal,
        "historico": historico,
        "total":     _hod_dinheiro(sum(h["valor"] for h in historico)),
    })


# ── Análise de Dados: agregações a partir dos lançamentos ────────────────────
# As categorias seguem a vinculação feita no Blue Fleet (_DRE_CATEGORIA_POR_NATUREZA).
# Frota de investidores fica FORA dos indicadores da Ativuz e é reportada à parte.

# Venda de veículo é desmobilização de ativo (seção INVESTIMENTOS na DRE do ERP),
# não receita operacional — fica fora da receita bruta e das margens.
_AD_RECEITA_BRUTA = ("RECEITAS DE LOCAÇÃO", "RECEITAS ADICIONAIS", "REEMBOLSOS")
_AD_DEDUCOES      = ("DEDUÇÕES",)
_AD_CUSTOS        = ("CUSTOS DIRETOS DA FROTA", "CUSTOS OPERACIONAIS")
_AD_DESPESAS      = ("ADMINISTRATIVAS", "COMERCIAIS", "TECNOLOGIA")
_AD_FINANCEIRO    = ("RECEITAS FINANCEIRAS", "DESPESAS FINANCEIRAS",
                     "RESULTADO NÃO OPERACIONAL")
_AD_INVESTIDORES  = ("FROTA INVESTIDORES",)


def _ad_valor(lanc):
    """Valor do lançamento com o sinal econômico correto (não o do fluxo de caixa)."""
    v = lanc["valor"]
    return -v if lanc["codigo"] in _DRE_CODIGOS_SINAL_INVERTIDO else v


def _ad_mes_ref(dt):
    return (dt.year, dt.month)


def _ad_dre_mensal(lancamentos):
    """
    Consolida lançamentos em {(ano, mes): {receita_bruta, deducoes, ...}}.

    Só entram as categorias de resultado: compra de veículos, financiamentos,
    imobilizado, mútuos, cauções e aportes são fluxo de caixa, não resultado.
    """
    meses = {}
    for l in lancamentos:
        cat = _dre_categoria(l["codigo"])
        if not cat:
            continue
        ref = _ad_mes_ref(l["dt"])
        m = meses.setdefault(ref, {
            "receita_bruta": 0.0, "deducoes": 0.0, "custos": 0.0,
            "despesas": 0.0, "financeiro": 0.0, "investidores": 0.0,
        })
        v = _ad_valor(l)
        if   cat in _AD_RECEITA_BRUTA: m["receita_bruta"] += v
        elif cat in _AD_DEDUCOES:      m["deducoes"]      += v
        elif cat in _AD_CUSTOS:        m["custos"]        += v
        elif cat in _AD_DESPESAS:      m["despesas"]      += v
        elif cat in _AD_FINANCEIRO:    m["financeiro"]    += v
        elif cat in _AD_INVESTIDORES:  m["investidores"]  += v

    for m in meses.values():
        m["receita_liquida"] = m["receita_bruta"] + m["deducoes"]
        m["lucro_bruto"]     = m["receita_liquida"] + m["custos"]
        m["ebitda"]          = m["lucro_bruto"] + m["despesas"]
        m["lucro_liquido"]   = m["ebitda"] + m["financeiro"]
    return meses


def _ad_ultimos_meses(meses_dict, n=12, ate=None):
    """As n chaves (ano, mes) mais recentes até 'ate' (default: último mês com dado)."""
    chaves = sorted(meses_dict)
    if ate:
        chaves = [c for c in chaves if c <= ate]
    return chaves[-n:]


def _ad_acumulado(meses_dict, chaves):
    """Soma dos indicadores no conjunto de meses informado."""
    tot = collections.defaultdict(float)
    for c in chaves:
        for k, v in meses_dict.get(c, {}).items():
            tot[k] += v
    return dict(tot)



def _ad_divida():
    """
    Saldo devedor dos financiamentos/consórcios em aberto, com corte curto x longo prazo.

    Curto prazo = parcelas que vencem nos próximos 12 meses; o resto é longo prazo.
    Contratos de veículos já vendidos ficam de fora.
    """
    try:
        sb   = _supabase()
        rows = sb.table("financiamentos_contratos").select("*").execute().data or []
    except Exception:
        return {"total": 0.0, "curto": 0.0, "longo": 0.0, "contratos": 0}

    hoje = datetime.now(_BRT).date()
    total = curto = longo = 0.0
    nao_iniciada = 0.0
    n = n_nao_iniciados = 0
    for r in rows:
        if r.get("vendido"):
            continue
        try:
            parcela  = float(r["valor_parcela"])
            restante = _fin_calcular_restante(r, hoje)
        except Exception:
            continue
        if restante <= 0:
            continue
        n += 1
        saldo = restante * parcela
        total += saldo
        curto += min(restante, 12) * parcela
        longo += max(0, restante - 12) * parcela

        # Contrato ainda não iniciado: nenhuma parcela venceu até hoje. Pesa no
        # saldo devedor sem ter contrapartida no EBITDA já realizado.
        p1 = r.get("data_parcela_1")
        comecou = True
        if p1:
            try:
                comecou = date.fromisoformat(str(p1)[:10]) <= hoje
            except Exception:
                comecou = True
        elif restante >= int(r["parcelas_total"]):
            comecou = False
        if not comecou:
            nao_iniciada += saldo
            n_nao_iniciados += 1

    return {"total": total, "curto": curto, "longo": longo, "contratos": n,
            "nao_iniciada": nao_iniciada, "em_curso": total - nao_iniciada,
            "contratos_nao_iniciados": n_nao_iniciados}


def _ad_frota_valor():
    """Valor FIPE atual da frota vs. valor de aquisição, e série mensal do FIPE total."""
    try:
        sb = _supabase()
        if sb is None:
            return {"fipe": 0.0, "aquisicao": 0.0, "veiculos": 0, "serie": []}
        res_v = sb.table("frota_veiculos").select(
            "placa, vl_aquisicao").eq("ativo", True).execute()
        res_h = sb.table("frota_fipe_historico").select(
            "placa, mes_ref, valor").execute()
    except Exception:
        return {"fipe": 0.0, "aquisicao": 0.0, "veiculos": 0, "serie": []}

    placas = {v["placa"] for v in (res_v.data or [])}
    aquisicao = sum(float(v.get("vl_aquisicao") or 0) for v in (res_v.data or []))

    _M = {"jan": 1, "fev": 2, "mar": 3, "abr": 4, "mai": 5, "jun": 6,
          "jul": 7, "ago": 8, "set": 9, "out": 10, "nov": 11, "dez": 12}

    def _ref_key(ref):
        """'MAI/26' → (2026, 5); None se não reconhecer."""
        try:
            mes, ano = str(ref).lower().replace("/", " ").split()
            return (2000 + int(ano), _M[mes[:3]])
        except Exception:
            return None

    por_mes = {}
    ultimo_por_placa = {}
    for row in (res_h.data or []):
        if row["placa"] not in placas:
            continue
        k = _ref_key(row.get("mes_ref"))
        if not k:
            continue
        valor = float(row["valor"])
        por_mes.setdefault(k, {})[row["placa"]] = valor
        if row["placa"] not in ultimo_por_placa or k > ultimo_por_placa[row["placa"]][0]:
            ultimo_por_placa[row["placa"]] = (k, valor)

    serie = [{"label": f"{k[1]:02d}/{str(k[0])[2:]}",
              "fipe":  round(sum(por_mes[k].values()), 2)}
             for k in sorted(por_mes)]

    fipe_atual = sum(v for _, v in ultimo_por_placa.values())
    return {"fipe": fipe_atual, "aquisicao": aquisicao,
            "veiculos": len(placas), "serie": serie}


def _ad_inadimplencia():
    """Total em atraso mais recente e a série semanal dos snapshots."""
    try:
        sb = _supabase()
        rows = (sb.table("inad_snapshots")
                  .select("semana,total_casos,total_valor")
                  .order("semana").execute().data or [])
    except Exception:
        return {"total": 0.0, "casos": 0, "serie": []}

    serie = []
    for r in rows:
        try:
            d = date.fromisoformat(r["semana"])
        except Exception:
            continue
        serie.append({"label": d.strftime("%d/%m"),
                      "valor": round(float(r["total_valor"] or 0), 2)})
    ult = rows[-1] if rows else {}
    return {"total": float(ult.get("total_valor") or 0),
            "casos": int(ult.get("total_casos") or 0),
            "serie": serie}


def _ad_receita_por_cliente(lancamentos, limite=10):
    """Top clientes por receita de locação no período dos lançamentos recebidos."""
    por = collections.defaultdict(float)
    for l in lancamentos:
        if _dre_categoria(l["codigo"]) not in ("RECEITAS DE LOCAÇÃO", "RECEITAS ADICIONAIS"):
            continue
        nome = l.get("cliente") or "—"
        por[nome] += _ad_valor(l)
    top = sorted(por.items(), key=lambda kv: -kv[1])[:limite]
    total = sum(por.values()) or 1.0
    return [{"nome": n, "valor": v, "valor_s": _brl(v),
             "pct": round(100 * v / total, 1)} for n, v in top]



# ── Análise de Dados ──────────────────────────────────────────────────────────
# Página em construção: estrutura e seções definidas, cálculos a implementar
# depois (receita/margens via _dre_calcular, alavancagem via
# _fin_calcular_restante reconstruído mês a mês, frota via frota_fipe_historico,
# inadimplência via inad_snapshots, receita por cliente via contratos_frota).

@app.route("/benchmarking")
def pagina_benchmarking():
    # Regime de caixa: é como a operação é acompanhada no dia a dia.
    # Só Ativuz + AZ: João Paulo Consórcios e Luz Divina são outras empresas e
    # não entram nos indicadores da casa.
    lancs = _dre_ler_lancamentos("pagamento", grupo=_DRE_GRUPO_PADRAO)
    meses = _ad_dre_mensal(lancs)
    disponiveis = sorted(meses)

    # Filtro: "12m" (padrão) ou um mês específico no formato AAAA-MM.
    periodo_req = (request.args.get("periodo") or "12m").strip()
    ref_sel = None
    if periodo_req != "12m":
        try:
            a, mm = periodo_req.split("-")
            cand = (int(a), int(mm))
            if cand in meses:
                ref_sel = cand
        except (ValueError, TypeError):
            ref_sel = None

    if ref_sel:
        chaves = [ref_sel]
        periodo_label = f"{ref_sel[1]:02d}/{ref_sel[0]}"
        periodo_val   = f"{ref_sel[0]}-{ref_sel[1]:02d}"
    else:
        chaves = _ad_ultimos_meses(meses, 12)
        periodo_val = "12m"
        periodo_label = (f"{chaves[0][1]:02d}/{chaves[0][0]} a "
                         f"{chaves[-1][1]:02d}/{chaves[-1][0]}") if chaves else "—"

    acum   = _ad_acumulado(meses, chaves) if chaves else {}
    rl     = acum.get("receita_liquida", 0.0)
    ebitda = acum.get("ebitda", 0.0)

    # A série do gráfico mostra sempre os últimos 13 meses; o mês filtrado
    # fica destacado para dar contexto em vez de virar uma barra solitária.
    serie = []
    for ref in _ad_ultimos_meses(meses, 13):
        v   = meses[ref]
        rlm = v["receita_liquida"]
        serie.append({
            "label":           f"{ref[1]:02d}/{str(ref[0])[2:]}",
            "receita_liquida": round(rlm, 2),
            "ebitda":          round(v["ebitda"], 2),
            "margem":          round(100 * v["ebitda"] / rlm, 1) if rlm else 0.0,
            "sel":             (ref == ref_sel),
        })

    divida = _ad_divida()
    frota  = _ad_frota_valor()
    inad   = _ad_inadimplencia()

    lancs_periodo = [l for l in lancs if _ad_mes_ref(l["dt"]) in chaves]
    clientes = _ad_receita_por_cliente(lancs_periodo)

    # Dívida/EBITDA anualiza quando o filtro é de um mês só, para o índice
    # continuar comparável com o padrão de mercado (dívida sobre EBITDA anual).
    ebitda_anual = ebitda if len(chaves) >= 12 else ebitda * 12 / max(1, len(chaves))
    div_ebitda   = (divida["total"] / ebitda_anual) if ebitda_anual > 0 else None
    div_ebitda_curso = (divida["em_curso"] / ebitda_anual) if ebitda_anual > 0 else None

    receita_periodo = sum(l["valor"] for l in lancs_periodo
                          if _dre_categoria(l["codigo"]) == "RECEITAS DE LOCAÇÃO")
    taxa_inad = (100 * inad["total"] / receita_periodo) if receita_periodo > 0 else None

    kpis = {
        "receita_liquida": _brl(rl),
        "ebitda":          _brl(ebitda),
        "margem_ebitda":   f"{100 * ebitda / rl:.1f}%".replace(".", ",") if rl else "—",
        "lucro_liquido":   _brl(acum.get("lucro_liquido", 0.0)),
        "investidores":    _brl(acum.get("investidores", 0.0)),
        "saldo_devedor":   _brl(divida["total"]),
        "div_curto":       _brl(divida["curto"]),
        "div_longo":       _brl(divida["longo"]),
        "div_ebitda":      f"{div_ebitda:.2f}x".replace(".", ",") if div_ebitda else "—",
        "div_ebitda_curso": f"{div_ebitda_curso:.2f}x".replace(".", ",") if div_ebitda_curso else "—",
        "div_nao_iniciada": _brl(divida["nao_iniciada"]),
        "div_em_curso":     _brl(divida["em_curso"]),
        "n_nao_iniciados":  divida["contratos_nao_iniciados"],
        "frota_fipe":      _brl(frota["fipe"]),
        "frota_aquisicao": _brl(frota["aquisicao"]),
        "frota_veiculos":  frota["veiculos"],
        "inad_total":      _brl(inad["total"]),
        "inad_casos":      inad["casos"],
        "taxa_inad":       f"{taxa_inad:.1f}%".replace(".", ",") if taxa_inad is not None else "—",
        "periodo":         periodo_label,
    }

    opcoes = [{"valor": f"{a}-{mm:02d}", "label": f"{mm:02d}/{a}"}
              for a, mm in reversed(disponiveis)]

    return render_template("benchmarking.html", active="benchmarking",
                           kpis=kpis, serie=serie, clientes=clientes,
                           serie_frota=frota["serie"], serie_inad=inad["serie"],
                           opcoes=opcoes, periodo_sel=periodo_val)


@app.route("/configuracoes")
def pagina_configuracoes():
    return render_template("configuracoes.html", active="configuracoes")


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5001, debug=False)
